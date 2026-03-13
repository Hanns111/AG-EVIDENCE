# -*- coding: utf-8 -*-
"""
Tests para src/validation/reporte_hallazgos.py — Tarea #29

Cobertura:
  - EscritorHallazgos: creación de hoja HALLAZGOS
  - _observaciones_a_filas: conversión y ordenamiento por severidad
  - _calcular_resumen: contadores por severidad
  - _clasificar_tipo: mapeo regla_aplicada → tipo hallazgo
  - FilaHallazgo.to_list: serialización de fila
  - ResumenHallazgos.to_dict: serialización de resumen
  - escribir_hallazgos: función de conveniencia
  - Integración con openpyxl: hoja creada, banner, tabla, pie
"""

from dataclasses import field
from typing import List

import pytest

from config.settings import (
    EvidenciaProbatoria,
    NivelObservacion,
    Observacion,
)
from src.validation.reporte_hallazgos import (
    MAPA_SEVERIDAD,
    NOMBRE_HOJA,
    TIPOS_HALLAZGO,
    VERSION_REPORTE,
    EscritorHallazgos,
    FilaHallazgo,
    ResumenHallazgos,
    _clasificar_tipo,
    escribir_hallazgos,
)

# Intentar importar openpyxl
try:
    from openpyxl import Workbook

    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ==============================================================================
# HELPERS
# ==============================================================================


def _obs(
    nivel: NivelObservacion = NivelObservacion.MENOR,
    descripcion: str = "Test hallazgo",
    regla: str = "VAL_TOTAL_001",
    archivo: str = "factura.pdf",
    pagina: int = 5,
    valor_detectado: str = "F001-00123",
    confianza: float = 0.85,
) -> Observacion:
    """Crea una Observacion con evidencia para tests."""
    ev = EvidenciaProbatoria(
        archivo=archivo,
        pagina=pagina,
        valor_detectado=valor_detectado,
        confianza=confianza,
    )
    return Observacion(
        nivel=nivel,
        agente="AG04_VALIDADOR",
        descripcion=descripcion,
        accion_requerida="Verificar manualmente",
        regla_aplicada=regla,
        evidencias=[ev],
    )


def _obs_sin_evidencia(
    nivel: NivelObservacion = NivelObservacion.INFORMATIVA,
    descripcion: str = "Info sin evidencia",
    regla: str = "",
) -> Observacion:
    """Crea Observacion sin evidencia."""
    return Observacion(
        nivel=nivel,
        agente="AG04_VALIDADOR",
        descripcion=descripcion,
        accion_requerida="",
        regla_aplicada=regla,
    )


def _lista_mixta() -> List[Observacion]:
    """Lista con observaciones de distintas severidades."""
    return [
        _obs(NivelObservacion.INFORMATIVA, "Info test", "VIAT_COBERTURA_001"),
        _obs(NivelObservacion.CRITICA, "Duplicidad detectada", "VAL_DUPLICIDAD_001"),
        _obs(NivelObservacion.MENOR, "IGV menor", "VAL_IGV_002"),
        _obs(NivelObservacion.MAYOR, "Tope excedido", "VIAT_TOPE_001"),
        _obs(NivelObservacion.INCIERTO, "Dato incierto", "VAL_CAMPOS_001"),
    ]


# ==============================================================================
# TESTS — _clasificar_tipo
# ==============================================================================


class TestClasificarTipo:
    def test_aritmetico_igv(self):
        assert _clasificar_tipo("VAL_IGV_001") == "ARITMÉTICO"

    def test_aritmetico_total(self):
        assert _clasificar_tipo("VAL_TOTAL_002") == "ARITMÉTICO"

    def test_aritmetico_suma_items(self):
        assert _clasificar_tipo("VAL_SUMA_ITEMS_001") == "ARITMÉTICO"

    def test_aritmetico_noches(self):
        assert _clasificar_tipo("VAL_NOCHES_001") == "ARITMÉTICO"

    def test_cruzado_duplicidad(self):
        assert _clasificar_tipo("VAL_DUPLICIDAD_001") == "CRUZADO"

    def test_cruzado_suma_anexo(self):
        assert _clasificar_tipo("VAL_SUMA_VS_ANEXO3_001") == "CRUZADO"

    def test_documental_campos(self):
        assert _clasificar_tipo("VAL_CAMPOS_001") == "DOCUMENTAL"

    def test_normativo_tope(self):
        assert _clasificar_tipo("VIAT_TOPE_001") == "NORMATIVO"

    def test_normativo_fecha(self):
        assert _clasificar_tipo("VIAT_FECHA_001") == "NORMATIVO"

    def test_normativo_monto(self):
        assert _clasificar_tipo("VIAT_MONTO_001") == "NORMATIVO"

    def test_normativo_cobertura(self):
        assert _clasificar_tipo("VIAT_COBERTURA_001") == "NORMATIVO"

    def test_normativo_boleta(self):
        assert _clasificar_tipo("VIAT_BOLETA_001") == "NORMATIVO"

    def test_documental_doc(self):
        assert _clasificar_tipo("VIAT_DOC_001") == "DOCUMENTAL"

    def test_otro_desconocido(self):
        assert _clasificar_tipo("DESCONOCIDO_001") == "OTRO"

    def test_vacio(self):
        assert _clasificar_tipo("") == "OTRO"

    def test_none_like(self):
        assert _clasificar_tipo("") == "OTRO"


# ==============================================================================
# TESTS — FilaHallazgo
# ==============================================================================


class TestFilaHallazgo:
    def test_to_list_completa(self):
        fila = FilaHallazgo(
            numero=1,
            severidad="CRÍTICO",
            tipo="ARITMÉTICO",
            descripcion="IGV incorrecto",
            accion_requerida="Recalcular",
            referencia_normativa="VAL_IGV_001",
            archivo="factura.pdf",
            pagina=5,
            comprobante="F001-00123",
            confianza=0.85,
        )
        lista = fila.to_list()
        assert lista[0] == 1
        assert lista[1] == "CRÍTICO"
        assert lista[2] == "ARITMÉTICO"
        assert lista[3] == "IGV incorrecto"
        assert lista[4] == "Recalcular"
        assert lista[5] == "VAL_IGV_001"
        assert lista[6] == "factura.pdf"
        assert lista[7] == 5
        assert lista[8] == "F001-00123"
        assert lista[9] == "85%"

    def test_pagina_cero_muestra_vacio(self):
        fila = FilaHallazgo(
            numero=1,
            severidad="BAJO",
            tipo="OTRO",
            descripcion="x",
            accion_requerida="",
            referencia_normativa="",
            archivo="",
            pagina=0,
            comprobante="",
            confianza=0.0,
        )
        lista = fila.to_list()
        assert lista[7] == ""  # página 0 → vacío
        assert lista[9] == ""  # confianza 0 → vacío


# ==============================================================================
# TESTS — ResumenHallazgos
# ==============================================================================


class TestResumenHallazgos:
    def test_to_dict(self):
        r = ResumenHallazgos(total=10, criticos=2, altos=3, medios=1, bajos=4, info=0)
        d = r.to_dict()
        assert d["total"] == 10
        assert d["criticos"] == 2
        assert d["altos"] == 3
        assert d["medios"] == 1
        assert d["bajos"] == 4
        assert d["info"] == 0

    def test_defaults(self):
        r = ResumenHallazgos()
        assert r.total == 0
        assert r.criticos == 0


# ==============================================================================
# TESTS — EscritorHallazgos (lógica sin openpyxl)
# ==============================================================================


class TestObservacionesAFilas:
    def setup_method(self):
        self.escritor = EscritorHallazgos()

    def test_orden_por_severidad(self):
        obs = _lista_mixta()
        filas = self.escritor._observaciones_a_filas(obs)
        assert len(filas) == 5
        # CRITICA primero, luego MAYOR, MENOR, INFORMATIVA, INCIERTO
        assert filas[0].severidad == "CRÍTICO"
        assert filas[1].severidad == "ALTO"
        assert filas[2].severidad == "MEDIO"
        assert filas[3].severidad == "BAJO"
        assert filas[4].severidad == "INFO"

    def test_numeracion_correlativa(self):
        obs = _lista_mixta()
        filas = self.escritor._observaciones_a_filas(obs)
        numeros = [f.numero for f in filas]
        assert numeros == [1, 2, 3, 4, 5]

    def test_tipo_mapeado(self):
        obs = [_obs(regla="VAL_IGV_001")]
        filas = self.escritor._observaciones_a_filas(obs)
        assert filas[0].tipo == "ARITMÉTICO"

    def test_descripcion_truncada(self):
        obs = [_obs(descripcion="A" * 300)]
        filas = self.escritor._observaciones_a_filas(obs)
        assert len(filas[0].descripcion) == 200

    def test_evidencia_extraida(self):
        obs = [_obs(archivo="test.pdf", pagina=10, valor_detectado="F001-999", confianza=0.9)]
        filas = self.escritor._observaciones_a_filas(obs)
        assert filas[0].archivo == "test.pdf"
        assert filas[0].pagina == 10
        assert filas[0].comprobante == "F001-999"
        assert filas[0].confianza == 0.9

    def test_sin_evidencia(self):
        obs = [_obs_sin_evidencia()]
        filas = self.escritor._observaciones_a_filas(obs)
        assert filas[0].archivo == ""
        assert filas[0].pagina == 0
        assert filas[0].comprobante == ""

    def test_lista_vacia(self):
        filas = self.escritor._observaciones_a_filas([])
        assert filas == []

    def test_comprobante_no_detectado_sin_patron(self):
        """Si valor_detectado no tiene patrón de serie, comprobante queda vacío."""
        obs = [_obs(valor_detectado="algun texto")]
        filas = self.escritor._observaciones_a_filas(obs)
        assert filas[0].comprobante == ""


class TestCalcularResumen:
    def setup_method(self):
        self.escritor = EscritorHallazgos()

    def test_contadores_mixtos(self):
        obs = _lista_mixta()
        resumen = self.escritor._calcular_resumen(obs)
        assert resumen.total == 5
        assert resumen.criticos == 1
        assert resumen.altos == 1
        assert resumen.medios == 1
        assert resumen.bajos == 1
        assert resumen.info == 1

    def test_vacio(self):
        resumen = self.escritor._calcular_resumen([])
        assert resumen.total == 0
        assert resumen.criticos == 0

    def test_solo_criticos(self):
        obs = [_obs(NivelObservacion.CRITICA)] * 3
        resumen = self.escritor._calcular_resumen(obs)
        assert resumen.criticos == 3
        assert resumen.altos == 0


# ==============================================================================
# TESTS — EscritorHallazgos (con openpyxl)
# ==============================================================================


@pytest.mark.skipif(not OPENPYXL_OK, reason="openpyxl no disponible")
class TestEscritorHallazgosExcel:
    def setup_method(self):
        self.escritor = EscritorHallazgos()
        self.wb = Workbook()

    def test_hoja_creada(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs, "TEST-001")
        assert ws is not None
        assert NOMBRE_HOJA in self.wb.sheetnames

    def test_banner_titulo(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs, "SINAD-001")
        assert "SINAD-001" in str(ws.cell(row=1, column=1).value)

    def test_banner_sin_sinad(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs, "")
        assert "REPORTE DE HALLAZGOS" in str(ws.cell(row=1, column=1).value)

    def test_banner_resumen(self):
        obs = _lista_mixta()
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        resumen_texto = str(ws.cell(row=2, column=1).value)
        assert "Total: 5" in resumen_texto
        assert "Críticos: 1" in resumen_texto

    def test_headers_tabla(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        # Row 4 should be headers (row 1-2 banner, row 3 blank)
        header_row = 4
        assert ws.cell(row=header_row, column=1).value == "#"
        assert ws.cell(row=header_row, column=2).value == "Severidad"
        assert ws.cell(row=header_row, column=4).value == "Descripción"
        assert ws.cell(row=header_row, column=10).value == "Confianza"

    def test_datos_fila(self):
        obs = [_obs(NivelObservacion.CRITICA, "IGV incorrecto", "VAL_IGV_001")]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        data_row = 5  # after banner (2) + blank (1) + header (1)
        assert ws.cell(row=data_row, column=1).value == 1
        assert ws.cell(row=data_row, column=2).value == "CRÍTICO"
        assert ws.cell(row=data_row, column=3).value == "ARITMÉTICO"

    def test_multiples_filas_ordenadas(self):
        obs = _lista_mixta()
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        # Primera fila de datos = CRÍTICO (más grave)
        assert ws.cell(row=5, column=2).value == "CRÍTICO"
        # Última = INFO
        assert ws.cell(row=9, column=2).value == "INFO"

    def test_pie_presente(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        # Pie: banner(2) + blank(1) + header(1) + data(1) + blank(1) + pie(1) = row 7
        pie_row = 7
        valor = str(ws.cell(row=pie_row, column=1).value or "")
        assert "AG-EVIDENCE" in valor or VERSION_REPORTE in valor

    def test_reemplaza_hoja_existente(self):
        """Si la hoja ya existe, la reemplaza."""
        self.wb.create_sheet(title=NOMBRE_HOJA)
        assert NOMBRE_HOJA in self.wb.sheetnames
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        assert ws is not None
        # Solo una hoja HALLAZGOS
        count = sum(1 for s in self.wb.sheetnames if s == NOMBRE_HOJA)
        assert count == 1

    def test_lista_vacia_genera_hoja(self):
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, [], "SINAD-EMPTY")
        assert ws is not None
        assert NOMBRE_HOJA in self.wb.sheetnames
        # Banner present, no data rows
        assert "REPORTE DE HALLAZGOS" in str(ws.cell(row=1, column=1).value)

    def test_nombre_hoja_custom(self):
        escritor = EscritorHallazgos(nombre_hoja="CUSTOM")
        ws = escritor.escribir_hoja_hallazgos(self.wb, [_obs()])
        assert "CUSTOM" in self.wb.sheetnames

    def test_anchos_columnas(self):
        obs = [_obs()]
        ws = self.escritor.escribir_hoja_hallazgos(self.wb, obs)
        # Descripción (col D) should be widest
        from openpyxl.utils import get_column_letter

        assert ws.column_dimensions["D"].width == 60


# ==============================================================================
# TESTS — Función de conveniencia
# ==============================================================================


@pytest.mark.skipif(not OPENPYXL_OK, reason="openpyxl no disponible")
class TestEscribirHallazgos:
    def test_convenencia_basica(self):
        wb = Workbook()
        obs = [_obs()]
        ws = escribir_hallazgos(wb, obs, "TEST-CONV")
        assert ws is not None
        assert NOMBRE_HOJA in wb.sheetnames

    def test_convenencia_nombre_custom(self):
        wb = Workbook()
        ws = escribir_hallazgos(wb, [_obs()], nombre_hoja="HALLAZGOS_V2")
        assert "HALLAZGOS_V2" in wb.sheetnames


# ==============================================================================
# TESTS — Constantes y version
# ==============================================================================


class TestConstantes:
    def test_version(self):
        assert VERSION_REPORTE == "1.0.0"

    def test_nombre_hoja(self):
        assert NOMBRE_HOJA == "HALLAZGOS"

    def test_mapa_severidad_completo(self):
        assert len(MAPA_SEVERIDAD) == 5
        assert MAPA_SEVERIDAD[NivelObservacion.CRITICA] == "CRÍTICO"
        assert MAPA_SEVERIDAD[NivelObservacion.MAYOR] == "ALTO"
        assert MAPA_SEVERIDAD[NivelObservacion.MENOR] == "MEDIO"
        assert MAPA_SEVERIDAD[NivelObservacion.INFORMATIVA] == "BAJO"
        assert MAPA_SEVERIDAD[NivelObservacion.INCIERTO] == "INFO"

    def test_escritor_version(self):
        e = EscritorHallazgos()
        assert e.version == "1.0.0"
