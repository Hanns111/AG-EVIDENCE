# -*- coding: utf-8 -*-
"""
Tests para src/extraction/excel_writer.py (Tarea #20)
=====================================================

Verifica:
  - Creación de hoja DIAGNOSTICO en Workbook
  - Banner con SINAD, status, confianza, acción
  - Tabla de secciones con colores semáforo
  - Tabla de detalle por campo con confianza y color
  - Reemplazo de hoja existente
  - Ajuste de anchos de columna
  - Métricas globales en pie de página
  - Función de conveniencia
  - Manejo de edge cases (vacíos, sin campos, etc.)

Requiere: openpyxl
"""

import os
import sys
from datetime import datetime, timezone

import pytest

# Asegurar que el directorio raíz del proyecto esté en el path
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

try:
    from openpyxl import Workbook

    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

from config.settings import MetodoExtraccion
from src.extraction.abstencion import (
    CampoExtraido,
    EvidenceStatus,
    RazonAbstencion,
    ResultadoAbstencion,
)
from src.extraction.confidence_router import (
    VERSION_ROUTER,
    DecisionCheckpoint,
    DiagnosticoExpediente,
    ResultadoRouter,
    SeccionDiagnostico,
)
from src.extraction.excel_writer import (
    AMARILLO,
    GRIS,
    NOMBRE_HOJA,
    ROJO,
    VERDE,
    VERSION_EXCEL_WRITER,
    EscritorDiagnostico,
    _color_por_confianza,
    _color_por_status,
    escribir_diagnostico,
)
from src.extraction.expediente_contract import (
    ConfianzaGlobal,
    IntegridadStatus,
)

# ==============================================================================
# FIXTURES
# ==============================================================================


@pytest.fixture
def campo_legible():
    """CampoExtraido con confianza alta (LEGIBLE)."""
    return CampoExtraido(
        nombre_campo="ruc_proveedor",
        valor="20123456789",
        archivo="expediente.pdf",
        pagina=3,
        confianza=0.92,
        metodo=MetodoExtraccion.OCR,
        snippet="RUC: 20123456789",
        tipo_campo="ruc",
        motor_ocr="PaddleOCR",
    )


@pytest.fixture
def campo_incompleto():
    """CampoExtraido con confianza media (INCOMPLETO)."""
    return CampoExtraido(
        nombre_campo="fecha_emision",
        valor="06/02/2026",
        archivo="expediente.pdf",
        pagina=5,
        confianza=0.55,
        metodo=MetodoExtraccion.OCR,
        snippet="Fecha: 06/02/2026",
        tipo_campo="fecha",
        motor_ocr="Tesseract",
    )


@pytest.fixture
def campo_abstencion():
    """CampoExtraido en abstención (ILEGIBLE)."""
    return CampoExtraido(
        nombre_campo="monto_total",
        valor=None,
        archivo="expediente.pdf",
        pagina=0,
        confianza=0.0,
        metodo=MetodoExtraccion.OCR,
        snippet="",
        tipo_campo="monto",
        regla_aplicada="ABSTENCION",
    )


@pytest.fixture
def resultado_router_basico(campo_legible, campo_incompleto, campo_abstencion):
    """ResultadoRouter con 3 campos (verde, amarillo, rojo)."""
    resultados_abs = [
        ResultadoAbstencion(
            campo=campo_legible,
            debe_abstenerse=False,
            umbral_aplicado=0.60,
        ),
        ResultadoAbstencion(
            campo=campo_incompleto,
            debe_abstenerse=False,
            umbral_aplicado=0.60,
        ),
        ResultadoAbstencion(
            campo=campo_abstencion,
            debe_abstenerse=True,
            razon_abstencion="Confianza 0.0 < umbral 0.60",
            razon_codigo=RazonAbstencion.CONFIANZA_BAJA,
            umbral_aplicado=0.60,
        ),
    ]
    return ResultadoRouter(
        status=IntegridadStatus.WARNING,
        confianza_global=ConfianzaGlobal.MEDIA,
        debe_detener=False,
        campos_evaluados=3,
        campos_abstenidos=1,
        campos_incompletos=1,
        campos_legibles=1,
        tasa_abstencion=1 / 3,
        resultados_abstencion=resultados_abs,
        alertas=["Tasa de abstención 33% >= 30%"],
        timestamp=datetime.now(timezone.utc).isoformat(),
    )


@pytest.fixture
def diagnostico_basico():
    """DiagnosticoExpediente con 6 secciones estándar."""
    return DiagnosticoExpediente(
        sinad="TEST2026-INT-0000001",
        timestamp=datetime.now(timezone.utc).isoformat(),
        version_router=VERSION_ROUTER,
        secciones=[
            SeccionDiagnostico(
                nombre="evaluacion_campos",
                status="WARNING",
                mensaje="2/3 campos legibles (67%)",
                detalles=["Legibles: 1", "Incompletos: 1", "Abstenidos: 1"],
                metricas={"total": 3, "legibles": 1},
            ),
            SeccionDiagnostico(
                nombre="enforcement",
                status="OK",
                mensaje="0 observaciones degradadas",
                detalles=[],
            ),
            SeccionDiagnostico(
                nombre="completitud",
                status="OK",
                mensaje="Sin problemas de completitud",
                detalles=[],
            ),
            SeccionDiagnostico(
                nombre="unicidad",
                status="OK",
                mensaje="Sin duplicados",
                detalles=[],
            ),
            SeccionDiagnostico(
                nombre="aritmetica",
                status="OK",
                mensaje="Sin errores aritméticos",
                detalles=[],
            ),
            SeccionDiagnostico(
                nombre="decision",
                status="WARNING",
                mensaje="CONTINUAR_CON_ALERTAS",
                detalles=["Tasa de abstención 33% >= 30%"],
            ),
        ],
    )


@pytest.fixture
def decision_completa(resultado_router_basico, diagnostico_basico):
    """DecisionCheckpoint completo con todas las partes."""
    return DecisionCheckpoint(
        accion="CONTINUAR_CON_ALERTAS",
        resultado=resultado_router_basico,
        diagnostico=diagnostico_basico,
        timestamp=datetime.now(timezone.utc).isoformat(),
    )


@pytest.fixture
def decision_critical():
    """DecisionCheckpoint con status CRITICAL."""
    diag = DiagnosticoExpediente(
        sinad="CRITICAL2026-INT-0000002",
        timestamp=datetime.now(timezone.utc).isoformat(),
        version_router=VERSION_ROUTER,
        secciones=[
            SeccionDiagnostico(
                nombre="evaluacion_campos",
                status="CRITICAL",
                mensaje="0/5 campos legibles (0%)",
                detalles=["Legibles: 0", "Abstenidos: 5"],
            ),
            SeccionDiagnostico(
                nombre="decision",
                status="CRITICAL",
                mensaje="DETENER",
                detalles=["Tasa de abstención 100% >= 50%"],
            ),
        ],
    )
    resultado = ResultadoRouter(
        status=IntegridadStatus.CRITICAL,
        confianza_global=ConfianzaGlobal.BAJA,
        debe_detener=True,
        razon_detencion="Tasa de abstención 100% >= 50%",
        campos_evaluados=5,
        campos_abstenidos=5,
        campos_legibles=0,
        tasa_abstencion=1.0,
        alertas=["Tasa de abstención 100% >= 50%"],
        timestamp=datetime.now(timezone.utc).isoformat(),
    )
    return DecisionCheckpoint(
        accion="DETENER",
        resultado=resultado,
        diagnostico=diag,
        timestamp=datetime.now(timezone.utc).isoformat(),
    )


@pytest.fixture
def decision_ok():
    """DecisionCheckpoint con status OK."""
    campo_ok = CampoExtraido(
        nombre_campo="ruc_proveedor",
        valor="20123456789",
        archivo="test.pdf",
        pagina=1,
        confianza=0.95,
        metodo=MetodoExtraccion.OCR,
        tipo_campo="ruc",
    )
    diag = DiagnosticoExpediente(
        sinad="OK2026-INT-0000003",
        timestamp=datetime.now(timezone.utc).isoformat(),
        version_router=VERSION_ROUTER,
        secciones=[
            SeccionDiagnostico(
                nombre="evaluacion_campos",
                status="OK",
                mensaje="5/5 campos legibles (100%)",
            ),
            SeccionDiagnostico(
                nombre="decision",
                status="OK",
                mensaje="CONTINUAR",
            ),
        ],
    )
    resultado = ResultadoRouter(
        status=IntegridadStatus.OK,
        confianza_global=ConfianzaGlobal.ALTA,
        campos_evaluados=5,
        campos_legibles=5,
        tasa_abstencion=0.0,
        resultados_abstencion=[
            ResultadoAbstencion(campo=campo_ok, debe_abstenerse=False),
        ],
        timestamp=datetime.now(timezone.utc).isoformat(),
    )
    return DecisionCheckpoint(
        accion="CONTINUAR",
        resultado=resultado,
        diagnostico=diag,
        timestamp=datetime.now(timezone.utc).isoformat(),
    )


# ==============================================================================
# TESTS: COLORES Y UTILIDADES
# ==============================================================================


class TestColores:
    """Tests para funciones de mapeo de colores."""

    def test_color_por_status_ok(self):
        assert _color_por_status("OK") == VERDE

    def test_color_por_status_warning(self):
        assert _color_por_status("WARNING") == AMARILLO

    def test_color_por_status_critical(self):
        assert _color_por_status("CRITICAL") == ROJO

    def test_color_por_status_skip(self):
        assert _color_por_status("SKIP") == GRIS

    def test_color_por_status_legible(self):
        assert _color_por_status("LEGIBLE") == VERDE

    def test_color_por_status_incompleto(self):
        assert _color_por_status("INCOMPLETO") == AMARILLO

    def test_color_por_status_ilegible(self):
        assert _color_por_status("ILEGIBLE") == ROJO

    def test_color_por_status_abstencion(self):
        assert _color_por_status("ABSTENCION") == ROJO

    def test_color_por_status_desconocido_default_gris(self):
        assert _color_por_status("OTRO") == GRIS

    def test_color_por_status_case_insensitive(self):
        assert _color_por_status("ok") == VERDE
        assert _color_por_status("Ok") == VERDE

    def test_color_por_confianza_alta(self):
        assert _color_por_confianza(0.95) == VERDE
        assert _color_por_confianza(0.70) == VERDE

    def test_color_por_confianza_media(self):
        assert _color_por_confianza(0.55) == AMARILLO
        assert _color_por_confianza(0.40) == AMARILLO

    def test_color_por_confianza_baja(self):
        assert _color_por_confianza(0.39) == ROJO
        assert _color_por_confianza(0.0) == ROJO

    def test_color_semaforo_frozen(self):
        """ColorSemaforo es inmutable."""
        with pytest.raises(AttributeError):
            VERDE.fondo = "000000"  # type: ignore


class TestColorSemaforo:
    """Tests para la dataclass ColorSemaforo."""

    def test_verde_valores(self):
        assert VERDE.fondo == "C6EFCE"
        assert VERDE.texto == "006100"
        assert VERDE.nombre == "verde"

    def test_rojo_valores(self):
        assert ROJO.fondo == "FFC7CE"
        assert ROJO.texto == "9C0006"
        assert ROJO.nombre == "rojo"

    def test_amarillo_valores(self):
        assert AMARILLO.fondo == "FFEB9C"
        assert AMARILLO.texto == "9C5700"
        assert AMARILLO.nombre == "amarillo"


# ==============================================================================
# TESTS: ESCRITOR DIAGNOSTICO (requiere openpyxl)
# ==============================================================================


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestEscritorDiagnostico:
    """Tests para la clase EscritorDiagnostico."""

    def test_crear_escritor_default(self):
        escritor = EscritorDiagnostico()
        assert escritor.nombre_hoja == NOMBRE_HOJA

    def test_crear_escritor_nombre_custom(self):
        escritor = EscritorDiagnostico(nombre_hoja="DIAG_CUSTOM")
        assert escritor.nombre_hoja == "DIAG_CUSTOM"

    def test_escribir_hoja_crea_sheet(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)
        assert NOMBRE_HOJA in wb.sheetnames
        assert ws.title == NOMBRE_HOJA

    def test_escribir_hoja_reemplaza_existente(self, decision_completa):
        wb = Workbook()
        wb.create_sheet(title=NOMBRE_HOJA)
        assert NOMBRE_HOJA in wb.sheetnames

        escritor = EscritorDiagnostico()
        escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Solo debe haber una hoja DIAGNOSTICO
        count = sum(1 for s in wb.sheetnames if s == NOMBRE_HOJA)
        assert count == 1

    def test_error_sin_diagnostico(self):
        wb = Workbook()
        decision = DecisionCheckpoint(diagnostico=None)
        escritor = EscritorDiagnostico()

        with pytest.raises(ValueError, match="no tiene diagnóstico"):
            escritor.escribir_hoja_diagnostico(wb, decision)

    def test_banner_contiene_sinad(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Buscar SINAD en las primeras filas
        found = False
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=6, values_only=True):
            for cell in row:
                if cell and "TEST2026-INT-0000001" in str(cell):
                    found = True
                    break
        assert found, "SINAD no encontrado en el banner"

    def test_banner_contiene_titulo(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)
        assert ws.cell(row=1, column=1).value == "DIAGNÓSTICO DE EXPEDIENTE"

    def test_banner_status_warning_tiene_color(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Fila 3 (datos), columna 2 (Status) debe tener fill amarillo
        celda_status = ws.cell(row=3, column=2)
        assert celda_status.value == "WARNING"
        assert celda_status.fill.start_color.rgb == "00" + AMARILLO.fondo

    def test_banner_accion_continuar_alertas_tiene_color(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Fila 3, columna 4 (Acción)
        celda_accion = ws.cell(row=3, column=4)
        assert celda_accion.value == "CONTINUAR_CON_ALERTAS"
        assert celda_accion.fill.start_color.rgb == "00" + AMARILLO.fondo


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestSeccionesDiagnostico:
    """Tests para la escritura de secciones del diagnóstico."""

    def test_secciones_presentes(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Buscar nombres de secciones legibles en la hoja
        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any("Evaluación de Campos" in c for c in contenido)
        assert any("Estándar Probatorio" in c for c in contenido)
        assert any("Completitud" in c for c in contenido)
        assert any("Unicidad" in c for c in contenido)
        assert any("Aritmética" in c for c in contenido)
        assert any("Decisión Final" in c for c in contenido)

    def test_seccion_warning_tiene_color_amarillo(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Buscar celda con "WARNING" en columna Status de secciones
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "WARNING":
                    assert cell.fill.start_color.rgb == "00" + AMARILLO.fondo
                    return
        pytest.fail("No se encontró celda WARNING en secciones")

    def test_seccion_ok_tiene_color_verde(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Buscar celda con "OK" en columna Status de secciones
        found = False
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "OK":
                    assert cell.fill.start_color.rgb == "00" + VERDE.fondo
                    found = True
                    break
            if found:
                break
        assert found, "No se encontró celda OK en secciones"


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestDetalleCampos:
    """Tests para la tabla de detalle por campo."""

    def test_campos_presentes(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any("ruc_proveedor" in c for c in contenido)
        assert any("fecha_emision" in c for c in contenido)
        assert any("monto_total" in c for c in contenido)

    def test_campo_legible_confianza_verde(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # Buscar "92%" (confianza del campo legible) en columna 3
        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "ruc_proveedor":
                celda_conf = row[2]  # Columna 3: Confianza
                assert celda_conf.value == "92%"
                assert celda_conf.fill.start_color.rgb == "00" + VERDE.fondo
                return
        pytest.fail("Campo ruc_proveedor no encontrado")

    def test_campo_incompleto_confianza_amarilla(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "fecha_emision":
                celda_conf = row[2]
                assert celda_conf.value == "55%"
                assert celda_conf.fill.start_color.rgb == "00" + AMARILLO.fondo
                return
        pytest.fail("Campo fecha_emision no encontrado")

    def test_campo_abstencion_confianza_roja(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "monto_total":
                celda_conf = row[2]
                assert celda_conf.value == "0%"
                assert celda_conf.fill.start_color.rgb == "00" + ROJO.fondo
                return
        pytest.fail("Campo monto_total no encontrado")

    def test_campo_abstencion_valor_muestra_parentesis(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "monto_total":
                assert row[1].value == "(abstención)"
                return
        pytest.fail("Campo monto_total no encontrado")

    def test_campo_status_legible(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "ruc_proveedor":
                celda_status = row[3]  # Columna 4: Status
                assert celda_status.value == "LEGIBLE"
                assert celda_status.fill.start_color.rgb == "00" + VERDE.fondo
                return
        pytest.fail("Campo ruc_proveedor no encontrado")

    def test_campo_muestra_motor(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "ruc_proveedor":
                assert row[4].value == "PaddleOCR"
                return
        pytest.fail("Campo ruc_proveedor no encontrado")

    def test_campo_muestra_archivo_pagina(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        for row in ws.iter_rows(min_row=1, max_row=40):
            if row[0].value == "ruc_proveedor":
                assert "expediente.pdf" in str(row[5].value)
                assert ":3" in str(row[5].value)
                return
        pytest.fail("Campo ruc_proveedor no encontrado")


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestCasosCriticos:
    """Tests para status CRITICAL y DETENER."""

    def test_critical_status_rojo(self, decision_critical):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_critical)

        # Buscar celda Status con "CRITICAL"
        celda_status = ws.cell(row=3, column=2)
        assert celda_status.value == "CRITICAL"
        assert celda_status.fill.start_color.rgb == "00" + ROJO.fondo

    def test_detener_accion_roja(self, decision_critical):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_critical)

        celda_accion = ws.cell(row=3, column=4)
        assert celda_accion.value == "DETENER"
        assert celda_accion.fill.start_color.rgb == "00" + ROJO.fondo

    def test_alertas_presentes(self, decision_critical):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_critical)

        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any("ALERTAS" in c for c in contenido)


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestCasosOK:
    """Tests para status OK y CONTINUAR."""

    def test_ok_status_verde(self, decision_ok):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_ok)

        celda_status = ws.cell(row=3, column=2)
        assert celda_status.value == "OK"
        assert celda_status.fill.start_color.rgb == "00" + VERDE.fondo

    def test_continuar_accion_verde(self, decision_ok):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_ok)

        celda_accion = ws.cell(row=3, column=4)
        assert celda_accion.value == "CONTINUAR"
        assert celda_accion.fill.start_color.rgb == "00" + VERDE.fondo


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestEscribirDesdeComponentes:
    """Tests para el método alternativo escribir_desde_componentes."""

    def test_desde_componentes_crea_hoja(self, diagnostico_basico, resultado_router_basico):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_desde_componentes(
            wb, diagnostico_basico, resultado_router_basico, "CONTINUAR_CON_ALERTAS"
        )
        assert NOMBRE_HOJA in wb.sheetnames

    def test_desde_componentes_sin_resultado(self, diagnostico_basico):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_desde_componentes(wb, diagnostico_basico)
        assert NOMBRE_HOJA in wb.sheetnames

    def test_desde_componentes_accion_default(self, diagnostico_basico):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_desde_componentes(wb, diagnostico_basico)

        # Acción debe ser "N/A" por defecto
        celda_accion = ws.cell(row=3, column=4)
        assert celda_accion.value == "N/A"


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestFuncionConveniencia:
    """Tests para la función escribir_diagnostico()."""

    def test_funcion_crea_hoja(self, decision_completa):
        wb = Workbook()
        ws = escribir_diagnostico(wb, decision_completa)
        assert NOMBRE_HOJA in wb.sheetnames

    def test_funcion_nombre_custom(self, decision_completa):
        wb = Workbook()
        ws = escribir_diagnostico(wb, decision_completa, nombre_hoja="MI_DIAG")
        assert "MI_DIAG" in wb.sheetnames

    def test_funcion_error_sin_diagnostico(self):
        wb = Workbook()
        decision = DecisionCheckpoint(diagnostico=None)
        with pytest.raises(ValueError):
            escribir_diagnostico(wb, decision)


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestAnchos:
    """Tests para ajuste de anchos de columna."""

    def test_anchos_configurados(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        assert ws.column_dimensions["A"].width == 28
        assert ws.column_dimensions["B"].width == 25
        assert ws.column_dimensions["C"].width == 14
        assert ws.column_dimensions["D"].width == 14
        assert ws.column_dimensions["E"].width == 16
        assert ws.column_dimensions["F"].width == 20


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestMetricas:
    """Tests para métricas en pie de página."""

    def test_metricas_presentes(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any("Campos evaluados" in c for c in contenido)
        assert any("Tasa abstención" in c for c in contenido)
        assert any(VERSION_EXCEL_WRITER in c for c in contenido)

    def test_version_router_en_metricas(self, decision_completa):
        wb = Workbook()
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision_completa)

        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any(VERSION_ROUTER in c for c in contenido)


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestEdgeCases:
    """Tests para casos edge."""

    def test_diagnostico_sin_secciones(self):
        wb = Workbook()
        diag = DiagnosticoExpediente(sinad="EMPTY", timestamp="2026-01-01")
        decision = DecisionCheckpoint(
            accion="CONTINUAR",
            diagnostico=diag,
            timestamp="2026-01-01",
        )
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision)
        assert NOMBRE_HOJA in wb.sheetnames

    def test_resultado_sin_campos_abstencion(self):
        wb = Workbook()
        diag = DiagnosticoExpediente(sinad="NOCAMPOS", timestamp="2026-01-01")
        resultado = ResultadoRouter(
            timestamp="2026-01-01",
            resultados_abstencion=[],
        )
        decision = DecisionCheckpoint(
            accion="CONTINUAR",
            diagnostico=diag,
            resultado=resultado,
            timestamp="2026-01-01",
        )
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision)

        # Debe mostrar mensaje "Sin campos evaluados"
        contenido = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    contenido.append(str(cell))

        assert any("Sin campos evaluados" in c for c in contenido)

    def test_workbook_con_otras_hojas_no_afectadas(self, decision_completa):
        wb = Workbook()
        ws_anexo = wb.active
        ws_anexo.title = "Anexo 3"
        ws_anexo.cell(row=1, column=1).value = "Dato original"

        escritor = EscritorDiagnostico()
        escritor.escribir_hoja_diagnostico(wb, decision_completa)

        # La hoja Anexo 3 debe seguir intacta
        assert "Anexo 3" in wb.sheetnames
        assert wb["Anexo 3"].cell(row=1, column=1).value == "Dato original"

    def test_valor_largo_truncado(self):
        """Valores mayores a 60 chars se truncan con '...'."""
        wb = Workbook()
        campo_largo = CampoExtraido(
            nombre_campo="descripcion",
            valor="A" * 100,
            archivo="test.pdf",
            pagina=1,
            confianza=0.80,
            metodo=MetodoExtraccion.OCR,
            tipo_campo="texto",
        )
        diag = DiagnosticoExpediente(sinad="LARGO", timestamp="2026-01-01")
        resultado = ResultadoRouter(
            timestamp="2026-01-01",
            resultados_abstencion=[
                ResultadoAbstencion(campo=campo_largo, debe_abstenerse=False),
            ],
        )
        decision = DecisionCheckpoint(
            accion="CONTINUAR",
            diagnostico=diag,
            resultado=resultado,
            timestamp="2026-01-01",
        )
        escritor = EscritorDiagnostico()
        ws = escritor.escribir_hoja_diagnostico(wb, decision)

        # Buscar campo y verificar truncado
        for row in ws.iter_rows(min_row=1, max_row=30):
            if row[0].value == "descripcion":
                assert len(str(row[1].value)) <= 63  # 60 + "..."
                assert str(row[1].value).endswith("...")
                return
        pytest.fail("Campo descripcion no encontrado")

    def test_multiples_escrituras_reemplazan(self, decision_completa, decision_ok):
        """Escribir dos veces reemplaza la hoja anterior."""
        wb = Workbook()
        escritor = EscritorDiagnostico()

        escritor.escribir_hoja_diagnostico(wb, decision_completa)
        escritor.escribir_hoja_diagnostico(wb, decision_ok)

        count = sum(1 for s in wb.sheetnames if s == NOMBRE_HOJA)
        assert count == 1

        # Debe tener datos del segundo (OK)
        ws = wb[NOMBRE_HOJA]
        celda_status = ws.cell(row=3, column=2)
        assert celda_status.value == "OK"


@pytest.mark.skipif(not OPENPYXL_DISPONIBLE, reason="openpyxl no instalado")
class TestIntegracionConExistentes:
    """Tests de integración con las clases existentes del proyecto."""

    def test_diagnostico_to_rows_coherente_con_writer(self, diagnostico_basico):
        """to_rows() produce datos coherentes con lo que el writer escribe."""
        rows = diagnostico_basico.to_rows()
        assert len(rows) > 0
        for row in rows:
            assert "seccion" in row
            assert "status" in row
            assert "mensaje" in row

    def test_resultado_abstencion_format_spec_coherente(self, campo_abstencion):
        """get_excel_format_spec() de ResultadoAbstencion es coherente."""
        res = ResultadoAbstencion(
            campo=campo_abstencion,
            debe_abstenerse=True,
            razon_abstencion="Confianza baja",
            razon_codigo=RazonAbstencion.CONFIANZA_BAJA,
        )
        spec = res.get_excel_format_spec()
        assert spec["bg_color"] == "FF0000"  # Rojo
        assert spec["comment"] == "Confianza baja"

    def test_campo_extraido_clasificar_status(
        self, campo_legible, campo_incompleto, campo_abstencion
    ):
        """clasificar_status() produce valores que el writer entiende."""
        assert campo_legible.clasificar_status() == EvidenceStatus.LEGIBLE
        assert campo_abstencion.clasificar_status() == EvidenceStatus.ILEGIBLE
        # Incompleto depende de umbrales, pero al menos no falla
        status = campo_incompleto.clasificar_status()
        assert status in (
            EvidenceStatus.LEGIBLE,
            EvidenceStatus.INCOMPLETO,
            EvidenceStatus.ILEGIBLE,
        )
