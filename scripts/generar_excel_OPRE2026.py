"""
Generador de Excel para expediente OPRE2026-INT-0131766
Prueba real de extracción de datos — AG-EVIDENCE v2.0

Hojas:
  1. Anexo3 — Rendición de cuentas por comisión de servicios
  2. DeclaracionJurada — Declaración Jurada del comisionado
  3. Comprobantes — Detalle tipo registro de compras SUNAT
  4. BoardingPass — Datos de vuelos, boarding pass y tiquete aéreo
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.join(
    "C:", os.sep, "Users", "Hans", "Proyectos", "AG-EVIDENCE",
    "data", "expedientes", "pruebas", "viaticos_2026",
    "OPRE2026-INT-0131766_12.FEB.2026_"
)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "RENDICION_OPRE2026-INT-0131766_v2.xlsx")

# ============================================================
# ESTILOS
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, is_money=False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP
    if is_money:
        cell.number_format = MONEY_FORMAT
        cell.alignment = RIGHT


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ============================================================
# HOJA 1: ANEXO 3 — Rendición de Cuentas
# ============================================================
def crear_hoja_anexo3(wb):
    ws = wb.active
    ws.title = "Anexo3"

    ws.merge_cells("A1:H1")
    ws["A1"] = "ANEXO N°3 — RENDICIÓN DE CUENTAS POR COMISIÓN DE SERVICIOS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Unidad Ejecutora", "024 MINISTERIO DE EDUCACION-SEDE CENTRAL - 024"),
        ("Nro. Identificación", "Q00079"),
        ("Comisionado", "SERNA ALVA LILYA PAOLA"),
        ("N° Planilla", "00036"),
        ("N° Exp. SIAF", "00001721"),
        ("N° Comprobante", "2601661"),
        ("Motivo", "Acompañamiento y asistencia comunicacional al Sr. Ministro durante actividades en las regiones de Piura y Tumbes"),
        ("DNI", "18140959"),
        ("CEL", "982770978"),
        ("Dirección", "AV. COSTANERA 1030 DPTO 507 - SAN MIGUEL"),
        ("CAS-JEFATURA / SINAD", "131766"),
        ("Salida", "04/02/2026"),
        ("Regreso", "07/02/2026"),
        ("N° Días/Horas", "3 días 8 horas"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Tabla de detalle de gastos
    row += 1
    ws.cell(row=row, column=1, value="DETALLE DEL GASTO").font = SUBTITLE_FONT
    row += 1

    headers = ["FECHA", "DOCUMENTO", "NÚMERO", "RAZÓN SOCIAL", "CONCEPTO", "IMPORTE S/"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    gastos = [
        ("04/02/2026", "Factura", "E001-1450",       "RAMOS TOLA MATEO",               "MOVILIDAD",     38.00),
        ("05/02/2026", "Factura", "F201-00007418",    "INVERSIONES TURISTICAS B Y S SAC","HOSPEDAJE",    140.00),
        ("04/02/2026", "Factura", "FS35-00071268",    "LASINO SA",                       "ALIMENTACION",  31.00),
        ("04/02/2026", "Factura", "FB04-004459",      "PICANTERIA SEÑOR CHICHERIO SAC",  "ALIMENTACION",  20.00),
        ("04/02/2026", "Factura", "FB04-004458",      "PICANTERIA SEÑOR CHICHERIO SAC",  "ALIMENTACION",  39.00),
        ("05/02/2026", "Factura", "FU01-011573",      "TIO LENGUADO RESTAURANTES EIRL",  "ALIMENTACION",  55.00),
        ("05/02/2026", "Factura", "FO12-00007900",    "EL CHALAN SAC",                   "ALIMENTACION",  23.90),
        ("06/02/2026", "Factura", "F040-00439678",    "COSTA DEL SOL SA",                "HOSPEDAJE",    241.92),
        ("07/02/2026", "Factura", "E001-29",          "GRUPO EMPRESARIAL PEMO SAC",      "HOSPEDAJE",    120.00),
        ("07/02/2026", "Factura", "F012-00018700",    "PROACCION EMPRESARIAL SAC",       "MOVILIDAD",     85.00),
        ("07/02/2026", "Factura", "F001-00009765",    "RESTAURANT A&W EIRL",             "ALIMENTACION", 136.00),
        ("07/02/2026", "Factura", "FO13-00006007",    "BGG COSTA SAC",                   "ALIMENTACION",  88.50),
    ]

    for g in gastos:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 6))
        row += 1

    # Resumen
    row += 1
    total_con_doc = sum(g[5] for g in gastos)
    total_sin_doc = 187.00
    resumen = [
        ("(1) GASTOS CON DOCUMENTACIÓN SUSTENTATORIA", total_con_doc),
        ("(2) GASTOS SIN DOCUMENTACIÓN SUSTENTATORIA", total_sin_doc),
        ("(3) TOTAL GASTADO (1 + 2)", total_con_doc + total_sin_doc),
        ("REEMBOLSO", 0.00),
        ("(4) DEVOLUCIÓN", 1440.00 - (total_con_doc + total_sin_doc)),
        ("(5) MONTO RECIBIDO (3 + 4)", 1440.00),
    ]

    for label, val in resumen:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=6, value=val)
        style_data_cell(cell, is_money=True)
        cell.font = Font(bold=True, size=10)
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# HOJA 2: DECLARACIÓN JURADA
# ============================================================
def crear_hoja_dj(wb):
    ws = wb.create_sheet("DeclaracionJurada")

    ws.merge_cells("A1:F1")
    ws["A1"] = "DECLARACIÓN JURADA DEL COMISIONADO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Unidad Ejecutora", "024 - MINISTERIO DE EDUCACION-SEDE CENTRAL"),
        ("Declarante", "LILYA PAOLA SERNA ALVA"),
        ("DNI", "18140959"),
        ("N° Expediente SINAD", "131766"),
        ("N° Planilla Viáticos", "00036"),
        ("Fecha de Salida", "04/02/2026"),
        ("Fecha de Regreso", "07/02/2026"),
        ("Fecha del Documento", "12/02/2026"),
        ("Motivo de Comisión", "Acompañamiento y asistencia comunicacional al Sr. Ministro "
                               "durante actividades en las regiones de Piura y Tumbes"),
        ("Declaración", "Declara bajo juramento que los gastos presentados en la rendición "
                        "de cuentas corresponden a gastos reales efectuados durante la comisión de servicio"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Tabla de gastos sin comprobante (Anexo N°4)
    row += 1
    ws.cell(row=row, column=1, value="DETALLE DE GASTOS SIN COMPROBANTE DE PAGO").font = SUBTITLE_FONT
    row += 1

    headers_dj = ["N°", "FECHA", "CONCEPTO DE GASTO", "CONCEPTO", "IMPORTE S/"]
    for i, h in enumerate(headers_dj, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_dj))
    row += 1

    gastos_sin_cp = [
        (1, "04/02/2026", "Movilidad de aeropuerto a hospedaje Piura", "MOVILIDAD", 35.00),
        (2, "04/02/2026", "Movilidad de hotel a construcción del colegio José Carlos Mariátegui Lachira", "MOVILIDAD", 12.00),
        (3, "04/02/2026", "Movilidad de Traslado de colegio José Carlos Mariátegui Lachira", "MOVILIDAD", 15.00),
        (4, "06/02/2026", "Movilidad de hotel a I.E. 004 en Tumbes", "MOVILIDAD", 20.00),
        (5, "06/02/2026", "Movilidad de Traslado de I.E. 004 a UGEL Tumbes y UGEL", "MOVILIDAD", 10.00),
        (6, "06/02/2026", "Compra de 4 botellas de agua en Tumbes", "ALIMENTACION", 20.00),
        (7, "06/02/2026", "Movilidad hospedaje - GORE Piura", "MOVILIDAD", 15.00),
        (8, "06/02/2026", "Compra de snacks, jugo especial, pan con queso y gaseosa", "ALIMENTACION", 30.00),
        (9, "07/02/2026", "Traslado hospedaje a Aeropuerto de Tumbes", "MOVILIDAD", 30.00),
    ]

    for g in gastos_sin_cp:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 5))
        row += 1

    # Total gastos sin comprobante
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="TOTAL GASTOS SIN COMPROBANTE DE PAGO").font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).border = THIN_BORDER
    total_sin_cp = sum(g[4] for g in gastos_sin_cp)
    cell = ws.cell(row=row, column=5, value=total_sin_cp)
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    # Resumen
    row += 2
    resumen_dj = [
        ("Total Movilidad", sum(g[4] for g in gastos_sin_cp if g[3] == "MOVILIDAD")),
        ("Total Alimentación", sum(g[4] for g in gastos_sin_cp if g[3] == "ALIMENTACION")),
        ("TOTAL GENERAL", total_sin_cp),
    ]
    for label, val in resumen_dj:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=5, value=val)
        style_data_cell(cell, is_money=True)
        if "GENERAL" in label:
            cell.font = Font(bold=True, size=10, color="FF0000")
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# HOJA 3: COMPROBANTES DE PAGO — Detalle tipo Registro de Compras SUNAT
# ============================================================
def crear_hoja_comprobantes(wb):
    ws = wb.create_sheet("Comprobantes")

    ws.merge_cells("A1:T1")
    ws["A1"] = "REGISTRO DE COMPROBANTES DE PAGO — DETALLE TIPO SUNAT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:T2")
    ws["A2"] = "Expediente: OPRE2026-INT-0131766 | Comisionada: SERNA ALVA LILYA PAOLA | DNI: 18140959"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = [
        "N°",                           # 1
        "Fecha Emisión",                 # 2
        "Tipo Comprobante",              # 3
        "Comprobante Electrónico",       # 4
        "Serie-Número",                  # 5
        "RUC Proveedor",                 # 6
        "Razón Social Proveedor",        # 7
        "Dirección Proveedor",           # 8
        "Cliente (Señor/es)",            # 9
        "RUC Cliente",                   # 10
        "Dirección Cliente",             # 11
        "Concepto / Qué consumió",       # 12
        "Detalle Ítems",                 # 13
        "Forma de Pago",                 # 14
        "Valor Venta (Base Imponible)",  # 15
        "IGV S/",                        # 16
        "% IGV Aplicado",               # 17
        "Otros Cargos (RC/Servicio)",    # 18
        "Importe Total S/",             # 19
        "Observaciones",                 # 20
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    comprobantes = [
        # === COMPROBANTE 1: TAXI IDA ===
        {
            "n": 1,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-1450",
            "ruc_prov": "10423461565",
            "razon_social_prov": "RAMOS TOLA MATEO",
            "dir_prov": "T 12 COO. AMERICA, San Juan de Miraflores - Lima - Lima",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193",
            "concepto": "MOVILIDAD (Transporte domicilio San Miguel a Aeropuerto Callao)",
            "detalle_items": "1 UND — Servicio de transporte domicilio San Miguel a Aeropuerto Callao: S/38.00",
            "forma_pago": "Al Contado",
            "valor_venta": 38.00,
            "igv": 0.00,
            "pct_igv": "EXONERADO",
            "otros_cargos": 0.00,
            "importe_total": 38.00,
            "obs": "Servicio exonerado de IGV. Persona natural con negocio.",
        },
        # === COMPROBANTE 2: HOSPEDAJE PIURA ===
        {
            "n": 2,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "F201-00007418",
            "ruc_prov": "20526539916",
            "razon_social_prov": "INVERSIONES TURISTICAS B Y S S.A.C. (SUCHE PARK HOTEL)",
            "dir_prov": "AV. LOS COCOS 472 MZA. N LOTE. 8 URB. GRAU, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193, LIMA - LIMA - SAN BORJA",
            "concepto": "HOSPEDAJE — Alojamiento 1 noche (04-Feb a 05-Feb 2026). Hab 303",
            "detalle_items": "1 NIU — Alojamiento. Huésped: SERNA ALVA, LILYA PAOLA. Reserva: 117784. Usuario: Elena Archenti",
            "forma_pago": "Contado",
            "valor_venta": 118.64,
            "igv": 21.36,
            "pct_igv": "18%",
            "otros_cargos": 0.00,
            "importe_total": 140.00,
            "obs": "IGV 18% régimen general. Telf: +(51 73) 613439. Web: www.sucheparkhotel.com",
        },
        # === COMPROBANTE 3: STARBUCKS PIURA ===
        {
            "n": 3,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FS35-00071268",
            "ruc_prov": "20388829452",
            "razon_social_prov": "LASINO S.A. (STARBUCKS COFFEE — STB35 PIURA)",
            "dir_prov": "AV. JAVIER PRADO OESTE 1650, LIMA LIMA SAN ISIDRO (sede fiscal)",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193, LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACIÓN (Refrigerio/Merienda)",
            "detalle_items": "1 Pavita & Queso: S/12.50 | 1 Beb Zumo Naranja: S/9.00 | 1 GRDE Té TUN HIB Ice: S/9.50",
            "forma_pago": "Efectivo (pagó S/101.00, vuelto S/70.00)",
            "valor_venta": 25.42,
            "igv": 4.58,
            "pct_igv": "18%",
            "otros_cargos": 1.00,
            "importe_total": 31.00,
            "obs": "RC (Recargo Consumo): S/1.00. ICBPER: S/0.00. Hora: 16:21:27. Cajero: 353024 MIGUEL L. Ticket: 20260204-05-000038211. CHK: 5137. Canal: Counter",
        },
        # === COMPROBANTE 4: PICANTERIA CHICHERIO (Cena 1) ===
        {
            "n": 4,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FB04-004459",
            "ruc_prov": "20606583541",
            "razon_social_prov": "PICANTERIA SEÑOR CHICHERIO S.A.C.",
            "dir_prov": "AV. FRANCISCO BOLOGNESI 390, PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO 193",
            "concepto": "ALIMENTACIÓN (Cena)",
            "detalle_items": "1 TOLLITO ALIÑADO: S/20.00",
            "forma_pago": "Efectivo S/20.00 — Al Contado",
            "valor_venta": 18.18,
            "igv": 1.82,
            "pct_igv": "10%",
            "otros_cargos": 0.00,
            "importe_total": 20.00,
            "obs": "IGV 10% MYPE (Ley 31556+32219). Hora: 05:39:32 PM. Cajero: MARIBEL. Mesa: Salón Principal. Atendido por: FERNANDA. Software: Byte Restaurantes",
        },
        # === COMPROBANTE 5: PICANTERIA CHICHERIO (Cena 2) ===
        {
            "n": 5,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FB04-004458",
            "ruc_prov": "20606583541",
            "razon_social_prov": "PICANTERIA SEÑOR CHICHERIO S.A.C.",
            "dir_prov": "AV. FRANCISCO BOLOGNESI 390, PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO 193",
            "concepto": "ALIMENTACIÓN (Cena)",
            "detalle_items": "1 INCA KOLA 1 LT.: S/10.00 | 1 CABRITO: S/29.00",
            "forma_pago": "Efectivo S/39.00 — Al Contado",
            "valor_venta": 35.45,
            "igv": 3.55,
            "pct_igv": "10%",
            "otros_cargos": 0.00,
            "importe_total": 39.00,
            "obs": "IGV 10% MYPE (Ley 31556+32219). Hora: 05:39:11 PM. Cajero: MARIBEL. Mesa: Salón Principal. Atendido por: MARIA. Software: Byte Restaurantes",
        },
        # === COMPROBANTE 6: TIO LENGUADO ===
        {
            "n": 6,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FU01-011573",
            "ruc_prov": "20610274774",
            "razon_social_prov": "TIO LENGUADO RESTAURANTES E.I.R.L.",
            "dir_prov": "LT. 31 MZ. C4 URB. BELLO HORIZONTE, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193, LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACIÓN (Almuerzo)",
            "detalle_items": "1 COCA COLA: S/5.00 | 1 SUDADO DE CABEZA DE MERO: S/50.00",
            "forma_pago": "Al Contado",
            "valor_venta": 50.00,
            "igv": 5.00,
            "pct_igv": "10%",
            "otros_cargos": 0.00,
            "importe_total": 55.00,
            "obs": "IGV 10% MYPE (Ley 31556+32219). Hora: 14:23:12. Cajero: MARIANA. Telf: 965880996",
        },
        # === COMPROBANTE 7: EL CHALAN ===
        {
            "n": 7,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FO12-00007900",
            "ruc_prov": "20102351038",
            "razon_social_prov": "EL CHALAN S.A.C.",
            "dir_prov": "CALLE TACNA N° 520, PIURA (Central) / Plaza de Armas Calle Tacna 520-526, Piura - Piura - Piura. Sede Productiva: Mza 250 Lote 6 Z I Zona Industrial",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193, LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACIÓN (Refrigerio)",
            "detalle_items": "1 CREMOLADA JUMBO UND: S/14.90 | 2 AGUA MINERAL C/GAS X 625ML UND: S/4.50 c/u = S/9.00",
            "forma_pago": "Contado — Venta Directa",
            "valor_venta": 19.43,
            "igv": 3.50,
            "pct_igv": "18%",
            "otros_cargos": 0.97,
            "importe_total": 23.90,
            "obs": "RC (Recargo Consumo): S/0.97. ICBPER: S/0.00. Hora: 4:31 PM. Vendedor: LILIANA VILCHERREZ AVELLANEDA. Telf: (073) 306483 / Delivery 987 343 632. Web: www.elchalan.com.pe",
        },
        # === COMPROBANTE 8: COSTA DEL SOL ===
        {
            "n": 8,
            "fecha": "06/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "F040-00439678",
            "ruc_prov": "20231843460",
            "razon_social_prov": "COSTA DEL SOL S.A.",
            "dir_prov": "DOMICILIO FISCAL: AV. SALAVERRY 3060 OFIC. 402-B, MAGDALENA DEL MAR - LIMA. SUCURSAL: JR. SAN MARTIN 275, TUMBES - TUMBES",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CALLE DEL COMERCIO 193 SAN BORJA LIMA SAN BORJA",
            "concepto": "HOSPEDAJE — Habitación y Desayuno, 1 noche (05-Feb a 06-Feb 2026). Hab 503",
            "detalle_items": "1 UND — HABITACION Y DESAYUNO. Huésped: SERNA ALVA, LILYA PAOLA. Reserva: 1558/2026. Ingreso: 5/02/2026. Salida: 6/02/2026. 1 noche",
            "forma_pago": "Contado",
            "valor_venta": 189.00,
            "igv": 34.02,
            "pct_igv": "18%",
            "otros_cargos": 18.90,
            "importe_total": 241.92,
            "obs": "Servicio 10% (DL 25988): S/18.90. IGV 18% régimen general. Tasa del día: 3.400. C.A.: 1378/2026. Sede: CDSW TUMBES. Telf: (01) 200 9222 / (072) 523 991. Web: www.costadelsolperu.com. Nota: derechos de crédito transferidos a patrimonio fideicometido.",
        },
        # === COMPROBANTE 9: PEMO HOSPEDAJE TUMBES ===
        {
            "n": 9,
            "fecha": "07/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-29",
            "ruc_prov": "20614255677",
            "razon_social_prov": "GRUPO EMPRESARIAL PEMO S.A.C.",
            "dir_prov": "PANAMERICANA NORTE KM. 1242, CORRALES - TUMBES - TUMBES",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193, LIMA - LIMA - SAN BORJA",
            "concepto": "HOSPEDAJE — Servicio de hospedaje del día 06 de febrero, habitación individual",
            "detalle_items": "1 UNIDAD — Servicio de hospedaje del día 06 de febrero, habitación individual. Huésped: LILYA PAOLA SERNA ALVA DNI 18140959",
            "forma_pago": "Contado",
            "valor_venta": 108.60,
            "igv": 11.40,
            "pct_igv": "10%",
            "otros_cargos": 0.00,
            "importe_total": 120.00,
            "obs": "OBSERVACIÓN: Comprobante parcialmente cortado/recortado en el escaneo. La imagen no muestra el documento completo. Se recomienda devolver al área usuaria para re-escaneo. Datos extraídos de la porción visible. | IGV 10% MYPE (Ley 31556+32219). Valor unitario: S/108.5972. Establecimiento en Panamericana Norte Km 1242, Corrales, Tumbes",
        },
        # === COMPROBANTE 10: TAXI RETORNO LIMA ===
        {
            "n": 10,
            "fecha": "07/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "F012-00018700",
            "ruc_prov": "20518220218",
            "razon_social_prov": "PROACCION EMPRESARIAL S.A.C. (TAXI OFICIAL DEL AEROPUERTO JORGE CHAVEZ DE LIMA)",
            "dir_prov": "AV. MARISCAL LA MAR NRO. 120 INT. 501, URB. SANTA CRUZ - MIRAFLORES - LIMA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193, SAN BORJA",
            "concepto": "MOVILIDAD (Aeropuerto Jorge Chávez a domicilio Av. Nicolás 1087, San Miguel)",
            "detalle_items": "1 — Servicio de taxi. Destino: Aeropuerto Internacional Jorge Chávez a Calle Av. Nicolás 1087, San Miguel, Perú",
            "forma_pago": "Contado",
            "valor_venta": 65.00,
            "igv": 0.00,
            "pct_igv": "EXONERADO",
            "otros_cargos": 0.00,
            "importe_total": 65.00,
            "obs": "Servicio exonerado de IGV. Operador: PROACCION EMPRESARIAL S.A.C.",
        },
        # === COMPROBANTE 11: EDUARDO EL BRUJO ===
        {
            "n": 11,
            "fecha": "07/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "F001-00009765",
            "ruc_prov": "20609176467",
            "razon_social_prov": "RESTAURANT A & W EMPRESA INDIVIDUAL DE RESPONSABILIDAD LIMITADA (EDUARDO EL BRUJO FAMILY)",
            "dir_prov": "AV. FAUSTINO PIAGGIO NRO. 072 (AL COSTADO DEL MALECON), TUMBES - CONTRALMIRANTE VILLAR - ZORRITOS",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 - LIMA LIMA SAN BORJA",
            "concepto": "ALIMENTACIÓN (Almuerzo)",
            "detalle_items": "1 Ceviche De Conchas Negras PERSONAL: S/70.00 | 1 Inka Zero botella: S/6.00 | 1 Pescado Frito Entero (Solo Ensalada Y Yuca Frita): S/60.00",
            "forma_pago": "Contado (Efectivo: S/100.00 + S/100.00 = S/200.00, Vuelto: S/64.00)",
            "valor_venta": 123.64,
            "igv": 12.36,
            "pct_igv": "10%",
            "otros_cargos": 0.00,
            "importe_total": 136.00,
            "obs": "OBSERVACIÓN: Factura consigna 2 platos principales para 1 persona: Ceviche de Conchas Negras (S/70.00) + Pescado Frito Entero (S/60.00) = S/130.00 solo en platos fuertes. El consumo resulta desproporcionado para comisión individual de 1 persona. Someter a evaluación del especialista. | IGV 10% MYPE (Ley 31556+32219). Operación: 1290010. Cajero: MARIBEL CHAPA. Mozo: AMAYA FIORELLA. Mesa: A20. 1 persona. Email: restaurantayw@gmail.com",
        },
        # === COMPROBANTE 12: RUTTA TRAVEL BREAK ===
        {
            "n": 12,
            "fecha": "07/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "FO13-00006007",
            "ruc_prov": "20556626902",
            "razon_social_prov": "BGG COSTA S.A.C. (RUTTA TRAVEL BREAK TUMBES)",
            "dir_prov": "AEROPUERTO CAP. FAP PEDRO CANGA RODRIGUEZ, CARRETERA PANAMERICANA NORTE KM 1276, TUMBES",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 - LIMA LIMA SAN BORJA 15, SAN BORJA",
            "concepto": "ALIMENTACIÓN (Refrigerio en aeropuerto Tumbes)",
            "detalle_items": "2 CROISSANT DE POLLO: S/50.00 | 1 TORTA DE CHOCOLATE: S/18.50 | 1 JUGO DE FRESA CON LECHE: S/20.00",
            "forma_pago": "Efectivo S/100.00, Vuelto S/11.50 — Contado",
            "valor_venta": 71.37,
            "igv": 12.85,
            "pct_igv": "18%",
            "otros_cargos": 4.28,
            "importe_total": 88.50,
            "obs": "RC (Recargo Consumo 6%): S/4.28. Op. Inafecta: S/0.00. Hora: 07:02 PM. Cajero/Mesero: CLARIVEL. 4 artículos. Telf: 905450684 / 970474090",
        },
    ]

    for cp in comprobantes:
        values = [
            cp["n"], cp["fecha"], cp["tipo_cp"], cp["electronico"],
            cp["serie_num"], cp["ruc_prov"], cp["razon_social_prov"],
            cp["dir_prov"], cp["cliente"], cp["ruc_cliente"],
            cp["dir_cliente"], cp["concepto"], cp["detalle_items"],
            cp["forma_pago"], cp["valor_venta"], cp["igv"],
            cp["pct_igv"], cp["otros_cargos"], cp["importe_total"], cp["obs"],
        ]
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i in (15, 16, 18, 19)
            style_data_cell(cell, is_money=is_money)
        row += 1

    # Fila de totales
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="TOTALES COMPROBANTES DE PAGO").font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).border = THIN_BORDER

    total_vv = sum(c["valor_venta"] for c in comprobantes)
    total_igv = sum(c["igv"] for c in comprobantes)
    total_otros = sum(c["otros_cargos"] for c in comprobantes)
    total_importe = sum(c["importe_total"] for c in comprobantes)

    for col, val in [(15, total_vv), (16, total_igv), (18, total_otros), (19, total_importe)]:
        cell = ws.cell(row=row, column=col, value=val)
        style_data_cell(cell, is_money=True)
        cell.font = Font(bold=True, size=10)

    auto_width(ws, max_width=50)
    return ws


# ============================================================
# HOJA 4: BOARDING PASS + TIQUETE AÉREO
# ============================================================
def crear_hoja_boarding(wb):
    ws = wb.create_sheet("BoardingPass")

    ws.merge_cells("A1:L1")
    ws["A1"] = "BOARDING PASS / TARJETA DE EMBARQUE + TIQUETE AÉREO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Info del pasajero
    info = [
        ("Pasajera", "LILYA PAOLA SERNA ALVA"),
        ("DNI", "18140959"),
        ("Tipo Pasajero", "Adulto"),
        ("Frequent Flyer", "Gold"),
        ("Aerolínea", "LATAM AIRLINES PERU S.A."),
        ("RUC Aerolínea", "20341841357"),
        ("Dirección Aerolínea", "Av. Santa Cruz 381 Piso 6, Miraflores, Lima, Perú"),
        ("Facturación tributaria a", "MINISTERIO DE EDUCACION — RUC 20131370998"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # === TIQUETE 1: IDA Lima→Piura ===
    row += 1
    ws.cell(row=row, column=1, value="TIQUETE AÉREO — IDA (Lima → Piura)").font = SUBTITLE_FONT
    row += 1

    tiq1_info = [
        ("Código de Reserva", "SYWLDR"),
        ("N° de Orden", "LA5441255XKBU"),
        ("N° de Ticket", "9442271796521"),
        ("Ciudad y Fecha Emisión", "Lima, Perú 02/02/2026"),
    ]
    for label, value in tiq1_info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    headers_v = [
        "Tramo", "N° Vuelo", "Fecha", "Origen", "Destino",
        "Hora Salida", "Hora Llegada", "Asiento", "Cabina", "Tarifa",
        "Hora Puerta Embarque", "Terminal"
    ]
    for i, h in enumerate(headers_v, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_v))
    row += 1

    vuelo_ida = (
        "IDA", "LA 2308", "04/02/2026",
        "LIMA (J. Chavez Intl.)", "PIURA (G. Concha Ibérico)",
        "13:10", "14:50", "9C", "Economy", "Light",
        "12:10 (finaliza 12:40)", "T.Nuevo"
    )
    for i, val in enumerate(vuelo_ida, 1):
        cell = ws.cell(row=row, column=i, value=val)
        style_data_cell(cell)
    row += 1

    # Desglose pago IDA
    row += 1
    ws.cell(row=row, column=1, value="Desglose de Pago — Tiquete IDA").font = Font(bold=True, size=10, color="2F5496")
    row += 1
    headers_p = ["Concepto", "Moneda", "Monto"]
    for i, h in enumerate(headers_p, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_p))
    row += 1

    pagos_ida = [
        ("Vuelo (tarifa)", "USD", 168.00),
        ("Tasas e impuestos", "USD", 42.84),
        ("   — PE (impuestos Perú)", "USD", 30.24),
        ("   — HW", "USD", 12.60),
        ("TOTAL TIQUETE IDA", "USD", 210.84),
        ("Forma de pago", "LATAM Wallet", "USD 210.84"),
    ]
    for p in pagos_ida:
        for i, val in enumerate(p, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = (i == 3 and isinstance(val, (int, float)))
            style_data_cell(cell, is_money=is_money)
            if "TOTAL" in str(p[0]):
                cell.font = Font(bold=True, size=10)
        row += 1

    # === TIQUETE 2: RETORNO Tumbes→Lima ===
    row += 2
    ws.cell(row=row, column=1, value="TIQUETE AÉREO — RETORNO (Tumbes → Lima)").font = SUBTITLE_FONT
    row += 1

    tiq2_info = [
        ("Código de Reserva", "GCVYSH"),
        ("N° de Orden", "LA5442620SSBM"),
        ("Ciudad y Fecha Emisión", "Lima, Perú 02/02/2026"),
        ("Nota", "Tiquete grupal — 8 pasajeros en misma reserva"),
    ]
    for label, value in tiq2_info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    for i, h in enumerate(headers_v, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_v))
    row += 1

    vuelo_ret = (
        "RETORNO", "LA 2341", "07/02/2026",
        "TUMBES (P. Canga Rodriguez)", "LIMA (J. Chavez Intl.)",
        "20:00", "21:50", "14F", "Economy", "Light",
        "19:00 (finaliza 19:30)", "Terminal 1"
    )
    for i, val in enumerate(vuelo_ret, 1):
        cell = ws.cell(row=row, column=i, value=val)
        style_data_cell(cell)
    row += 1

    # Desglose pago RETORNO (grupal)
    row += 1
    ws.cell(row=row, column=1, value="Desglose de Pago — Tiquete RETORNO (grupal 8 pasajeros)").font = Font(bold=True, size=10, color="2F5496")
    row += 1
    for i, h in enumerate(headers_p, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_p))
    row += 1

    pagos_ret = [
        ("Vuelo (tarifa, 8 pasajeros)", "USD", 1000.00),
        ("Tasas e impuestos", "USD", 229.12),
        ("   — PE (impuestos Perú)", "USD", 180.00),
        ("TOTAL TIQUETE RETORNO (8 pax)", "USD", 1229.12),
        ("Forma de pago", "LATAM Wallet", "USD 1,229.12"),
    ]
    for p in pagos_ret:
        for i, val in enumerate(p, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = (i == 3 and isinstance(val, (int, float)))
            style_data_cell(cell, is_money=is_money)
            if "TOTAL" in str(p[0]):
                cell.font = Font(bold=True, size=10)
        row += 1

    # Lista de pasajeros en tiquete retorno grupal
    row += 2
    ws.cell(row=row, column=1, value="Pasajeros en Tiquete Retorno Grupal").font = Font(bold=True, size=10, color="2F5496")
    row += 1
    headers_pax = ["N°", "Nombre Pasajero", "Tipo", "DNI/Documento", "N° Ticket"]
    for i, h in enumerate(headers_pax, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_pax))
    row += 1

    pasajeros = [
        (1, "ALEKSANDAR RAFAJLOVSKY BURGOS", "Adulto", "70058315",  "5442271800090"),
        (2, "JULIO PISUA GONZALES",           "Adulto", "06073045",  "5442271800091"),
        (3, "CLAUDIA HUAIRA VEGA",            "Adulto", "77096663",  "5442271800089"),
        (4, "JEFFERSON CARRERA HUAMAN",        "Adulto", "46692501",  "5442271800085"),
        (5, "LUIS RUIZ AYMARA",                "Adulto", "10234713",  "5442271800086"),
        (6, "MINERVA MORA ALVINO",             "Adulto", "42947931",  "5442271800087"),
        (7, "MARCOS MEDINA ANTONIO",            "Adulto", "43589542",  "5442271800088"),
        (8, "LILYA SERNA ALVA",                "Adulto", "18140959",  "5442271800084"),
    ]

    for pax in pasajeros:
        for i, val in enumerate(pax, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell)
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()

    print("Creando Hoja 1: Anexo 3...")
    crear_hoja_anexo3(wb)

    print("Creando Hoja 2: Declaración Jurada...")
    crear_hoja_dj(wb)

    print("Creando Hoja 3: Comprobantes (detalle SUNAT)...")
    crear_hoja_comprobantes(wb)

    print("Creando Hoja 4: Boarding Pass + Tiquete Aéreo...")
    crear_hoja_boarding(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Excel generado: {OUTPUT_FILE}")
    print(f"   Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
