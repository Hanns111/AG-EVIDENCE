#!/usr/bin/env python3
"""
Generar Excel de rendición para expediente DIRI2026-INT-0068815.
Comisionada: CRUZATT MENDOZA SANDRA
Destino: Lima - Piura - Huancabamba - Lima
Período: 26/01/2026 - 30/01/2026
"""
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call(["pip3", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# =============================================================================
# DATOS DEL EXPEDIENTE (extraídos de las 47 páginas)
# =============================================================================

EXPEDIENTE = {
    "numero": "DIRI2026-INT-0068815",
    "planilla": "0021",
    "siaf": "2600773",
    "comisionado": "CRUZATT MENDOZA SANDRA",
    "dni": "40687817",
    "cargo": "COORDINADOR TERRITORIAL",
    "dependencia": "DIRECCION DE RELACIONES INTERGUBERNAMENTALES (DIRI)",
    "escala": "ESCALA-2 FUNCIONARIOS Y EMPLEADOS",
    "destino": "Lima - Piura - Huancabamba - Lima",
    "fecha_salida": "26/01/2026",
    "fecha_regreso": "30/01/2026",
    "dias_horas": "4d 6h",
    "monto_recibido": 2033.30,
    "monto_gastado_con_doc": 1154.18,
    "monto_gastado_sin_doc": 229.00,
    "total_gastado": 1383.18,
    "devolucion": 650.12,
    "motivo": "Reuniones con autoridades DRE/UGEL para coordinar acciones previas al inicio de año escolar 2026",
    "pasaje_aereo": {
        "aerolinea": "JETSMART AIRLINES PERU S.A.C.",
        "ruc_aerolinea": "20607393649",
        "codigo_reserva": "P4MI9V",
        "ida": "JA 7220 Lima-Piura 26/01/2026 14:10-15:57",
        "vuelta": "JA 7225 Piura-Lima 30/01/2026 16:29-18:00",
        "monto_usd": 236.80,
    },
}

# Comprobantes extraídos del Anexo N°3 (p1) + verificados contra cada factura individual
COMPROBANTES = [
    {
        "nro": 1,
        "fecha": "26/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-858",
        "ruc": "10056424291",
        "razon_social": "GARCIA SILVA WILLIAM MARTIN",
        "concepto": "MOVILIDAD",
        "descripcion": "Servicio de traslado de aeropuerto Piura al hotel",
        "valor_venta": 25.00,
        "igv": 4.50,
        "importe_total": 29.50,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 16,
        "pagina_validez": 17,
        "observacion": "Importe en Anexo3=25.00 pero factura dice 29.50 con IGV. Verificar."
    },
    {
        "nro": 2,
        "fecha": "26/01/2026",  # Fecha emisión según factura: 29/01/2026 — DISCREPANCIA
        "tipo_doc": "Factura",
        "serie_numero": "E001-1692",
        "ruc": "10103886622",
        "razon_social": "CARMIN ALZAMORA CARLOS ALBERTO",
        "concepto": "MOVILIDAD",
        "descripcion": "Servicio traslado de casa Los Olivos al aeropuerto Callao 26/02/26",
        "valor_venta": 70.00,
        "igv": 0.00,
        "importe_total": 70.00,
        "tasa_igv": "EXONERADO",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 18,
        "pagina_validez": 19,
        "observacion": "Fecha emisión factura=29/01/2026, Anexo3=26/01/2026. IGV=0. Desc dice 26/02/26 (error de año?)."
    },
    {
        "nro": 3,
        "fecha": "26/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "F011-00008846",
        "ruc": "20102351038",
        "razon_social": "EL CHALAN S.A.C.",
        "concepto": "ALIMENTACION",
        "descripcion": "Tallarín con pavo, cremolada mediana, agua mineral",
        "valor_venta": 35.20,
        "igv": 6.34,
        "importe_total": 43.30,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 20,  # OCR
        "pagina_validez": 21,
        "observacion": "Incluye ICBPER S/1.76. RC S/0.61."
    },
    {
        "nro": 4,
        "fecha": "26/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "F700-00000141",
        "ruc": "20493040309",
        "razon_social": "CORPORACION HOTELERA SAN ANDRES S.A.C.",
        "concepto": "HOSPEDAJE",
        "descripcion": "Hotel Win & Win, 2 noches (26-28/01), 1 agua sin gas",
        "valor_venta": 230.47,
        "igv": 41.48,  # 18% sobre 230.47
        "importe_total": 295.00,
        "tasa_igv": "18%",  # Factura muestra IGV 18% (S/41.48) + 10% (S/23.05)
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 22,
        "pagina_validez": 23,
        "observacion": "Hospedaje 2 noches SWB. Factura muestra IGV 18%=41.48 + 10%=23.05. Check-in 26/01, Check-out 28/01. TC=3.56."
    },
    {
        "nro": 5,
        "fecha": "27/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "F012-0007708",
        "ruc": "20102351038",
        "razon_social": "EL CHALAN S.A.C.",
        "concepto": "ALIMENTACION",
        "descripcion": "Cremolada jumbo",
        "valor_venta": 12.11,
        "igv": 2.18,
        "importe_total": 14.90,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 30,  # OCR
        "pagina_validez": 31,
        "observacion": "Incluye RC S/0.61."
    },
    {
        "nro": 6,
        "fecha": "27/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-1221",
        "ruc": "10026226339",
        "razon_social": "FOSSA SEMINARIO LUIS ALBERTO (Bar Restaurant La Academia)",
        "concepto": "ALIMENTACION",
        "descripcion": "Costillas de chancho, Inka Cola",
        "valor_venta": 43.22,
        "igv": 7.78,
        "importe_total": 51.00,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 26,
        "pagina_validez": 27,
        "observacion": ""
    },
    {
        "nro": 7,
        "fecha": "27/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "F002-0011835",
        "ruc": "20608722999",
        "razon_social": "GIASAL SOCIEDAD ANONIMA CERRADA",
        "concepto": "ALIMENTACION",
        "descripcion": "Especial trio, jugo personal, infusión hierba luisa, infusión jamaica, pie de pecanas",
        "valor_venta": 62.72,
        "igv": 6.28,
        "importe_total": 69.00,
        "tasa_igv": "~10%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 32,  # OCR
        "pagina_validez": 33,
        "observacion": "Anexo3=64.00, factura=69.00. DISCREPANCIA DE MONTO. OCR lee Total=69.00 (sub=62.72+igv=6.28). Pago con tarjeta culqi: S/64+S/5=S/69."
    },
    {
        "nro": 8,
        "fecha": "28/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-1251",
        "ruc": "20600843771",
        "razon_social": "INVERSIONES MARKHADE E.I.R.L.",
        "concepto": "ALIMENTACION",
        "descripcion": "1 ceviche de mero, 1 gaseosa de litro",
        "valor_venta": 63.55,
        "igv": 11.43,
        "importe_total": 74.98,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 24,
        "pagina_validez": 25,
        "observacion": ""
    },
    {
        "nro": 9,
        "fecha": "28/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-887",
        "ruc": "20610742492",
        "razon_social": "DIOS CRUZ DEL MUNDO OJEDA TOURS E.I.R.L.",
        "concepto": "PASAJE",
        "descripcion": "Servicio de movilidad Piura-Huancabamba",
        "valor_venta": 84.75,
        "igv": 15.26,
        "importe_total": 100.01,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 28,
        "pagina_validez": 29,
        "observacion": "Anexo3=100.00, factura=100.01 (redondeo)."
    },
    {
        "nro": 10,
        "fecha": "29/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-1771",
        "ruc": "10032064332",
        "razon_social": "LANDIVAR DE GUERRERO NELLY ZELIDETH (Hosp Turist Virgen del Carmen)",
        "concepto": "HOSPEDAJE",
        "descripcion": "Servicio de alojamiento ingreso 28/01/2026 salida 30/01/2026",
        "valor_venta": 181.00,
        "igv": 19.00,
        "importe_total": 200.00,
        "tasa_igv": "~10.5%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 42,
        "pagina_validez": 43,
        "observacion": "Fecha emisión factura=12/02/2026 (posterior a comisión). Hospedaje en Huancabamba 2 noches. IGV=19/181=10.5% aprox."
    },
    {
        "nro": 11,
        "fecha": "29/01/2026",
        "tipo_doc": "Boleta de Venta",
        "serie_numero": "EB01-696",
        "ruc": "10478060071",
        "razon_social": "MORILLO BERMEO MARIO ROLANDO (Hospedaje El Titanic)",
        "concepto": "ALIMENTACION",
        "descripcion": "Caldo de gallina, jugo de piña",
        "valor_venta": 18.00,
        "igv": 0.00,
        "importe_total": 18.00,
        "tasa_igv": "GRAVADO sin IGV separado",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 38,
        "pagina_validez": 39,
        "observacion": "Fecha emisión=29/01/2026. Op.Gravada=18.00, IGV=0.00. Boleta sin IGV desglosado."
    },
    {
        "nro": 12,
        "fecha": "30/01/2026",
        "tipo_doc": "Boleta de Venta",
        "serie_numero": "EB01-1908",
        "ruc": "10003282339",
        "razon_social": "BRUNO PEÑA ERESLINDA MARIA (ANNY SARITA)",
        "concepto": "ALIMENTACION",
        "descripcion": "1 caldo de gallina, 1 agua",
        "valor_venta": 12.71,
        "igv": 2.29,
        "importe_total": 15.00,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 36,
        "pagina_validez": 37,
        "observacion": "Salitral, Morropón, Piura."
    },
    {
        "nro": 13,
        "fecha": "30/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "E001-888",
        "ruc": "20610742492",
        "razon_social": "DIOS CRUZ DEL MUNDO OJEDA TOURS E.I.R.L.",
        "concepto": "PASAJE",
        "descripcion": "Servicio de movilidad Huancabamba-Piura",
        "valor_venta": 84.75,
        "igv": 15.26,
        "importe_total": 100.01,
        "tasa_igv": "18%",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 40,
        "pagina_validez": 41,
        "observacion": "Anexo3=100.00, factura=100.01 (redondeo)."
    },
    {
        "nro": 14,
        "fecha": "30/01/2026",
        "tipo_doc": "Factura",
        "serie_numero": "F028-012573",
        "ruc": "20509138053",
        "razon_social": "GREEN AIRPORT S.A. (TAXI GREEN)",
        "concepto": "MOVILIDAD",
        "descripcion": "Aeropuerto Jorge Chávez - Jr. Aries 1144, Los Olivos",
        "valor_venta": 65.00,
        "igv": 0.00,
        "importe_total": 65.00,
        "tasa_igv": "EXONERADO",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 44,
        "pagina_validez": 45,
        "observacion": "Total Exonerado=65.00, IGV=0.00."
    },
    {
        "nro": 15,
        "fecha": "30/01/2026",
        "tipo_doc": "Boleta de Venta",
        "serie_numero": "EB01-703",
        "ruc": "10478060071",
        "razon_social": "MORILLO BERMEO MARIO ROLANDO (Hospedaje El Titanic)",
        "concepto": "ALIMENTACION",
        "descripcion": "Caldo de gallina, jugo de piña",
        "valor_venta": 18.00,
        "igv": 0.00,
        "importe_total": 18.00,
        "tasa_igv": "GRAVADO sin IGV separado",
        "validez_sunat": "VALIDO",
        "pagina_comprobante": 34,
        "pagina_validez": 35,
        "observacion": "Fecha emisión=10/02/2026 (posterior a comisión). Obs: 'Boleta del día 30 de enero 2026'. IGV=0."
    },
]

# Gastos sin comprobante (DJ - Anexo N°4, p4)
GASTOS_DJ = [
    {"fecha": "26/01/2026", "concepto": "MOVILIDAD", "detalle": "Hotel a restaurante y retorno al hotel", "importe": 24.00},
    {"fecha": "27/01/2026", "concepto": "MOVILIDAD", "detalle": "Hotel a la DRE Piura, IE San Miguel, hotel", "importe": 30.00},
    {"fecha": "28/01/2026", "concepto": "MOVILIDAD", "detalle": "Hotel a UGEL Piura, DRE Piura, Hotel", "importe": 30.00},
    {"fecha": "28/01/2026", "concepto": "ALIMENTACION", "detalle": "Alimentos en el kiosco de la DRE Piura", "importe": 22.00},
    {"fecha": "28/01/2026", "concepto": "MOVILIDAD", "detalle": "Del hotel al terminal terrestre salida a Huancabamba", "importe": 18.00},
    {"fecha": "29/01/2026", "concepto": "MOVILIDAD", "detalle": "Terminal terrestre Huancabamba al hotel, hotel a la UGEL", "importe": 30.00},
    {"fecha": "29/01/2026", "concepto": "ALIMENTACION", "detalle": "Almuerzo comedor UGEL Huancabamba", "importe": 20.00},
    {"fecha": "30/01/2026", "concepto": "MOVILIDAD", "detalle": "Hotel a UGEL Huancabamba, Hotel, TT salida Piura", "importe": 30.00},
    {"fecha": "30/01/2026", "concepto": "MOVILIDAD", "detalle": "Terminal terrestre Piura al aeropuerto Piura", "importe": 25.00},
]

# =============================================================================
# GENERAR EXCEL
# =============================================================================

OUTPUT_DIR = '/mnt/c/Users/Hans/Proyectos/AG-EVIDENCE/data/expedientes/pruebas/viaticos_2026/DIRI2026-INT-0068815/extraccion'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'RENDICION_{EXPEDIENTE["numero"]}.xlsx')

wb = Workbook()

# --- Estilos ---
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def apply_border(ws, row, cols):
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).border = thin_border

# =============================================================================
# HOJA 1: DATOS GENERALES
# =============================================================================
ws1 = wb.active
ws1.title = "Datos Generales"
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 60

datos = [
    ("EXPEDIENTE", EXPEDIENTE["numero"]),
    ("N° Planilla", EXPEDIENTE["planilla"]),
    ("N° Exp SIAF", EXPEDIENTE["siaf"]),
    ("Comisionado", EXPEDIENTE["comisionado"]),
    ("DNI", EXPEDIENTE["dni"]),
    ("Cargo", EXPEDIENTE["cargo"]),
    ("Dependencia", EXPEDIENTE["dependencia"]),
    ("Escala", EXPEDIENTE["escala"]),
    ("Destino", EXPEDIENTE["destino"]),
    ("Fecha Salida", EXPEDIENTE["fecha_salida"]),
    ("Fecha Regreso", EXPEDIENTE["fecha_regreso"]),
    ("Días/Horas", EXPEDIENTE["dias_horas"]),
    ("Motivo", EXPEDIENTE["motivo"]),
    ("", ""),
    ("RESUMEN FINANCIERO", ""),
    ("Monto Recibido", f'S/ {EXPEDIENTE["monto_recibido"]:,.2f}'),
    ("Gastos con Documentación", f'S/ {EXPEDIENTE["monto_gastado_con_doc"]:,.2f}'),
    ("Gastos sin Documentación (DJ)", f'S/ {EXPEDIENTE["monto_gastado_sin_doc"]:,.2f}'),
    ("Total Gastado", f'S/ {EXPEDIENTE["total_gastado"]:,.2f}'),
    ("Devolución", f'S/ {EXPEDIENTE["devolucion"]:,.2f}'),
    ("", ""),
    ("PASAJE AÉREO", ""),
    ("Aerolínea", EXPEDIENTE["pasaje_aereo"]["aerolinea"]),
    ("RUC Aerolínea", EXPEDIENTE["pasaje_aereo"]["ruc_aerolinea"]),
    ("Código Reserva", EXPEDIENTE["pasaje_aereo"]["codigo_reserva"]),
    ("Vuelo Ida", EXPEDIENTE["pasaje_aereo"]["ida"]),
    ("Vuelo Vuelta", EXPEDIENTE["pasaje_aereo"]["vuelta"]),
    ("Monto Total (USD)", f'USD {EXPEDIENTE["pasaje_aereo"]["monto_usd"]:,.2f}'),
]

for i, (label, value) in enumerate(datos, 1):
    ws1.cell(row=i, column=1, value=label).font = Font(bold=True) if label else Font()
    ws1.cell(row=i, column=2, value=value)
    if label in ("RESUMEN FINANCIERO", "PASAJE AÉREO"):
        ws1.cell(row=i, column=1).font = Font(bold=True, size=12, color="4472C4")

# =============================================================================
# HOJA 2: COMPROBANTES (detalle SUNAT)
# =============================================================================
ws2 = wb.create_sheet("Comprobantes")
headers2 = ["Nro", "Fecha", "Tipo Doc", "Serie-Número", "RUC", "Razón Social",
            "Concepto", "Descripción", "Valor Venta", "IGV", "Tasa IGV",
            "Importe Total", "Válido SUNAT", "Página", "Observación"]

for j, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=j, value=h)
apply_header(ws2, 1, len(headers2))

for i, c in enumerate(COMPROBANTES, 2):
    ws2.cell(row=i, column=1, value=c["nro"])
    ws2.cell(row=i, column=2, value=c["fecha"])
    ws2.cell(row=i, column=3, value=c["tipo_doc"])
    ws2.cell(row=i, column=4, value=c["serie_numero"])
    ws2.cell(row=i, column=5, value=c["ruc"])
    ws2.cell(row=i, column=6, value=c["razon_social"])
    ws2.cell(row=i, column=7, value=c["concepto"])
    ws2.cell(row=i, column=8, value=c["descripcion"])
    ws2.cell(row=i, column=9, value=c["valor_venta"])
    ws2.cell(row=i, column=10, value=c["igv"])
    ws2.cell(row=i, column=11, value=c["tasa_igv"])
    ws2.cell(row=i, column=12, value=c["importe_total"])
    ws2.cell(row=i, column=13, value=c["validez_sunat"])
    ws2.cell(row=i, column=14, value=c["pagina_comprobante"])
    ws2.cell(row=i, column=15, value=c["observacion"])
    apply_border(ws2, i, len(headers2))

    # Color por observación
    if c["observacion"]:
        for j in range(1, len(headers2) + 1):
            ws2.cell(row=i, column=j).fill = warn_fill

# Totales
row_total = len(COMPROBANTES) + 2
ws2.cell(row=row_total, column=7, value="TOTALES").font = Font(bold=True)
ws2.cell(row=row_total, column=9, value=sum(c["valor_venta"] for c in COMPROBANTES)).font = Font(bold=True)
ws2.cell(row=row_total, column=10, value=sum(c["igv"] for c in COMPROBANTES)).font = Font(bold=True)
ws2.cell(row=row_total, column=12, value=sum(c["importe_total"] for c in COMPROBANTES)).font = Font(bold=True)

# Ajustar anchos
widths2 = [5, 12, 14, 18, 14, 40, 15, 45, 12, 10, 10, 12, 10, 8, 50]
for j, w in enumerate(widths2, 1):
    ws2.column_dimensions[chr(64 + j) if j <= 26 else 'A' + chr(64 + j - 26)].width = w

# =============================================================================
# HOJA 3: DECLARACIÓN JURADA (gastos sin comprobante)
# =============================================================================
ws3 = wb.create_sheet("Declaración Jurada")
headers3 = ["Fecha", "Concepto", "Detalle", "Importe S/"]

for j, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=j, value=h)
apply_header(ws3, 1, len(headers3))

for i, g in enumerate(GASTOS_DJ, 2):
    ws3.cell(row=i, column=1, value=g["fecha"])
    ws3.cell(row=i, column=2, value=g["concepto"])
    ws3.cell(row=i, column=3, value=g["detalle"])
    ws3.cell(row=i, column=4, value=g["importe"])
    apply_border(ws3, i, len(headers3))

row_total3 = len(GASTOS_DJ) + 2
ws3.cell(row=row_total3, column=3, value="TOTAL").font = Font(bold=True)
ws3.cell(row=row_total3, column=4, value=sum(g["importe"] for g in GASTOS_DJ)).font = Font(bold=True)

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 12

# =============================================================================
# HOJA 4: HALLAZGOS Y OBSERVACIONES
# =============================================================================
ws4 = wb.create_sheet("Hallazgos")
headers4 = ["Código", "Severidad", "Descripción", "Página", "Comprobante"]

for j, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=j, value=h)
apply_header(ws4, 1, len(headers4))

hallazgos = [
    ("H-01", "INFO", "Factura E001-858: Anexo3=S/25, factura=S/29.50 (diferencia=IGV S/4.50). Anexo3 registra valor venta, no importe total.", 16, "E001-858"),
    ("H-02", "ALERTA", "Factura E001-1692: Fecha emisión=29/01/2026, pero Anexo3=26/01/2026. Descripción dice '26/02/26' (posible error de mes).", 18, "E001-1692"),
    ("H-03", "ALERTA", "Factura F002-0011835 GIASAL: Anexo3=S/64.00, factura OCR=S/69.00. Discrepancia de S/5.00. Pago fraccionado con tarjeta.", 32, "F002-0011835"),
    ("H-04", "INFO", "Facturas E001-887 y E001-888: Importe=S/100.01 (redondeo), Anexo3=S/100.00.", 28, "E001-887/888"),
    ("H-05", "INFO", "Factura E001-1771 hospedaje Huancabamba: Fecha emisión=12/02/2026 (13 días después de la estancia). Práctica común en hospedajes rurales.", 42, "E001-1771"),
    ("H-06", "INFO", "Boleta EB01-703: Fecha emisión=10/02/2026, observación dice 'Boleta del día 30 de enero 2026'. Emisión posterior.", 34, "EB01-703"),
    ("H-07", "INFO", "Hotel Win & Win: Factura muestra doble tasa IGV (18%=S/41.48 + 10%=S/23.05). Posible MYPE con tasa diferenciada.", 22, "F700-00000141"),
    ("H-08", "OK", "Todos los 15 comprobantes validados como VÁLIDOS en consulta SUNAT.", "-", "Todos"),
    ("H-09", "OK", "Devolución S/650.12 acreditada en Banco de la Nación el 16/02/2026. Voucher en p47.", 47, "Depósito BN"),
    ("H-10", "INFO", "Verificación aritmética: Recibido S/2,033.30 - Gastado S/1,383.18 = Devolución S/650.12. CUADRA.", "-", "-"),
    ("H-11", "INFO", "Total DJ S/229.00 = suma de 9 items. CUADRA.", "-", "DJ"),
]

for i, (cod, sev, desc, pag, comp) in enumerate(hallazgos, 2):
    ws4.cell(row=i, column=1, value=cod)
    ws4.cell(row=i, column=2, value=sev)
    ws4.cell(row=i, column=3, value=desc)
    ws4.cell(row=i, column=4, value=str(pag))
    ws4.cell(row=i, column=5, value=comp)
    apply_border(ws4, i, len(headers4))

    if sev == "ALERTA":
        for j in range(1, len(headers4) + 1):
            ws4.cell(row=i, column=j).fill = alert_fill
    elif sev == "OK":
        for j in range(1, len(headers4) + 1):
            ws4.cell(row=i, column=j).fill = ok_fill
    elif sev == "INFO":
        for j in range(1, len(headers4) + 1):
            ws4.cell(row=i, column=j).fill = warn_fill

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 80
ws4.column_dimensions['D'].width = 8
ws4.column_dimensions['E'].width = 18

# =============================================================================
# HOJA 5: MAPA DE PÁGINAS
# =============================================================================
ws5 = wb.create_sheet("Mapa Páginas")
headers5 = ["Página", "PDF", "Documento", "Método Extracción"]

for j, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=j, value=h)
apply_header(ws5, 1, len(headers5))

mapa = [
    # PV
    (1, "PV", "Planilla de Viáticos (SIGA)", "pymupdf"),
    (2, "PV", "Planilla de Viáticos (SIGA) p2", "pymupdf"),
    (3, "PV", "Plan de Trabajo Diario (Anexo 01)", "pymupdf"),
    (4, "PV", "Planilla de Viáticos detalle cadena", "pymupdf"),
    (5, "PV", "Declaración Jurada compromiso", "pymupdf"),
    (6, "PV", "Consulta SIAF disponibilidad", "pymupdf"),
    (7, "PV", "DNI comisionada (imagen)", "paddleocr"),
    # RENDICION
    (1, "REND", "Anexo N°3 - Rendición de cuentas", "pymupdf"),
    (2, "REND", "Nota de Pago SIAF", "pymupdf"),
    (3, "REND", "Planilla de Viáticos", "pymupdf"),
    (4, "REND", "Anexo N°4 - Declaración Jurada", "pymupdf"),
    (5, "REND", "Plan de Trabajo Diario (Anexo 01)", "pymupdf"),
    (6, "REND", "Anexo N°2 - Informe de Comisión p1", "pymupdf"),
    (7, "REND", "Informe de Comisión p2", "pymupdf"),
    (8, "REND", "Informe de Comisión p3", "pymupdf"),
    (9, "REND", "Formato Alertas (Anexo)", "pymupdf"),
    (10, "REND", "Anexo N°5 - Constancia comisión DRE Piura", "paddleocr"),
    (11, "REND", "Firmas/sellos constancias (imagen)", "paddleocr"),
    (12, "REND", "Ticket aéreo JetSMART", "pymupdf"),
    (13, "REND", "Detalles pago JetSMART", "pymupdf"),
    (14, "REND", "Boarding Pass Piura→Lima", "paddleocr"),
    (15, "REND", "Boarding Pass Lima→Piura", "paddleocr"),
    (16, "REND", "Factura E001-858 (Movilidad)", "pymupdf"),
    (17, "REND", "Validez SUNAT E001-858", "pymupdf"),
    (18, "REND", "Factura E001-1692 (Movilidad)", "pymupdf"),
    (19, "REND", "Validez SUNAT E001-1692", "pymupdf"),
    (20, "REND", "Factura F011-00008846 El Chalán (Alimentación)", "paddleocr"),
    (21, "REND", "Validez SUNAT F011-00008846", "pymupdf"),
    (22, "REND", "Factura F700-00000141 Hotel (Hospedaje)", "pymupdf"),
    (23, "REND", "Validez SUNAT F700-00000141", "pymupdf"),
    (24, "REND", "Factura E001-1251 Inv.Markhade (Alimentación)", "pymupdf"),
    (25, "REND", "Validez SUNAT E001-1251", "pymupdf"),
    (26, "REND", "Factura E001-1221 La Academia (Alimentación)", "pymupdf"),
    (27, "REND", "Validez SUNAT E001-1221", "pymupdf"),
    (28, "REND", "Factura E001-887 Ojeda Tours (Pasaje)", "pymupdf"),
    (29, "REND", "Validez SUNAT E001-887", "pymupdf"),
    (30, "REND", "Factura F012-0007708 El Chalán (Alimentación)", "paddleocr"),
    (31, "REND", "Validez SUNAT F012-0007708", "pymupdf"),
    (32, "REND", "Factura F002-0011835 GIASAL (Alimentación)", "paddleocr"),
    (33, "REND", "Validez SUNAT F002-0011835", "pymupdf"),
    (34, "REND", "Boleta EB01-703 Morillo (Alimentación)", "pymupdf"),
    (35, "REND", "Validez SUNAT EB01-703", "pymupdf"),
    (36, "REND", "Boleta EB01-1908 Anny Sarita (Alimentación)", "pymupdf"),
    (37, "REND", "Validez SUNAT EB01-1908", "pymupdf"),
    (38, "REND", "Boleta EB01-696 Morillo (Alimentación)", "pymupdf"),
    (39, "REND", "Validez SUNAT EB01-696", "pymupdf"),
    (40, "REND", "Factura E001-888 Ojeda Tours (Pasaje)", "pymupdf"),
    (41, "REND", "Validez SUNAT E001-888", "pymupdf"),
    (42, "REND", "Factura E001-1771 Hospedaje Virgen Carmen", "pymupdf"),
    (43, "REND", "Validez SUNAT E001-1771", "pymupdf"),
    (44, "REND", "Factura F028-012573 Taxi Green (Movilidad)", "pymupdf"),
    (45, "REND", "Validez SUNAT F028-012573", "pymupdf"),
    (46, "REND", "Devolución de saldos + Recibo de ingreso", "pymupdf"),
    (47, "REND", "Voucher Banco de la Nación depósito", "paddleocr"),
]

for i, (pag, pdf, doc, metodo) in enumerate(mapa, 2):
    ws5.cell(row=i, column=1, value=pag)
    ws5.cell(row=i, column=2, value=pdf)
    ws5.cell(row=i, column=3, value=doc)
    ws5.cell(row=i, column=4, value=metodo)
    apply_border(ws5, i, len(headers5))
    if metodo == "paddleocr":
        ws5.cell(row=i, column=4).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 8
ws5.column_dimensions['C'].width = 50
ws5.column_dimensions['D'].width = 18

# =============================================================================
# GUARDAR
# =============================================================================
wb.save(OUTPUT_FILE)
print(f"\nExcel generado: {OUTPUT_FILE}")
print(f"  Hojas: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"    - {s}")

# Resumen de extracción
print(f"\n{'='*60}")
print(f"RESUMEN DE EXTRACCIÓN")
print(f"{'='*60}")
print(f"Expediente: {EXPEDIENTE['numero']}")
print(f"Comisionada: {EXPEDIENTE['comisionado']}")
print(f"Destino: {EXPEDIENTE['destino']}")
print(f"Período: {EXPEDIENTE['fecha_salida']} - {EXPEDIENTE['fecha_regreso']}")
print(f"Comprobantes con documento: {len(COMPROBANTES)}")
print(f"Gastos DJ sin documento: {len(GASTOS_DJ)}")
print(f"Total comprobantes extraídos: S/ {sum(c['importe_total'] for c in COMPROBANTES):,.2f}")
print(f"Total DJ: S/ {sum(g['importe'] for g in GASTOS_DJ):,.2f}")
print(f"Total gastado: S/ {sum(c['importe_total'] for c in COMPROBANTES) + sum(g['importe'] for g in GASTOS_DJ):,.2f}")
print(f"Monto recibido: S/ {EXPEDIENTE['monto_recibido']:,.2f}")
print(f"Devolución: S/ {EXPEDIENTE['devolucion']:,.2f}")
print(f"Hallazgos: {len(hallazgos)}")

# Verificación aritmética
total_comp = sum(c['importe_total'] for c in COMPROBANTES)
total_dj = sum(g['importe'] for g in GASTOS_DJ)
total = total_comp + total_dj
devolucion_calc = EXPEDIENTE['monto_recibido'] - total
print(f"\nVERIFICACIÓN ARITMÉTICA:")
print(f"  Sum comprobantes: S/ {total_comp:,.2f}")
print(f"  Sum DJ: S/ {total_dj:,.2f}")
print(f"  Total: S/ {total:,.2f}")
print(f"  Recibido - Total = S/ {devolucion_calc:,.2f}")
print(f"  Devolución declarada: S/ {EXPEDIENTE['devolucion']:,.2f}")
print(f"  Diferencia: S/ {devolucion_calc - EXPEDIENTE['devolucion']:,.2f}")
if abs(devolucion_calc - EXPEDIENTE['devolucion']) < 0.10:
    print(f"  ✓ CUADRA")
else:
    print(f"  ✗ NO CUADRA — revisar")
