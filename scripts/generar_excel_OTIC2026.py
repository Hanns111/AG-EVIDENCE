"""
Generador de Excel para expediente OTIC2026-INT-0115085
Rendicion de viaticos — Tacna (02-04 Feb 2026)
Comisionado: ZENOZAIN FLORES JACK EDWARDS (DNI 40765970)

Hojas:
  1. Anexo3 — Rendicion de cuentas por comision de servicios (refleja Anexo 3 tal cual)
  2. DJ — Declaracion Jurada (gastos sin comprobante)
  3. Comprobantes — Detalle tipo registro de compras SUNAT (FUENTE: cada factura)
  4. BoardingPass — Datos de vuelos y tiquete aereo

FUENTE DE DATOS PARA HOJA COMPROBANTES:
  Documento fuente = LA FACTURA (no el Anexo 3).
  Cada comprobante refleja exactamente lo que dice la factura original.
  Datos extraidos via lectura visual directa de cada pagina del PDF (PyMuPDF + imagen).
  Paginas 49-58 del PDF contienen las facturas escaneadas como imagen.

METODO DE EXTRACCION:
  - PyMuPDF (fitz) para renderizar paginas escaneadas como imagen PNG a 300-600 DPI
  - Lectura visual directa de cada imagen de factura
  - Confirmacion de RUCs y montos con zoom a areas especificas

Version: 4 (10 facturas completas desde documento fuente)
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "expedientes",
    "pruebas",
    "viaticos_2026",
    "OTIC2026-INT-0115085_11_02_26",
)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "RENDICION_OTIC2026-INT-0115085.xlsx")

# ============================================================
# ESTILOS
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FORMAT = "#,##0.00"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
ALERT_FONT = Font(name="Calibri", bold=True, size=10, color="CC6600")
NOTE_FONT = Font(name="Calibri", italic=True, size=9, color="666666")
OK_FONT = Font(name="Calibri", bold=True, size=10, color="008000")
WARN_FONT = Font(name="Calibri", bold=True, size=10, color="FF0000")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, is_money=False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP
    if is_money:
        cell.number_format = MONEY_FORMAT
        cell.alignment = RIGHT


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ============================================================
# HOJA 1: ANEXO 3 (refleja el documento Anexo 3 tal cual)
# ============================================================
def crear_hoja_anexo3(wb):
    ws = wb.active
    ws.title = "Anexo3"

    ws.merge_cells("A1:H1")
    ws["A1"] = "ANEXO N.3 - RENDICION DE CUENTAS POR COMISION DE SERVICIOS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Unidad Ejecutora", "026 - PROGRAMA EDUCACION BASICA PARA TODOS"),
        ("Nro. Identificacion", "000081"),
        ("Comisionado", "ZENOZAIN FLORES JACK EDWARDS"),
        ("N. Planilla", "00302"),
        ("N. Exp SIAF", "00001288"),
        ("N. Comprobante", "2600858"),
        (
            "Motivo",
            "Participar como enlace tecnico-operativo del aplicativo Matricula Digital en la UGEL Tacna",
        ),
        ("Direccion/Oficina", "OTIC - Oficina de Tecnologias de la Informacion y Comunicacion"),
        ("DNI", "40765970"),
        ("Domicilio", "JR. SUCCHA 338, BRENA, LIMA"),
        ("CEL", "902736762"),
        ("SINAD", "0115085"),
        (
            "Centro de Costo",
            "026.30.04.01 - OFICINA DE TECNOLOGIAS DE LA INFORMACION Y COMUNICACION",
        ),
        ("Solicitante", "REYES GUTIERREZ JOSE ALBERTO"),
        ("Salida", "02/02/2026 06:00 hrs"),
        ("Regreso", "04/02/2026 22:00 hrs"),
        ("N. Dias/Horas", "2 dias 16 horas"),
        ("Escala", "ESCALA-2 FUNCIONARIOS Y EMPLEADOS"),
        ("Total Viaticos Asignados", "S/960.00"),
        ("N. Cuenta Bancaria", "AHORROS-00110814023202656717"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="DETALLE DEL GASTO (segun Anexo 3)").font = SUBTITLE_FONT
    row += 1

    headers = ["N.", "FECHA", "DOCUMENTO", "NUMERO", "RAZON SOCIAL", "CONCEPTO", "IMPORTE S/"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    gastos = [
        (
            1,
            "02/02/2026",
            "Factura",
            "E001-8998",
            "EMP. DE SERV. TURISTICOS EL MESON S.R.L.",
            "HOSPEDAJE",
            130.00,
        ),
        (
            2,
            "02/02/2026",
            "Factura",
            "E001-390",
            "MANTARI VASQUEZ WALTER BRADY",
            "MOVILIDAD",
            70.00,
        ),
        (
            3,
            "02/02/2026",
            "Factura",
            "FN5E-194",
            "OPERACIONES ARCOS DORADOS DE PERU SA",
            "ALIMENTACION",
            25.00,
        ),
        (4, "02/02/2026", "Factura", "FA01-1245", "ROOFTOP", "ALIMENTACION", 95.00),
        (
            5,
            "03/02/2026",
            "Factura",
            "E001-08998",
            "EMP. DE SERV. TURISTICOS EL MESON S.R.L.",
            "HOSPEDAJE",
            130.00,
        ),
        (6, "03/02/2026", "Factura", "FPP1-027804", "GLORIETA TACNENA SAC", "ALIMENTACION", 96.00),
        (
            7,
            "04/02/2026",
            "Factura",
            "E001-394",
            "MANTARI VASQUEZ WALTER BRADY",
            "MOVILIDAD",
            70.00,
        ),
        (
            8,
            "04/02/2026",
            "Factura",
            "F005-1302",
            "GANADERA MALAGA 1967 E.I.R.L.",
            "ALIMENTACION",
            59.00,
        ),
        (9, "04/02/2026", "Factura", "F001-2309", "NUQANCHIK S.A.C.", "ALIMENTACION", 127.00),
        (10, "04/02/2026", "Factura", "F009-4337", "13 MONJAS S.A.C.", "ALIMENTACION", 26.00),
    ]

    for g in gastos:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 7))
        row += 1

    row += 1
    resumen = [
        ("(1) GASTOS CON DOCUMENTACION SUSTENTATORIA", 828.00),
        ("(2) GASTOS SIN DOCUMENTACION SUSTENTATORIA (DJ)", 128.00),
        ("(3) TOTAL GASTADO (1 + 2)", 956.00),
        ("(4) DEVOLUCION", 4.00),
        ("(5) MONTO RECIBIDO (3 + 4)", 960.00),
    ]

    for label, val in resumen:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=7, value=val)
        style_data_cell(cell, is_money=True)
        cell.font = Font(bold=True, size=10)
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# HOJA 2: DECLARACION JURADA
# ============================================================
def crear_hoja_dj(wb):
    ws = wb.create_sheet("DeclaracionJurada")

    ws.merge_cells("A1:F1")
    ws["A1"] = "ANEXO N.4 - DECLARACION JURADA"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Unidad Ejecutora", "026 - PROGRAMA EDUCACION BASICA PARA TODOS"),
        ("Declarante", "JACK EDWARDS ZENOZAIN FLORES"),
        ("DNI", "40765970"),
        ("Domicilio", "JR. SUCCHA 338, BRENA, LIMA"),
        ("Fecha del documento", "09/02/2026"),
        ("Declaracion", "Gastos donde fue imposible obtener comprobantes de pago"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="DETALLE DE GASTOS SIN COMPROBANTE").font = SUBTITLE_FONT
    row += 1

    headers = ["N.", "FECHA", "CONCEPTO DE GASTO", "TIPO", "IMPORTE S/"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    gastos_dj = [
        (1, "03/02/2026", "YOGURT, JUGO, AGUA, GATORADE", "ALIMENTACION", 35.00),
        (2, "03/02/2026", "HOTEL A UGEL / UGEL A COLEGIOS", "MOVILIDAD", 30.00),
        (3, "03/02/2026", "COLEGIOS A HOTEL", "MOVILIDAD", 25.00),
        (4, "04/02/2026", "AGUA, SNACK, GALLETAS, FRUTOS SECOS", "ALIMENTACION", 38.00),
    ]

    for g in gastos_dj:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 5))
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="TOTAL S/").font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell = ws.cell(row=row, column=5, value=128.00)
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    auto_width(ws)
    return ws


# ============================================================
# HOJA 3: COMPROBANTES — FUENTE: CADA FACTURA (documento fuente)
# ============================================================
def crear_hoja_comprobantes(wb):
    ws = wb.create_sheet("Comprobantes")

    ws.merge_cells("A1:T1")
    ws["A1"] = "REGISTRO DE COMPROBANTES DE PAGO - DETALLE TIPO SUNAT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:T2")
    ws["A2"] = (
        "Expediente: OTIC2026-INT-0115085 | Comisionado: ZENOZAIN FLORES JACK EDWARDS | DNI: 40765970"
    )
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:T3")
    ws["A3"] = (
        "FUENTE: Documento fuente = cada factura individual. Datos extraidos por lectura visual directa del PDF (PyMuPDF imagen)."
    )
    ws["A3"].font = NOTE_FONT
    ws["A3"].alignment = Alignment(horizontal="center")

    row = 5
    headers = [
        "N.",
        "Fecha Emision",
        "Tipo Comprobante",
        "Comprobante Electronico",
        "Serie-Numero",
        "RUC Proveedor",
        "Razon Social Proveedor",
        "Direccion Proveedor",
        "Cliente (Senior/es)",
        "RUC/DNI Cliente",
        "Direccion Cliente",
        "Concepto / Descripcion",
        "Detalle Items",
        "Forma de Pago",
        "Valor Venta (Base Imponible)",
        "IGV S/",
        "% IGV Aplicado",
        "Otros cargos",
        "Importe Total S/",
        "Observaciones",
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    DIR_CLIENTE = "CAL. DEL COMERCIO 193 RES. LAS TORRES DE SAN BORJA, LIMA - LIMA - SAN BORJA"

    # -------------------------------------------------------
    # 10 COMPROBANTES — datos de cada factura (documento fuente)
    # -------------------------------------------------------
    comprobantes = [
        {  # FACTURA 1 — El Meson Hotel (Pag 50 y 53 del PDF)
            "n": 1,
            "fecha": "03/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "E001-8998",
            "ruc_prov": "20119395543",
            "razon_social_prov": "EMP. DE SERVICIOS TURISTICOS EL MESON S.R.L. (EL MESON HOTEL)",
            "dir_prov": "CAL. UNANUE 175, TACNA - TACNA - TACNA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": DIR_CLIENTE,
            "concepto": "ALOJAMIENTO 02 Y 03/02/26 (2 noches)",
            "detalle_items": "2.00 UNIDAD - Alojamiento @ S/118.181818 c/u. ICBPER: S/0.00",
            "forma_pago": "Contado",
            "valor_venta": 236.36,
            "igv": 23.64,
            "pct_igv": "10% (MYPE hotel)",
            "otros": 0.00,
            "importe_total": 260.00,
            "obs": "Observacion en factura: JACK EDWARDS ZENOZAIN FLORES - DNI: 40765970. Factura cubre 2 noches.",
        },
        {  # FACTURA 2 — Mantari Vasquez taxi ida (Pag 49)
            "n": 2,
            "fecha": "02/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "E001-390",
            "ruc_prov": "10409510928",
            "razon_social_prov": "MANTARI VASQUEZ WALTER BRADY",
            "dir_prov": "F 7 RINCONADA DE PANDO, SAN MIGUEL - LIMA - LIMA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": DIR_CLIENTE,
            "concepto": "SERVICIO DE TAXI BRENA JR SUCCHA 338 A AEROPUERTO",
            "detalle_items": "1.00 UNIDAD - Taxi casa a aeropuerto S/70.00",
            "forma_pago": "Al Contado",
            "valor_venta": 70.00,
            "igv": 0.00,
            "pct_igv": "0% (EXONERADO)",
            "otros": 0.00,
            "importe_total": 70.00,
            "obs": "Persona natural con negocio (taxi). IGV exonerado.",
        },
        {  # FACTURA 3 — Arcos Dorados / McDonald's (Pag 51)
            "n": 3,
            "fecha": "02/02/2026",
            "tipo_cp": "FACTURA DE VENTA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "FN5E-00000194",
            "ruc_prov": "20376289215",
            "razon_social_prov": "OPERACIONES ARCOS DORADOS DE PERU SA (McDonald's)",
            "dir_prov": "AV. NESTOR GAMBETTA NA*1570 LOCAL AC 30,039 Y A C 30 112, CALLAO - LIMA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CALLE EL COMERCIO 193",
            "concepto": "ALIMENTACION",
            "detalle_items": "1 CG BigMac S/23.02, Coca Cola Gde, Papa Gra, Sin Adicional. ORD #959. Cajero: 74025330 Jelstin Winkler Jara. Hora 14:07",
            "forma_pago": "N/E en factura",
            "valor_venta": 23.02,
            "igv": 4.15,
            "pct_igv": "18%",
            "otros": 0.23,
            "importe_total": 27.40,
            "obs": "RCC (1.00%) = S/0.23. EVS = S/27.40. Restaurante LIM. Resolucion SUNAT Nro. 0180050001988.",
        },
        {  # FACTURA 4 — Rooftop77 Restaurant (Pag 52)
            "n": 4,
            "fecha": "02/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "FA01-00001245",
            "ruc_prov": "20612528641",
            "razon_social_prov": "ROOFTOP77 RESTAURANT E.I.R.L.",
            "dir_prov": "CALLE APURIMAC 277 - 3ER PISO, TACNA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 RES. LAS TORRES DE SAN BORJA, LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (SALON SAN MARTIN)",
            "detalle_items": "1 TABLA PREMIUM GRILL: 1.00 x S/95.00 = S/95.00. Hora 11:23 PM. Descuento S/0.00",
            "forma_pago": "Efectivo S/95.00",
            "valor_venta": 86.36,
            "igv": 8.64,
            "pct_igv": "10% (MYPE restaurante)",
            "otros": 0.00,
            "importe_total": 95.00,
            "obs": "Telf: 956225505. Resolucion SUNAT Nro. 034-005-0005294. Op. Exonerada S/0.00.",
        },
        {  # FACTURA 5 — Glorieta Tacnena (Pag 54)
            "n": 5,
            "fecha": "03/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "FPP1-027804",
            "ruc_prov": "20532631433",
            "razon_social_prov": "GLORIETA TACNENA SOCIEDAD ANONIMA CERRADA (GLORIETA TACNENA S.A.C.)",
            "dir_prov": "AV. JORGE BASADRE GROHMANN NRO. 335 FND PAGO AYMARA, TACNA - TACNA - TACNA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 RES. LAS TORRES DE SAN BORJA - LIMA LIMA SAN BORJA",
            "concepto": "ALIMENTACION",
            "detalle_items": "[1] ZZ E003 Papa a la Huancaina S/15, [1] ZZ E001 Tamal S/7, [1] ZZ BB03 Jarra de Chicha S/14, [1] ZZ P004 Parrillada de Cordero (3P) S/60",
            "forma_pago": "Efectivo (Contado) S/96.00",
            "valor_venta": 87.27,
            "igv": 8.73,
            "pct_igv": "10% (MYPE restaurante)",
            "otros": 0.00,
            "importe_total": 96.00,
            "obs": "Tel: 052-245097 / 962964957 / 978000714. Vuelto S/0.00. Fecha Venc: 03/02/2026.",
        },
        {  # FACTURA 6 — Mantari Vasquez taxi retorno (Pag 58)
            "n": 6,
            "fecha": "08/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "E001-394",
            "ruc_prov": "10409510928",
            "razon_social_prov": "MANTARI VASQUEZ WALTER BRADY",
            "dir_prov": "F 7 RINCONADA DE PANDO, SAN MIGUEL - LIMA - LIMA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": DIR_CLIENTE,
            "concepto": "SERVICIO TAXI AEROPUERTO DE LIMA A BRENA JR SUCCHA 338",
            "detalle_items": "1.00 UNIDAD - Taxi aeropuerto a casa S/70.00. Servicio prestado el 04/02/2026.",
            "forma_pago": "Al Contado",
            "valor_venta": 70.00,
            "igv": 0.00,
            "pct_igv": "0% (EXONERADO)",
            "otros": 0.00,
            "importe_total": 70.00,
            "obs": "Fecha emision en factura: 08/02/2026. Persona natural taxi. IGV exonerado.",
        },
        {  # FACTURA 7 — Ganadera Malaga / La Lecheria (Pag 55)
            "n": 7,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "F005-1302",
            "ruc_prov": "20605128514",
            "razon_social_prov": "GANADERA MALAGA 1967 E.I.R.L. (LA LECHERIA)",
            "dir_prov": "CAL. HIPOLITO UNANUE NRO. 336, TACNA - TACNA - TACNA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CAL. DEL COMERCIO RES. LAS TORRES DE SAN BORJA 193, SAN BORJA - LIMA - DEPARTAMENTO LIMA",
            "concepto": "ALIMENTACION (POR CONSUMO)",
            "detalle_items": "1 POR CONSUMO: P.U. S/59.00, IMP. S/59.00. Cajero: GIU. Para Delivery 3. Hora: 12:27:49",
            "forma_pago": "IZIPAY (Contado)",
            "valor_venta": 53.64,
            "igv": 5.36,
            "pct_igv": "10% (MYPE restaurante)",
            "otros": 0.00,
            "importe_total": 59.00,
            "obs": "Cel: 052-608859. Gratuitas S/0.00. Exoneradas S/0.00.",
        },
        {  # FACTURA 8 — Nuqanchik / Biru (Pag 56)
            "n": 8,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "F001-00002309",
            "ruc_prov": "20604749515",
            "razon_social_prov": "NUQANCHIK S.A.C. (BIRU)",
            "dir_prov": "AV. GUSTAVO PINTO NRO. 18, TACNA - TACNA - TACNA",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 RES. LAS TORRES DE SAN BORJA - LIMA LIMA SAN BORJA",
            "concepto": "ALIMENTACION",
            "detalle_items": "Agua Mineral Con G (San Luis 350ml) S/5, Coca Cola 600ml S/7, Cebiche Carretillero (Corvina con chicharron de pesca) S/60, Arroz con Mariscos y Pescado Grillado Corvina S/54",
            "forma_pago": "Tarjeta Visa S/126.00 + Propina Visa S/1.00",
            "valor_venta": 114.55,
            "igv": 11.45,
            "pct_igv": "10% (MYPE restaurante)",
            "otros": 1.00,
            "importe_total": 127.00,
            "obs": "Tel: 977322071. Operacion 00026232. Mesero: Carlos Jimenez. Mesa: M16. Cant. Personas: 1. Total venta S/126 + Propina S/1 = S/127. Oper Visa 0183-311374.",
        },
        {  # FACTURA 9 — 13 Monjas (Pag 57)
            "n": 9,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA ELECTRONICA",
            "electronico": "SI",
            "serie_num": "F009-00004337",
            "ruc_prov": "20605278745",
            "razon_social_prov": "13 MONJAS S.A.C.",
            "dir_prov": "CAL. SANTA CATALINA NRO. 300 (B1) AREQ UIPA - AREQUIPA - AREQUIPA (13 Monjas Tacna: OTR. CARRETERA PANAMERICANA SUR S/N - SALA DE EMBARQUE, TACNA)",
            "cliente": "PROGRAMA EDUCACION BASICA PARA TODOS",
            "ruc_cliente": "20380795907",
            "dir_cliente": "CAL. DEL COMERCIO NRO. 193 RES. LAS TORRES DE SAN BORJA - LIMA LIMA S",
            "concepto": "ALIMENTACION",
            "detalle_items": "Coca Cola Sabor Original 600ml S/7, Empanada Carne 1 unidad S/12, Inca Kola Sabor Original 600ml S/7. Hora: 16:20:53",
            "forma_pago": "Tarjeta (Open-Visa y Mastercars) S/26.00",
            "valor_venta": 21.13,
            "igv": 2.12,
            "pct_igv": "10% (MYPE restaurante)",
            "otros": 2.75,
            "importe_total": 26.00,
            "obs": "Tel: 940 592 601. Cajero: Caja. RC (Recargo al Consumo) = S/2.75. Sub-Total S/21.13 + IGV S/2.12 + RC S/2.75 = S/26.00. Sede Tacna (principal Arequipa).",
        },
    ]

    for cp in comprobantes:
        values = [
            cp["n"],
            cp["fecha"],
            cp["tipo_cp"],
            cp["electronico"],
            cp["serie_num"],
            cp["ruc_prov"],
            cp["razon_social_prov"],
            cp["dir_prov"],
            cp["cliente"],
            cp["ruc_cliente"],
            cp["dir_cliente"],
            cp["concepto"],
            cp["detalle_items"],
            cp["forma_pago"],
            cp["valor_venta"],
            cp["igv"],
            cp["pct_igv"],
            cp["otros"],
            cp["importe_total"],
            cp["obs"],
        ]
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i in (15, 16, 18, 19) and isinstance(val, (int, float))
            style_data_cell(cell, is_money=is_money)
        row += 1

    # Totales
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="TOTALES (9 facturas)").font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).border = THIN_BORDER

    total_vv = sum(c["valor_venta"] for c in comprobantes)
    total_igv = sum(c["igv"] for c in comprobantes)
    total_otros = sum(c["otros"] for c in comprobantes)
    total_importe = sum(c["importe_total"] for c in comprobantes)

    cell = ws.cell(row=row, column=15, value=round(total_vv, 2))
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    cell = ws.cell(row=row, column=16, value=round(total_igv, 2))
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    cell = ws.cell(row=row, column=18, value=round(total_otros, 2))
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    cell = ws.cell(row=row, column=19, value=round(total_importe, 2))
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    auto_width(ws, max_width=50)
    return ws


# ============================================================
# HOJA 4: BOARDING PASS + TIQUETE AEREO
# ============================================================
def crear_hoja_boarding(wb):
    ws = wb.create_sheet("BoardingPass")

    ws.merge_cells("A1:H1")
    ws["A1"] = "BOARDING PASS / TARJETA DE EMBARQUE + TIQUETE AEREO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Pasajero", "ZENOZAIN FLORES JACK EDWARDS"),
        ("Nombre en tiquete Sky", "JACKS EDWARS ZENOZAIN FLORS (variacion menor)"),
        ("DNI", "40765970"),
        ("Tipo Pasajero", "Adulto"),
        ("Aerolinea", "SKY AIRLINE PERU S.A.C."),
        ("RUC Aerolinea", "20603446543"),
        ("Telefono Aerolinea", "+511 3913600"),
        ("Codigo de Reserva", "GEHZAF"),
        ("Numero de Ticket", "6052109130259"),
        ("Fecha Emision Tiquete", "29/01/2026"),
        ("Numero de Orden", "bbbn125132"),
        ("Tarifa", "Light"),
        ("Comprador / Pagador", "PROGRAMA EDUCACION BASICA PARA TODOS"),
        (
            "Pasajero Acompanante",
            "SOLEDAD ADELA CANAZA ESPEJO (DNI 00490589) - Ticket 6052109130258",
        ),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="DETALLE DE VUELOS").font = SUBTITLE_FONT
    row += 1

    headers = [
        "Tramo",
        "N. Vuelo",
        "Fecha",
        "Origen",
        "Destino",
        "Hora Salida",
        "Hora Llegada",
        "Duracion",
        "Asiento",
        "Tarjeta Embarque",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    vuelos = [
        (
            "IDA",
            "H2 5190",
            "02/02/2026",
            "LIMA (LIM)",
            "TACNA (TCQ)",
            "16:10 (prog. 14:35, retraso)",
            "17:55 aprox",
            "1h 55m",
            "22F",
            "GEHZAF-Light-BN65",
        ),
        (
            "RETORNO",
            "H2 5191",
            "04/02/2026",
            "TACNA (TCQ)",
            "LIMA (LIM)",
            "17:15",
            "19:15",
            "2h 00m",
            "21F",
            "GEHZAF-Light-BN75",
        ),
    ]

    for v in vuelos:
        for i, val in enumerate(v, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell)
        row += 1

    row += 2
    ws.cell(
        row=row, column=1, value="DESGLOSE DE PAGO DEL TIQUETE (ZENOZAIN FLORES)"
    ).font = SUBTITLE_FONT
    row += 1

    headers_p = ["Concepto", "Moneda", "Monto"]
    for i, h in enumerate(headers_p, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_p))
    row += 1

    pagos = [
        ("Pasaje aereo: Lima - Tacna (Ida y Vuelta)", "USD", 288.00),
        ("Impuesto por Ventas (Peru)", "USD", 51.84),
        ("Tasa de Salida Aeroportuaria TUUA (Peru)", "USD", 18.74),
        ("TOTAL POR PASAJERO", "USD", 358.58),
    ]

    for p in pagos:
        for i, val in enumerate(p, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i == 3 and isinstance(val, (int, float))
            style_data_cell(cell, is_money=is_money)
            if p[0].startswith("TOTAL"):
                cell.font = Font(bold=True, size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="RESUMEN TOTAL RESERVA (2 PASAJEROS)").font = SUBTITLE_FONT
    row += 1

    resumen = [
        ("Pasajero 1: SOLEDAD ADELA CANAZA ESPEJO", "USD", 358.58),
        ("Pasajero 2: JACKS EDWARS ZENOZAIN FLORS", "USD", 358.58),
        ("TOTAL RESERVA GEHZAF", "USD", 717.16),
    ]

    for r in resumen:
        for i, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i == 3 and isinstance(val, (int, float))
            style_data_cell(cell, is_money=is_money)
            if r[0].startswith("TOTAL"):
                cell.font = Font(bold=True, size=10)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = (
        "NOTAS: (1) Vuelo IDA H2 5190 con retraso: programado 14:35, despego real ~16:10 (informe de comision). "
        "(2) Reserva compartida con CANAZA ESPEJO. "
        "(3) Equipaje: 1 bolso de mano + 1 equipaje cabina (Light). "
        "(4) Asientos aleatorios: 22F (ida) y 21F (retorno)."
    )
    cell.font = NOTE_FONT
    cell.alignment = WRAP

    auto_width(ws)
    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()

    print("Creando Hoja 1: Anexo 3...")
    crear_hoja_anexo3(wb)

    print("Creando Hoja 2: Declaracion Jurada...")
    crear_hoja_dj(wb)

    print("Creando Hoja 3: Comprobantes (9 facturas, documento fuente)...")
    crear_hoja_comprobantes(wb)

    print("Creando Hoja 4: Boarding Pass + Tiquete Aereo...")
    crear_hoja_boarding(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Excel generado: {OUTPUT_FILE}")
    print(f"   Hojas: {wb.sheetnames}")
    print("   9 facturas (documento fuente) con desglose completo")
    print("   4 items DJ = S/128.00")
    print("   2 vuelos Sky Airline H2 5190/5191")


if __name__ == "__main__":
    main()
