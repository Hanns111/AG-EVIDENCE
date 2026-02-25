"""
Generador de Excel para expediente ODI2026-INT-0139051
Prueba real de extracción de datos — AG-EVIDENCE v2.0

Hojas:
  1. Anexo3 — Rendición de cuentas por comisión de servicios
  2. DJ — Declaración Jurada (gastos sin comprobante)
  3. Comprobantes — Detalle tipo registro de compras SUNAT
  4. BoardingPass — Datos de vuelos y tiquete aéreo
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "expedientes",
    "pruebas",
    "viaticos_2026",
    "ODI2026-INT-0139051_11_02_26",
)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "RENDICION_ODI2026-INT-0139051.xlsx")

# ============================================================
# ESTILOS
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FORMAT = "#,##0.00"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, is_money=False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP
    if is_money:
        cell.number_format = MONEY_FORMAT
        cell.alignment = RIGHT


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ============================================================
# HOJA 1: ANEXO 3 — Rendición de Cuentas
# ============================================================
def crear_hoja_anexo3(wb):
    ws = wb.active
    ws.title = "Anexo3"

    # Encabezado del documento
    ws.merge_cells("A1:H1")
    ws["A1"] = "ANEXO N°3 — RENDICIÓN DE CUENTAS POR COMISIÓN DE SERVICIOS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Datos generales
    info = [
        ("Unidad Ejecutora", "024 - MINISTERIO DE EDUCACION-SEDE CENTRAL"),
        ("Nro. Identificación", "000079"),
        ("Comisionado", "QUEREVALU MIÑAN JOSE FAUSTINO"),
        ("N° Planilla", "00050"),
        ("N° Exp. SIAF", "00001705"),
        ("N° Comprobante", "2601710"),
        (
            "Motivo",
            "Avanzada y acompañamiento social a visita técnica del SM a instituciones educativas de Piura",
        ),
        ("Dirección", "Conjunto residencial Iquique, Callao"),
        ("DNI", "41969732"),
        ("CEL", "997789672"),
        ("SINAD", "0139051"),
        ("Salida", "04/02/2026"),
        ("Regreso", "05/02/2026"),
        ("N° Días/Horas", "1 día 16 horas"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Tabla de detalle de gastos
    row += 1
    ws.cell(row=row, column=1, value="DETALLE DEL GASTO").font = SUBTITLE_FONT
    row += 1

    headers = ["FECHA", "DOCUMENTO", "NÚMERO", "RAZÓN SOCIAL", "CONCEPTO", "IMPORTE S/"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    gastos = [
        ("04/02/2026", "Factura", "E001-11614", "CONTADORES PUBLICOS EIRL", "ALIMENTACION", 70.00),
        ("04/02/2026", "Factura", "E001-11611", "CONTADORES PUBLICOS EIRL", "ALIMENTACION", 50.00),
        ("04/02/2026", "Factura", "E001-1134", "MANRIQUE RIOS JULIO NICANOR", "HOSPEDAJE", 300.00),
        ("05/02/2026", "Factura", "E001-11622", "CONTADORES PUBLICOS EIRL", "ALIMENTACION", 40.00),
        ("05/02/2026", "Factura", "E001-11629", "CONTADORES PUBLICOS EIRL", "ALIMENTACION", 60.00),
        ("05/02/2026", "Factura", "E001-11630", "CONTADORES PUBLICOS EIRL", "ALIMENTACION", 50.00),
    ]

    for g in gastos:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 6))
        row += 1

    # Resumen
    row += 1
    resumen = [
        ("(1) GASTOS CON DOCUMENTACIÓN SUSTENTATORIA", 570.00),
        ("(2) GASTOS SIN DOCUMENTACIÓN SUSTENTATORIA", 70.00),
        ("(3) TOTAL GASTADO (1 + 2)", 640.00),
        ("REEMBOLSO", 0.00),
        ("(4) DEVOLUCIÓN", 0.00),
        ("(5) MONTO RECIBIDO (3 + 4)", 640.00),
    ]

    for label, val in resumen:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=6, value=val)
        style_data_cell(cell, is_money=True)
        cell.font = Font(bold=True, size=10)
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# HOJA 2: DECLARACIÓN JURADA (Anexo 4)
# ============================================================
def crear_hoja_dj(wb):
    ws = wb.create_sheet("DeclaracionJurada")

    ws.merge_cells("A1:F1")
    ws["A1"] = "ANEXO N°4 — DECLARACIÓN JURADA"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Unidad Ejecutora", "024 - MINISTERIO DE EDUCACION-SEDE CENTRAL"),
        ("Nro. Identificación", "0079"),
        ("Declarante", "JOSE FAUSTINO QUEREVALU MIÑAN"),
        ("DNI", "41969732"),
        ("Domicilio", "PJE. EL SOL 397 Int. 203 - URB. BLOCK P"),
        ("Fecha del documento", "10/02/2026"),
        ("Declaración", "Gastos donde fue imposible obtener comprobantes de pago"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="DETALLE DE GASTOS SIN COMPROBANTE").font = SUBTITLE_FONT
    row += 1

    headers = ["FECHA", "CONCEPTO DE GASTO", "CONCEPTO", "IMPORTE S/"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    gastos_dj = [
        ("04/02/2026", "COMPRA DE SNAKS, FRUTA Y AGUA", "ALIMENTACION", 20.00),
        ("04/02/2026", "TRASLADO DE PIURA A CATACAOS", "MOVILIDAD", 15.00),
        ("05/02/2026", "COMPRA DE SNAKS, FRUTA Y AGUA", "ALIMENTACION", 20.00),
        ("05/02/2026", "TRASLADO DE CATACAOS A PIURA", "MOVILIDAD", 15.00),
    ]

    for g in gastos_dj:
        for i, val in enumerate(g, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i == 4))
        row += 1

    # Total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="TOTAL S/").font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell = ws.cell(row=row, column=4, value=70.00)
    style_data_cell(cell, is_money=True)
    cell.font = Font(bold=True, size=10)

    auto_width(ws)
    return ws


# ============================================================
# HOJA 3: COMPROBANTES DE PAGO — Detalle tipo Registro de Compras SUNAT
# ============================================================
def crear_hoja_comprobantes(wb):
    ws = wb.create_sheet("Comprobantes")

    ws.merge_cells("A1:T1")
    ws["A1"] = "REGISTRO DE COMPROBANTES DE PAGO — DETALLE TIPO SUNAT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Info general del expediente
    ws.merge_cells("A2:T2")
    ws["A2"] = (
        "Expediente: ODI2026-INT-0139051 | Comisionado: QUEREVALU MIÑAN JOSE FAUSTINO | DNI: 41969732"
    )
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = [
        "N°",  # 1
        "Fecha Emisión",  # 2
        "Tipo Comprobante",  # 3
        "Comprobante Electrónico",  # 4
        "Serie-Número",  # 5
        "RUC Proveedor",  # 6
        "Razón Social Proveedor",  # 7
        "Dirección Proveedor",  # 8
        "Cliente (Señor/es)",  # 9
        "RUC Cliente",  # 10
        "Dirección Cliente",  # 11
        "Concepto / Qué consumió",  # 12
        "Detalle Ítems",  # 13
        "Forma de Pago",  # 14
        "Valor Venta (Base Imponible)",  # 15
        "IGV S/",  # 16
        "% IGV Aplicado",  # 17
        "ICBPER S/",  # 18
        "Importe Total S/",  # 19
        "Observaciones",  # 20
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    # =================================================================
    # COMPROBANTE 1: E001-11614 — CONTADORES PUBLICOS EIRL (Almuerzo día 1)
    # =================================================================
    comprobantes = [
        {
            "n": 1,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-11614",
            "ruc_prov": "20102883391",
            "razon_social_prov": "CONTADORES PUBLICOS EIRL (CHALAN DEL NORTE DESDE 1976)",
            "dir_prov": "JR. ICA 649 CENT. PIURA A 2 PUERTAS DE LA CURACAO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (Almuerzo)",
            "detalle_items": "1 DUO MARINO (S/50.85) + 1 JARRA DE MARACUYA (S/8.47)",
            "forma_pago": "Contado",
            "valor_venta": 59.32,
            "igv": 10.68,
            "pct_igv": "18%",
            "icbper": 0.00,
            "importe_total": 70.00,
            "obs": "Hora impresión: 15:12. Factura electrónica SUNAT.",
        },
        # COMPROBANTE 2: E001-11611
        {
            "n": 2,
            "fecha": "04/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-11611",
            "ruc_prov": "20102883391",
            "razon_social_prov": "CONTADORES PUBLICOS EIRL (CHALAN DEL NORTE DESDE 1976)",
            "dir_prov": "JR. ICA 649 CENT. PIURA A 2 PUERTAS DE LA CURACAO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (Desayuno/Refrigerio)",
            "detalle_items": "2 PANES CON CHICHARRON (S/16.95 c/u) + 1 JUGO DE PAPAYA (S/8.47)",
            "forma_pago": "Contado",
            "valor_venta": 42.37,
            "igv": 7.63,
            "pct_igv": "18%",
            "icbper": 0.00,
            "importe_total": 50.00,
            "obs": "Hora impresión: 14:56. Factura electrónica SUNAT.",
        },
        # COMPROBANTE 3: E001-1134 — HOSPEDAJE
        {
            "n": 3,
            "fecha": "04/02/2026 - 05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-1134",
            "ruc_prov": "10028235173",
            "razon_social_prov": "MANRIQUE RIOS JULIO NICANOR (HOSPEDAJE MOON NIGHT)",
            "dir_prov": "CAL. JUNIN 899 A 2 CASAS DE DIARIO EL TIEMPO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "HOSPEDAJE (2 noches: 04 y 05 de febrero 2026)",
            "detalle_items": "2 UNID. Alojamiento del día 04 y 05 de febrero 2026 (S/136.363 c/u)",
            "forma_pago": "Contado",
            "valor_venta": 272.73,
            "igv": 27.27,
            "pct_igv": "10%",
            "icbper": 0.00,
            "importe_total": 300.00,
            "obs": "Fecha emisión: 05/02/2026. Hospedaje 2 días. IGV 10% CORRECTO — Aplica Ley 31556 + Ley 32219 (MYPE restaurantes/hoteles/alojamientos, vigente 2025-2026: IGV 8% + IPM 2% = 10% total).",
        },
        # COMPROBANTE 4: E001-11622
        {
            "n": 4,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-11622",
            "ruc_prov": "20102883391",
            "razon_social_prov": "CONTADORES PUBLICOS EIRL (CHALAN DEL NORTE DESDE 1976)",
            "dir_prov": "JR. ICA 649 CENT. PIURA A 2 PUERTAS DE LA CURACAO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (Desayuno día 2)",
            "detalle_items": "2 PANES CON POLLO (S/12.71 c/u) + 1 JUGO SURTIDO (S/8.475)",
            "forma_pago": "Contado",
            "valor_venta": 33.90,
            "igv": 6.10,
            "pct_igv": "18%",
            "icbper": 0.00,
            "importe_total": 40.00,
            "obs": "Hora impresión: 11:07. Factura electrónica SUNAT.",
        },
        # COMPROBANTE 5: E001-11629
        {
            "n": 5,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-11629",
            "ruc_prov": "20102883391",
            "razon_social_prov": "CONTADORES PUBLICOS EIRL (CHALAN DEL NORTE DESDE 1976)",
            "dir_prov": "JR. ICA 649 CENT. PIURA A 2 PUERTAS DE LA CURACAO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (Almuerzo día 2)",
            "detalle_items": "1 CEVICHE (S/12.71) + 1 LOMO A LO POBRE (S/33.90) + 1 GASEOSA (S/4.24)",
            "forma_pago": "Contado",
            "valor_venta": 50.85,
            "igv": 9.15,
            "pct_igv": "18%",
            "icbper": 0.00,
            "importe_total": 60.00,
            "obs": "Hora impresión: 14:41. Factura electrónica SUNAT.",
        },
        # COMPROBANTE 6: E001-11630
        {
            "n": 6,
            "fecha": "05/02/2026",
            "tipo_cp": "FACTURA",
            "electronico": "SÍ",
            "serie_num": "E001-11630",
            "ruc_prov": "20102883391",
            "razon_social_prov": "CONTADORES PUBLICOS EIRL (CHALAN DEL NORTE DESDE 1976)",
            "dir_prov": "JR. ICA 649 CENT. PIURA A 2 PUERTAS DE LA CURACAO, PIURA - PIURA - PIURA",
            "cliente": "MINISTERIO DE EDUCACION",
            "ruc_cliente": "20131370998",
            "dir_cliente": "CAL. DEL COMERCIO 193 LIMA - LIMA - SAN BORJA",
            "concepto": "ALIMENTACION (Cena/Refrigerio día 2)",
            "detalle_items": "1 PAVO CON TALLARINES Y CHIFLES (S/38.14) + 1 GASEOSA (S/4.235)",
            "forma_pago": "Contado",
            "valor_venta": 42.38,
            "igv": 7.63,
            "pct_igv": "18%",
            "icbper": 0.00,
            "importe_total": 50.00,
            "obs": "Hora impresión: 17:04. Factura electrónica SUNAT.",
        },
    ]

    for cp in comprobantes:
        values = [
            cp["n"],
            cp["fecha"],
            cp["tipo_cp"],
            cp["electronico"],
            cp["serie_num"],
            cp["ruc_prov"],
            cp["razon_social_prov"],
            cp["dir_prov"],
            cp["cliente"],
            cp["ruc_cliente"],
            cp["dir_cliente"],
            cp["concepto"],
            cp["detalle_items"],
            cp["forma_pago"],
            cp["valor_venta"],
            cp["igv"],
            cp["pct_igv"],
            cp["icbper"],
            cp["importe_total"],
            cp["obs"],
        ]
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i in (15, 16, 18, 19)
            style_data_cell(cell, is_money=is_money)
        row += 1

    # Fila de totales
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="TOTALES COMPROBANTES CON SUSTENTO").font = Font(
        bold=True, size=10
    )
    ws.cell(row=row, column=1).border = THIN_BORDER

    total_vv = sum(c["valor_venta"] for c in comprobantes)
    total_igv = sum(c["igv"] for c in comprobantes)
    total_icbper = sum(c["icbper"] for c in comprobantes)
    total_importe = sum(c["importe_total"] for c in comprobantes)

    for col, val in [(15, total_vv), (16, total_igv), (18, total_icbper), (19, total_importe)]:
        cell = ws.cell(row=row, column=col, value=val)
        style_data_cell(cell, is_money=True)
        cell.font = Font(bold=True, size=10)

    # Verificación aritmética
    row += 2
    ws.cell(row=row, column=1, value="VERIFICACIÓN ARITMÉTICA IGV").font = SUBTITLE_FONT
    row += 1
    headers_v = [
        "Comprobante",
        "Base Imponible",
        "IGV Factura",
        "% Real",
        "Régimen Aplicable",
        "% Esperado",
        "IGV Esperado",
        "Diferencia",
        "ESTADO",
    ]
    for i, h in enumerate(headers_v, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_v))
    row += 1

    # Regímenes de IGV vigentes 2025-2026
    # - General: 18% (IGV 16% + IPM 2%)
    # - MYPE Restaurantes/Hoteles/Alojamientos: 10% (IGV 8% + IPM 2%) — Ley 31556 + Ley 32219
    # - Amazonía: 0% (Ley 27037)

    for cp in comprobantes:
        pct_real = round((cp["igv"] / cp["valor_venta"]) * 100, 2) if cp["valor_venta"] > 0 else 0

        # Determinar régimen esperado basado en tipo de proveedor
        es_mype_restaurante_hotel = cp["concepto"].startswith("HOSPEDAJE") or cp[
            "concepto"
        ].startswith("ALIMENTACION")
        # MYPE restaurantes/hoteles → 10% (Ley 31556 + 32219)
        if es_mype_restaurante_hotel and abs(pct_real - 10.0) < 1.0:
            regimen = "MYPE Rest/Hotel (Ley 31556+32219)"
            pct_esperado = 10.0
        else:
            regimen = "General"
            pct_esperado = 18.0

        igv_esperado = round(cp["valor_venta"] * (pct_esperado / 100), 2)
        diff = round(cp["igv"] - igv_esperado, 2)

        estado = ""
        if abs(pct_real - pct_esperado) < 0.5:
            estado = "OK"
        elif abs(pct_real - 10.0) < 0.5 and es_mype_restaurante_hotel:
            estado = "OK (MYPE 10%)"
        else:
            estado = f"REVISAR — {pct_real}% no coincide con {pct_esperado}%"

        vals = [
            cp["serie_num"],
            cp["valor_venta"],
            cp["igv"],
            f"{pct_real}%",
            regimen,
            f"{pct_esperado}%",
            igv_esperado,
            diff,
            estado,
        ]
        for i, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell, is_money=(i in (2, 3, 7, 8)))
            if "REVISAR" in str(val):
                cell.font = Font(bold=True, color="FF0000", size=10)
            elif val == "OK" or val == "OK (MYPE 10%)":
                cell.font = Font(bold=True, color="008000", size=10)
        row += 1

    auto_width(ws, max_width=50)
    return ws


# ============================================================
# HOJA 4: BOARDING PASS + TIQUETE AÉREO
# ============================================================
def crear_hoja_boarding(wb):
    ws = wb.create_sheet("BoardingPass")

    ws.merge_cells("A1:H1")
    ws["A1"] = "BOARDING PASS / TARJETA DE EMBARQUE + TIQUETE AÉREO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Info del pasajero
    info = [
        ("Pasajero", "QUEREVALU MINAN JOSE FAUSTINO"),
        ("DNI", "41969732"),
        ("Tipo Pasajero", "Adulto"),
        ("Aerolínea", "LATAM AIRLINES PERU"),
        ("Código de Reserva", "EKKBPG (Boarding) / DOJYHH (Tiquete)"),
        ("N° de Orden", "LA5448833SRSP"),
        ("N° de Ticket", "0452272096281 / 5442272045356"),
        ("Frequent Flyer", "LA 51419697323 — GLD (Gold)"),
        ("Emisión Tiquete", "Lima, Perú 03/02/2026"),
        ("RUC Aerolínea", "20341841357 (Latam Airlines Perú S.A.)"),
        ("Dirección Aerolínea", "Av. Santa Cruz 381 Piso 6, Miraflores, Lima, Perú"),
        ("Facturación tributaria a", "MINISTERIO DE EDUCACION — RUC 20131370998"),
    ]

    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Tabla de vuelos
    row += 1
    ws.cell(row=row, column=1, value="DETALLE DE VUELOS").font = SUBTITLE_FONT
    row += 1

    headers = [
        "Tramo",
        "N° Vuelo",
        "Fecha",
        "Origen",
        "Destino",
        "Hora Salida",
        "Hora Llegada",
        "Puerta",
        "Asiento",
        "Cabina",
        "Tarifa",
        "Hora Presentarse",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    vuelos = [
        (
            "IDA",
            "LA 2302",
            "04/02/2026",
            "LIMA (J. Chavez Intl.)",
            "PIURA (G. Concha Ibérico)",
            "05:35",
            "07:15",
            "C17",
            "8C (LIM)",
            "Economy",
            "Basic",
            "04:35",
        ),
        (
            "RETORNO",
            "LA 2126",
            "05/02/2026",
            "PIURA (G. Concha Ibérico)",
            "LIMA (J. Chavez Intl.)",
            "19:35",
            "21:00",
            "1",
            "8C (PIU)",
            "Economy",
            "Basic",
            "18:35",
        ),
    ]

    for v in vuelos:
        for i, val in enumerate(v, 1):
            cell = ws.cell(row=row, column=i, value=val)
            style_data_cell(cell)
        row += 1

    # Desglose de pago del tiquete
    row += 2
    ws.cell(row=row, column=1, value="DESGLOSE DE PAGO DEL TIQUETE").font = SUBTITLE_FONT
    row += 1

    headers_p = ["Concepto", "Moneda", "Monto"]
    for i, h in enumerate(headers_p, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers_p))
    row += 1

    pagos = [
        ("Vuelo (tarifa)", "USD", 323.00),
        ("Tasas y/o impuestos", "USD", 77.82),
        ("   — PE (impuestos Perú)", "USD", 58.14),
        ("TOTAL TIQUETE", "USD", 400.82),
        ("Forma de pago", "LATAM Wallet", "USD 400.82"),
    ]

    for p in pagos:
        for i, val in enumerate(p, 1):
            cell = ws.cell(row=row, column=i, value=val)
            is_money = i == 3 and isinstance(val, (int, float))
            style_data_cell(cell, is_money=is_money)
            if p[0] == "TOTAL TIQUETE":
                cell.font = Font(bold=True, size=10)
        row += 1

    auto_width(ws)
    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()

    print("Creando Hoja 1: Anexo 3...")
    crear_hoja_anexo3(wb)

    print("Creando Hoja 2: Declaración Jurada...")
    crear_hoja_dj(wb)

    print("Creando Hoja 3: Comprobantes (detalle SUNAT)...")
    crear_hoja_comprobantes(wb)

    print("Creando Hoja 4: Boarding Pass + Tiquete Aéreo...")
    crear_hoja_boarding(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Excel generado: {OUTPUT_FILE}")
    print(f"   Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
