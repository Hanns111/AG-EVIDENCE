#!/usr/bin/env python3
"""
Genera Excel de rendicion para expediente OTIC2026-INT-0072834
MERA CASAS JULIO IVAN - Canete - 18-21/01/2026

4 Hojas:
1. Anexo 3 - Rendicion de Cuentas
2. DJ - Anexo 4 (Declaracion Jurada)
3. Comprobantes de Pago (detalle SUNAT)
4. Boletos y Boarding Pass
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ============================================================
# ESTILOS COMUNES
# ============================================================
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=10, color='FFFFFF')
data_font = Font(size=9)
money_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment


def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)


# ============================================================
# INFO GENERAL DEL EXPEDIENTE
# ============================================================
exp_info = {
    'expediente': 'OTIC2026-INT-0072834',
    'comisionado': 'MERA CASAS JULIO IVAN',
    'dni': '09620528',
    'dependencia': 'OTIC - OFICINA DE TECNOLOGIAS DE LA INFORMACION Y COMUNICACION',
    'cargo': 'PERSONAL CAS - ASISTENCIA TECNICO-OPERATIVO',
    'destino': 'CANETE - SAN VICENTE DE CANETE (UGEL 08 CANETE)',
    'fecha_salida': '18/01/2026',
    'fecha_regreso': '21/01/2026',
    'dias_horas': '3d 3h',
    'monto_recibido': 1200.00,
    'planilla': '00030',
    'exp_siaf': '2600400',
    'nro_comprobante': '000297',
}

# ============================================================
# HOJA 1: ANEXO 3 - RENDICION DE CUENTAS
# ============================================================
ws1 = wb.active
ws1.title = 'Anexo 3 - Rendicion'

# Titulo
ws1.merge_cells('A1:G1')
ws1['A1'] = 'ANEXO N 3 - RENDICION DE CUENTAS POR COMISION DE SERVICIOS'
ws1['A1'].font = Font(bold=True, size=12)
ws1['A1'].alignment = Alignment(horizontal='center')

# Info del comisionado
info_rows = [
    ('Expediente:', exp_info['expediente']),
    ('Comisionado:', exp_info['comisionado']),
    ('DNI:', exp_info['dni']),
    ('Dependencia:', exp_info['dependencia']),
    ('Destino:', exp_info['destino']),
    ('Salida:', exp_info['fecha_salida']),
    ('Regreso:', exp_info['fecha_regreso']),
    ('Dias/Horas:', exp_info['dias_horas']),
    ('Planilla:', exp_info['planilla']),
    ('Exp SIAF:', exp_info['exp_siaf']),
]
for i, (label, val) in enumerate(info_rows, start=3):
    ws1.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
    ws1.cell(row=i, column=2, value=val).font = data_font

# Tabla de gastos
header_row = 14
headers = ['N', 'FECHA', 'TIPO DOC', 'RAZON SOCIAL', 'NUMERO', 'CONCEPTO', 'IMPORTE S/']
for col, h in enumerate(headers, 1):
    ws1.cell(row=header_row, column=col, value=h)
style_header(ws1, header_row, len(headers))

# 19 comprobantes del Anexo 3
anexo3_data = [
    (1, '18/01/2026', 'Factura', 'VIERA HUAPAYA ARMANDO FRITZ', 'E001-1335', 'MOVILIDAD', 50.00),
    (2, '18/01/2026', 'Factura', 'FLORES SIANCAS VDA. DE GUEVARA RUTH CONSUELO', 'E001-449', 'HOSPEDAJE', 90.00),
    (3, '18/01/2026', 'Factura', 'RODRIGUEZ AYALA HECTOR', 'FF02-00000107', 'ALIMENTACION', 22.50),
    (4, '18/01/2026', 'Factura', 'EMPRESA DE TRANSPORTES PERU BUS SA', 'F26M-00020463', 'PASAJE', 25.00),
    (5, '19/01/2026', 'Factura', 'INVERSIONES FRUTALINA S.A.C', 'FZ01-3485', 'ALIMENTACION', 28.00),
    (6, '19/01/2026', 'Factura', 'LOPEZ CAMPOS DAIRY TAHIRI', 'E001-5438', 'ALIMENTACION', 60.00),
    (7, '19/01/2026', 'Factura', 'CASASMARKET S.A.C.', 'FF01-00001284', 'ALIMENTACION', 9.50),
    (8, '19/01/2026', 'Factura', 'CAM MEJIA JIMMY CESAR', 'F002-00000394', 'ALIMENTACION', 57.00),
    (9, '19/01/2026', 'Factura', 'FLORES SIANCAS VDA. DE GUEVARA RUTH CONSUELO', 'E001-0449', 'HOSPEDAJE', 90.00),
    (10, '20/01/2026', 'Factura', 'INVERSIONES FRUTALINA S.A.C', 'FZ01-3494', 'ALIMENTACION', 33.00),
    (11, '20/01/2026', 'Factura', 'CAM MEJIA JIMMY CESAR', 'F002-00000395', 'ALIMENTACION', 53.00),
    (12, '20/01/2026', 'Factura', 'PIRIS', 'F001-881', 'ALIMENTACION', 41.00),
    (13, '20/01/2026', 'Factura', 'FLORES SIANCAS VDA. DE GUEVARA RUTH CONSUELO', 'E001-00449', 'HOSPEDAJE', 90.00),
    (14, '21/01/2026', 'Factura', 'INVERSIONES FRUTALINA S.A.C', 'FZ01-3504', 'ALIMENTACION', 20.00),
    (15, '21/01/2026', 'Factura', 'LOPEZ CAMPOS DAIRY TAHIRI', 'E001-5447', 'ALIMENTACION', 70.00),
    (16, '21/01/2026', 'Factura', 'RODRIGUEZ AYALA HECTOR', 'FF02-00000116', 'ALIMENTACION', 20.00),
    (17, '21/01/2026', 'Factura', 'FLORES SIANCAS VDA. DE GUEVARA RUTH CONSUELO', 'E001-000449', 'HOSPEDAJE', 90.00),
    (18, '21/01/2026', 'Factura', 'TAXI RUMI EXPRESS S.A.C.', 'FE01-00001043', 'MOVILIDAD', 60.00),
    (19, '21/01/2026', 'Factura', 'EMPRESA DE TRANSPORTES PERU BUS SA', 'F54L-00012299', 'PASAJE', 15.00),
]

for i, row_data in enumerate(anexo3_data, start=header_row + 1):
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        if col == 7:
            cell.number_format = money_format
    style_data(ws1, i, len(headers))

# Totales
total_row = header_row + 1 + len(anexo3_data)
totals = [
    ('(1) GASTOS CON DOCUMENTACION SUSTENTATORIA', 924.00),
    ('(2) GASTOS SIN DOCUMENTACION SUSTENTATORIA (DJ)', 98.00),
    ('(3) TOTAL GASTADO (1 + 2)', 1022.00),
    ('(4) DEVOLUCION', 178.00),
    ('(5) MONTO RECIBIDO (3 + 4)', 1200.00),
]
for j, (label, amount) in enumerate(totals):
    r = total_row + j
    ws1.merge_cells(f'A{r}:F{r}')
    ws1.cell(row=r, column=1, value=label).font = Font(bold=True, size=9)
    ws1.cell(row=r, column=7, value=amount).number_format = money_format
    ws1.cell(row=r, column=7).font = Font(bold=True, size=9)
    for c in range(1, 8):
        ws1.cell(row=r, column=c).border = thin_border

# Anchos
for col, w in zip('ABCDEFG', [4, 12, 10, 42, 18, 16, 14]):
    ws1.column_dimensions[col].width = w

print('Hoja 1 (Anexo 3) completada - 19 items')

# ============================================================
# HOJA 2: DECLARACION JURADA (DJ / ANEXO 4)
# ============================================================
ws2 = wb.create_sheet('DJ - Anexo 4')

ws2.merge_cells('A1:E1')
ws2['A1'] = 'ANEXO N 4 - DECLARACION JURADA DE GASTOS SIN COMPROBANTE'
ws2['A1'].font = Font(bold=True, size=12)
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.cell(row=3, column=1, value='Comisionado:').font = Font(bold=True, size=9)
ws2.cell(row=3, column=2, value=exp_info['comisionado']).font = data_font
ws2.cell(row=4, column=1, value='DNI:').font = Font(bold=True, size=9)
ws2.cell(row=4, column=2, value=exp_info['dni']).font = data_font
ws2.cell(row=5, column=1, value='Domicilio:').font = Font(bold=True, size=9)
ws2.cell(row=5, column=2, value='PJE. LOS CARDOS 153 - URB. MICAELA BASTIDAS').font = data_font

header_row2 = 7
headers2 = ['N', 'FECHA', 'CONCEPTO', 'DETALLE', 'IMPORTE S/']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=header_row2, column=col, value=h)
style_header(ws2, header_row2, len(headers2))

dj_data = [
    (1, '18/01/2026', 'MOVILIDAD', 'Movilidad terminal Peru Bus - Hotel', 5.00),
    (2, '18/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (3, '18/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Hotel', 4.00),
    (4, '19/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (5, '19/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (6, '19/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Restaurant', 4.00),
    (7, '19/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (8, '19/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Hotel', 4.00),
    (9, '19/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (10, '19/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Hotel', 4.00),
    (11, '20/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (12, '20/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (13, '20/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Restaurant', 4.00),
    (14, '20/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (15, '20/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Hotel', 4.00),
    (16, '20/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (17, '20/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Hotel', 4.00),
    (18, '21/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (19, '21/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (20, '21/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Restaurant', 4.00),
    (21, '21/01/2026', 'MOVILIDAD', 'Movilidad Restaurant - Ugel Canete', 4.00),
    (22, '21/01/2026', 'MOVILIDAD', 'Movilidad Ugel Canete - Hotel', 4.00),
    (23, '21/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Restaurant', 4.00),
    (24, '21/01/2026', 'MOVILIDAD', 'Movilidad Hotel - Terminal Terrestre Peru Bus', 5.00),
]

for i, row_data in enumerate(dj_data, start=header_row2 + 1):
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        if col == 5:
            cell.number_format = money_format
    style_data(ws2, i, len(headers2))

total_row2 = header_row2 + 1 + len(dj_data)
ws2.merge_cells(f'A{total_row2}:D{total_row2}')
ws2.cell(row=total_row2, column=1, value='TOTAL S/').font = Font(bold=True, size=9)
ws2.cell(row=total_row2, column=5, value=98.00).number_format = money_format
ws2.cell(row=total_row2, column=5).font = Font(bold=True, size=9)
for c in range(1, 6):
    ws2.cell(row=total_row2, column=c).border = thin_border

for col, w in zip('ABCDE', [4, 12, 14, 50, 14]):
    ws2.column_dimensions[col].width = w

print('Hoja 2 (DJ) completada - 24 items')

# ============================================================
# HOJA 3: COMPROBANTES DE PAGO (datos SUNAT)
# ============================================================
ws3 = wb.create_sheet('Comprobantes de Pago')

ws3.merge_cells('A1:T1')
ws3['A1'] = 'DETALLE DE COMPROBANTES DE PAGO - EXPEDIENTE OTIC2026-INT-0072834'
ws3['A1'].font = Font(bold=True, size=11)
ws3['A1'].alignment = Alignment(horizontal='center')

header_row3 = 3
headers3 = [
    'N', 'FECHA EMISION', 'TIPO DOC', 'SERIE-NUMERO',
    'RUC EMISOR', 'RAZON SOCIAL EMISOR', 'DIRECCION EMISOR',
    'CONCEPTO', 'DESCRIPCION ITEMS',
    'OP. GRAVADAS', 'OP. EXONERADAS', 'OP. INAFECTAS', 'OP. GRATUITAS',
    'BASE IMPONIBLE', 'IGV', 'TASA IGV', 'ICBPER', 'TOTAL',
    'FORMA PAGO', 'OBSERVACIONES'
]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=header_row3, column=col, value=h)
style_header(ws3, header_row3, len(headers3))

# 16 facturas fisicas reales (hospedaje es 1 sola factura por 4 noches)
comprobantes = [
    (1, '18/01/2026', 'Factura Electronica', 'E001-1335',
     '10097112393', 'VIERA HUAPAYA ARMANDO FRITZ',
     'AV. FAUSTINO SILVA 319 P.J. CIUDAD DE DIOS ZONA A, SAN JUAN DE MIRAFLORES - LIMA',
     'MOVILIDAD', 'LOS OLIVOS - LA VICTORIA (EMP.SOYUS)',
     50.00, 0.00, 0.00, 0.00, 50.00, 0.00, '0%', 0.00, 50.00,
     'Contado', 'Servicio de transporte'),

    (2, '21/01/2026', 'Factura Electronica', 'E001-449',
     '10153497394', 'FLORES SIANCAS VDA. DE GUEVARA RUTH CONSUELO',
     'HOSTAL-RESTAURANT-COLIBRI, MZ A URB. LOS LIBERTADORES INT LT 9, SAN VICENTE DE CANETE',
     'HOSPEDAJE', 'Alojamiento 18/01 al 21/01/2026 - Sr. Julio Mera Casa (4 noches x S/81.82+IGV)',
     327.27, 0.00, 0.00, 0.00, 327.27, 32.73, '10%', 0.00, 360.00,
     'Contado', 'MYPE IGV 10%. Anexo3 registra 4 lineas S/90 c/u = S/360'),

    (3, '18/01/2026', 'Factura Electronica', 'FF02-00000107',
     '10411107529', 'RODRIGUEZ AYALA HECTOR (Marlys)',
     'JR. 2 DE MAYO No 392, SAN VICENTE DE CANETE - CANETE - LIMA',
     'ALIMENTACION', '1/4 Pollo Brasa S/18 + Gaseosa 1/2 LT S/4.50',
     19.07, 0.00, 0.00, 0.00, 19.07, 3.43, '18%', 0.00, 22.50,
     'Contado', ''),

    (4, '18/01/2026', 'Factura Electronica', 'F26M-00020463',
     '20350227630', 'EMPRESA DE TRANSPORTES PERU BUS S.A.',
     'CAR. PANAMERICANA SUR HR. 142.5 VIA LIMA-ICA',
     'PASAJE', 'SERV TRANS RUTA Lima - Canete (Embarque 18/01 16:20)',
     0.00, 25.00, 0.00, 0.00, 0.00, 0.00, 'EXONERADO', 0.00, 25.00,
     'Contado', 'Boleto IDA. Transporte interprovincial exonerado IGV'),

    (5, '19/01/2026', 'Factura Electronica', 'FZ01-3485',
     '20612547310', 'INVERSIONES FRUTALINA S.A.C.',
     'JR. SEPULVEDA NRO. 231 URB. SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Extracto Surtido S/10 + Cafe Combo S/1 + Combo Salchicha S/17',
     25.45, 0.00, 0.00, 0.00, 25.45, 2.55, '10%', 0.00, 28.00,
     'Contado', 'MYPE IGV 10%'),

    (6, '19/01/2026', 'Factura Electronica', 'E001-5438',
     '10465399669', 'LOPEZ CAMPOS DAIRY TAHIRI',
     'URB. SANTA ROSA DE HUALCA MZA. C LOTE 19, SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Chifa Frita + Porcion de Arroz + Gaseosa Personal',
     60.00, 0.00, 0.00, 0.00, 60.00, 0.00, '0%', 0.00, 60.00,
     'Contado', 'Sin desglose IGV visible en factura'),

    (7, '19/01/2026', 'Factura Electronica', 'FF01-00001284',
     '20604417538', 'COMPANIA TOVAR E.I.R.L. (ECOMARKET)',
     'JR. 2 DE MAYO NRO 530, SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Papas nativas S/2 + Leche fresca S/2.50 + Agua san luis S/3 + Inca kola S/2',
     8.05, 0.00, 0.00, 0.00, 8.05, 1.45, '18%', 0.00, 9.50,
     'Contado', ''),

    (8, '19/01/2026', 'Factura Electronica', 'F002-00000394',
     '10421686331', 'CAM MEJIA JIMMY CESAR (CHIFA KIMFUN)',
     'JR. SAN AGUSTIN NRO 219-A SAN VICENTE - CANETE - LIMA',
     'ALIMENTACION', 'Taypa Especial S/40 + Sopa Sustancia S/17',
     51.81, 0.00, 0.00, 0.00, 51.81, 5.19, '10%', 0.00, 57.00,
     'Contado', 'MYPE IGV 10%'),

    (9, '20/01/2026', 'Factura Electronica', 'FZ01-3494',
     '20612547310', 'INVERSIONES FRUTALINA S.A.C.',
     'JR. SEPULVEDA NRO. 231 URB. SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Combo Tradiciona S/23 + Pina+Melon S/10',
     30.00, 0.00, 0.00, 0.00, 30.00, 3.00, '10%', 0.00, 33.00,
     'Contado', 'MYPE IGV 10%'),

    (10, '20/01/2026', 'Factura Electronica', 'F002-00000395',
     '10421686331', 'CAM MEJIA JIMMY CESAR (CHIFA KIMFUN)',
     'JR. SAN AGUSTIN NRO 219-A SAN VICENTE - CANETE - LIMA',
     'ALIMENTACION', '1/2 Inca Kola S/5 + Sopa Woaming S/18 + Pollo Saltado Hongo Chino S/30',
     48.18, 0.00, 0.00, 0.00, 48.18, 4.82, '10%', 0.00, 53.00,
     'Contado', 'MYPE IGV 10%'),

    (11, '20/01/2026', 'Factura Electronica', 'F001-881',
     '10515561217', 'FERNANDEZ PULCHS MANUEL GREGORIO (PIRIS)',
     'JR. SEPULVEDA Nro. 381, SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Ceviche Pescado S/37 + Inca Kola 600ml S/4',
     37.28, 0.00, 0.00, 0.00, 37.28, 3.72, '10%', 0.00, 41.00,
     'Contado', 'MYPE IGV 10%'),

    (12, '21/01/2026', 'Factura Electronica', 'FZ01-3504',
     '20612547310', 'INVERSIONES FRUTALINA S.A.C.',
     'JR. SEPULVEDA NRO. 231 URB. SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Pina+Melon S/10 + Hamburguesa Clasi S/10',
     18.18, 0.00, 0.00, 0.00, 18.18, 1.82, '10%', 0.00, 20.00,
     'Contado', 'MYPE IGV 10%'),

    (13, '21/01/2026', 'Factura Electronica', 'E001-5447',
     '10465398869', 'LOPEZ CAMPOS DAIRY TAHIRI',
     'URB. SANTA ROSA DE HUALCA MZA. C LOTE 19, SAN VICENTE DE CANETE',
     'ALIMENTACION', 'Causa Rellena + Arroz Chaufa de Mariscos',
     70.00, 0.00, 0.00, 0.00, 70.00, 0.00, '0%', 0.00, 70.00,
     'Contado', 'Sin desglose IGV visible en factura'),

    (14, '21/01/2026', 'Factura Electronica', 'FF02-00000116',
     '10411107529', 'RODRIGUEZ AYALA HECTOR (Marlys)',
     'JR. 2 DE MAYO No 392, SAN VICENTE DE CANETE',
     'ALIMENTACION', '1/4 Pollo Brasa S/17 + Gaseosa Mediana S/3',
     16.95, 0.00, 0.00, 0.00, 16.95, 3.05, '18%', 0.00, 20.00,
     'Contado', ''),

    (15, '21/01/2026', 'Factura Electronica', 'FE01-00001043',
     '20605749381', 'TAXI RUMI EXPRESS S.A.C.',
     'MZA. F LOTE 01 PROV VISTA ALEGRE IV Q2, CALLAO',
     'MOVILIDAD', 'Taxi agencia Peru Bus hacia Los Olivos',
     0.00, 60.00, 0.00, 0.00, 0.00, 0.00, 'EXONERADO', 0.00, 60.00,
     'Contado', 'Op. Exonerada'),

    (16, '21/01/2026', 'Factura Electronica', 'F54L-00012299',
     '20350227630', 'EMPRESA DE TRANSPORTES PERU BUS S.A.',
     'CAR. PANAMERICANA SUR HR. 142.5 VIA LIMA-ICA',
     'PASAJE', 'SERV TRANS RUTA Canete - Lima (Embarque 21/01 18:01)',
     0.00, 15.00, 0.00, 0.00, 0.00, 0.00, 'EXONERADO', 0.00, 15.00,
     'Contado', 'Boleto RETORNO. Transporte interprovincial exonerado IGV'),
]

for i, row_data in enumerate(comprobantes, start=header_row3 + 1):
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        if col in (10, 11, 12, 13, 14, 15, 17, 18):
            cell.number_format = money_format
    style_data(ws3, i, len(headers3))

# Nota sobre hospedaje
nota_row = header_row3 + 1 + len(comprobantes) + 1
ws3.merge_cells(f'A{nota_row}:T{nota_row}')
ws3.cell(row=nota_row, column=1,
         value='NOTA: El hospedaje (FLORES SIANCAS) aparece como 4 lineas de S/90 en Anexo 3 '
               '(una por noche), pero fisicamente es UNA SOLA factura E001-449 por S/360 '
               '(4 noches). Aqui se registra la factura real unica.').font = Font(italic=True, size=9, color='FF0000')

# Totales
total_row3 = nota_row + 1
ws3.merge_cells(f'A{total_row3}:Q{total_row3}')
ws3.cell(row=total_row3, column=1, value='TOTAL COMPROBANTES (16 facturas fisicas)').font = Font(bold=True, size=10)
total_sum = sum(c[17] for c in comprobantes)  # index 17 = TOTAL (0-based)
ws3.cell(row=total_row3, column=18, value=total_sum).number_format = money_format
ws3.cell(row=total_row3, column=18).font = Font(bold=True, size=10)

# Resumen IGV
igv_row = total_row3 + 2
ws3.merge_cells(f'A{igv_row}:T{igv_row}')
ws3.cell(row=igv_row, column=1, value='RESUMEN DE TASAS IGV DETECTADAS:').font = Font(bold=True, size=10)

igv_items = [
    'IGV 18% (regimen general): Comprobantes #3, #7, #14 (Rodriguez Ayala, Ecomarket)',
    'IGV 10% (MYPE restaurante/hotel): Comprobantes #2, #5, #8, #9, #10, #11, #12 (Flores Siancas, Frutalina, Cam Mejia, Piris)',
    'IGV 0% / Sin desglose: Comprobantes #1, #6, #13 (Viera Huapaya, Lopez Campos)',
    'EXONERADO (transporte interprovincial): Comprobantes #4, #15, #16 (Peru Bus, Taxi Rumi)',
]
for j, item in enumerate(igv_items):
    r = igv_row + 1 + j
    ws3.merge_cells(f'A{r}:T{r}')
    ws3.cell(row=r, column=1, value=f'  - {item}').font = Font(size=9)

# Anchos
col_widths = [4, 12, 18, 18, 14, 38, 48, 14, 48, 12, 12, 12, 12, 12, 10, 10, 8, 12, 10, 45]
for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

print(f'Hoja 3 (Comprobantes) completada - {len(comprobantes)} facturas, Total: S/{total_sum:.2f}')

# ============================================================
# HOJA 4: RESUMEN DE BOLETOS Y BOARDING PASS
# ============================================================
ws4 = wb.create_sheet('Boletos y Boarding Pass')

ws4.merge_cells('A1:L1')
ws4['A1'] = 'RESUMEN DE BOLETOS DE VIAJE Y BOARDING PASS'
ws4['A1'].font = Font(bold=True, size=12)
ws4['A1'].alignment = Alignment(horizontal='center')

info4 = [
    ('Expediente:', exp_info['expediente']),
    ('Comisionado:', exp_info['comisionado']),
    ('Ruta Comision:', 'LIMA - CANETE (SAN VICENTE DE CANETE) - LIMA'),
    ('Tipo transporte:', 'TERRESTRE (Bus interprovincial)'),
]
for i, (label, val) in enumerate(info4, start=3):
    ws4.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
    ws4.cell(row=i, column=2, value=val).font = data_font

header_row4 = 8
headers4 = [
    'N', 'TIPO', 'EMPRESA', 'RUC', 'SERIE-NUMERO',
    'FECHA EMISION', 'FECHA EMBARQUE', 'HORA EMBARQUE',
    'RUTA', 'PASAJERO', 'DNI', 'IMPORTE S/'
]
for col, h in enumerate(headers4, 1):
    ws4.cell(row=header_row4, column=col, value=h)
style_header(ws4, header_row4, len(headers4))

boletos = [
    (1, 'BOLETO TERRESTRE (IDA)', 'EMPRESA DE TRANSPORTES PERU BUS S.A.',
     '20350227630', 'F26M-00020463', '18/01/2026', '18/01/2026', '16:20',
     'LIMA - CANETE', 'MERA CASAS JULIO IVAN', '09620528', 25.00),
    (2, 'BOLETO TERRESTRE (RETORNO)', 'EMPRESA DE TRANSPORTES PERU BUS S.A.',
     '20350227630', 'F54L-00012299', '21/01/2026', '21/01/2026', '18:01',
     'CANETE - LIMA', 'MERA CASAS JULIO IVAN', '09620528', 15.00),
]

for i, row_data in enumerate(boletos, start=header_row4 + 1):
    for col, val in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=col, value=val)
        if col == 12:
            cell.number_format = money_format
    style_data(ws4, i, len(headers4))

total_row4 = header_row4 + 1 + len(boletos)
ws4.merge_cells(f'A{total_row4}:K{total_row4}')
ws4.cell(row=total_row4, column=1, value='TOTAL PASAJES').font = Font(bold=True, size=10)
ws4.cell(row=total_row4, column=12, value=40.00).number_format = money_format
ws4.cell(row=total_row4, column=12).font = Font(bold=True, size=10)
for c in range(1, 13):
    ws4.cell(row=total_row4, column=c).border = thin_border

# Notas
nota_row4 = total_row4 + 2
ws4.merge_cells(f'A{nota_row4}:L{nota_row4}')
ws4.cell(row=nota_row4, column=1,
         value='NOTA: Este expediente NO tiene boletos aereos ni boarding pass. '
               'El transporte fue terrestre (bus interprovincial Lima-Canete via Peru Bus).'
         ).font = Font(italic=True, size=9, color='0070C0')

nota_row4b = nota_row4 + 1
ws4.merge_cells(f'A{nota_row4b}:L{nota_row4b}')
ws4.cell(row=nota_row4b, column=1,
         value='Las facturas de Peru Bus incluyen datos del pasajero (nombre y DNI), '
               'ruta, fecha y hora de embarque.').font = Font(italic=True, size=9, color='0070C0')

# Constancias de comision
const_title_row = nota_row4b + 2
ws4.merge_cells(f'A{const_title_row}:L{const_title_row}')
ws4.cell(row=const_title_row, column=1,
         value='CONSTANCIAS DE COMISION DE SERVICIO (Anexo 5):').font = Font(bold=True, size=10)

const_header = const_title_row + 1
const_headers = ['N', 'FECHA', 'LUGAR', 'DEPENDENCIA DESTINO', 'OBSERVACIONES']
for col, h in enumerate(const_headers, 1):
    ws4.cell(row=const_header, column=col, value=h)
style_header(ws4, const_header, len(const_headers))

constancias = [
    (1, '19/01/2026', 'UGEL 08 CANETE', 'UGEL 08 CANETE',
     'Firmada y sellada por Prof. Fidel Ramos En Borja'),
    (2, '20/01/2026', 'UGEL 08 CANETE', 'UGEL 08 CANETE',
     'Firmada y sellada por Prof. Fidel Ramos En Borja'),
    (3, '21/01/2026', 'UGEL 08 CANETE', 'UGEL 08 CANETE',
     'Firmada y sellada por Prof. Fidel Ramos En Borja'),
]

for i, row_data in enumerate(constancias, start=const_header + 1):
    for col, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=col, value=val)
    style_data(ws4, i, len(const_headers))

# Registro Anexo 6
sup_row = const_header + 1 + len(constancias) + 1
ws4.merge_cells(f'A{sup_row}:L{sup_row}')
ws4.cell(row=sup_row, column=1,
         value='CONSTANCIA DE COMISION - SUPERVISION Y MONITOREO (Anexo 6):').font = Font(bold=True, size=10)
sup_detail = sup_row + 1
ws4.cell(row=sup_detail, column=1, value='Registro:').font = Font(bold=True, size=9)
ws4.cell(row=sup_detail, column=2,
         value='19/01/26, 20/01/26, 21/01/26 - UGEL 8 CANETE. DNI 09620528. '
               'Firmado por Prof. Fidel Ramos En Borja').font = data_font

# Anchos
col_widths4 = [4, 28, 38, 14, 18, 12, 14, 14, 20, 28, 12, 14]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

print('Hoja 4 (Boletos/Boarding) completada')

# ============================================================
# GUARDAR
# ============================================================
output_dir = r'C:\Users\Hans\Proyectos\AG-EVIDENCE\data\expedientes\pruebas\viaticos_2026\OTIC2026-INT-0072834'
output_file = os.path.join(output_dir, 'RENDICION_OTIC2026-INT-0072834.xlsx')
wb.save(output_file)
print(f'\n{"="*60}')
print(f'EXCEL GENERADO EXITOSAMENTE')
print(f'{"="*60}')
print(f'Archivo: {output_file}')
print(f'Hojas:   {wb.sheetnames}')
print(f'  1. Anexo 3 - Rendicion:     19 items, S/924.00 con docs')
print(f'  2. DJ - Anexo 4:            24 items, S/98.00 sin docs')
print(f'  3. Comprobantes de Pago:    16 facturas fisicas, S/{total_sum:.2f}')
print(f'  4. Boletos y Boarding Pass: 2 boletos terrestres, S/40.00')
print(f'{"="*60}')
