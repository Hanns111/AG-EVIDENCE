#!/usr/bin/env python3
"""
Generador de Excel v2 para expediente DEBEDSAR2026-INT-0146130
Comisionado: MARTIARENA CARHUARUPAY, Víctor
Destino: Lima → Tarapoto → Chachapoyas → Tarapoto → Lima
Fechas: 07/02/2026 - 10/02/2026

4 Hojas: Anexo 3 | Comprobantes de Pago | Declaración Jurada | Boletos y Boarding Pass

=== REGLAS ESTRICTAS v2 ===
1. Datos de Comprobantes = EXCLUSIVAMENTE del documento fuente (factura/boleta física)
2. NULL = dato no visible para el motor de extracción
3. SIN inferencias — si el texto dice "NUEVA", se pone "NUEVA", no se completa
4. SIN cruces — NO comparar Anexo 3 con documento fuente en observaciones
5. SIN correcciones manuales — si VLM lee algo incorrecto, se reporta tal cual
6. Observaciones = solo texto relevante DEL PROPIO comprobante (pie de página, etc.)

Herramientas utilizadas:
- PyMuPDF (fitz): Extracción de texto digital embebido en PDFs
- Qwen2.5-VL-7B (via Ollama, 500 DPI): Extracción visual de páginas escaneadas
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# DATOS EXTRAÍDOS — ANEXO 3 (Fuente: PDF Rendición, pág 1)
# Motor: PyMuPDF (texto digital embebido)
# ============================================================
DATOS_EXPEDIENTE = {
    "sinad": "DEBEDSAR2026-INT-0146130",
    "comisionado": "MARTIARENA CARHUARUPAY VICTOR",
    "dni": "25185850",
    "planilla": "00399",
    "exp_siaf": "2601514",
    "ue": "026 PROGRAMA EDUCACION BASICA PARA TODOS",
    "salida": "07/02/2026",
    "regreso": "10/02/2026",
    "dias_horas": "3d 7h",
    "motivo": "SUPERVISION PUA 2026",
    "destino": "LIMA - TARAPOTO - CHACHAPOYAS - TARAPOTO - LIMA",
    "viatico_otorgado": 1640.00,
    "total_comprobantes": 1081.50,
    "total_dj": 100.00,
    "total_gastado": 1181.50,
    "devolucion": 458.50,
}

# Comprobantes del Anexo 3 (texto digital, PyMuPDF pág 1)
ANEXO3_COMPROBANTES = [
    {
        "nro": 1,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "PALOMINO AGUILAR DAVID",
        "numero": "E001-3990",
        "concepto": "MOVILIDAD",
        "importe": 70.00,
    },
    {
        "nro": 2,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "PAPACHOS RESTAURANTES S.A.C.",
        "numero": "F205-00012299",
        "concepto": "ALIMENTACION",
        "importe": 42.00,
    },
    {
        "nro": 3,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "DIAZ RIOS ETY DORIS",
        "numero": "F001-00007330",
        "concepto": "ALIMENTACION",
        "importe": 22.00,
    },
    {
        "nro": 4,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "HOTELERO DORDEAN SAC",
        "numero": "F002-4814",
        "concepto": "HOSPEDAJE",
        "importe": 150.00,
    },
    {
        "nro": 5,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "EMPRESA DE TRANSPORTES Y TURISMO RIOJA S.A",
        "numero": "FT05-00016765",
        "concepto": "PASAJE",
        "importe": 50.00,
    },
    {
        "nro": 6,
        "fecha": "07/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "EMPRESA DE TRANSPORTE EVANGELIO PODER DE DIOS",
        "numero": "FT06-00002741",
        "concepto": "PASAJE",
        "importe": 70.00,
    },
    {
        "nro": 7,
        "fecha": "08/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "HOTELERO DORDEAN SAC",
        "numero": "F002-04814",
        "concepto": "HOSPEDAJE",
        "importe": 150.00,
    },
    {
        "nro": 8,
        "fecha": "08/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "LA REAL CECINA E.I.R.L.",
        "numero": "E001-8818",
        "concepto": "ALIMENTACION",
        "importe": 37.00,
    },
    {
        "nro": 9,
        "fecha": "08/02/2026",
        "tipo_doc": "Factura",
        "razon_social": '"CORPORACION COMERCIAL Y DE SERVICIOS PARIS" E.I.R.L.',
        "numero": "FW01-00003756",
        "concepto": "ALIMENTACION",
        "importe": 24.00,
    },
    {
        "nro": 10,
        "fecha": "09/02/2026",
        "tipo_doc": "Boleta de Venta",
        "razon_social": "CIEZA BUSTAMANTE DARIO",
        "numero": "0001-005367",
        "concepto": "ALIMENTACION",
        "importe": 25.00,
    },
    {
        "nro": 11,
        "fecha": "09/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "GRUPO EMPRESARIAL GLORIA A DIOS SAC",
        "numero": "FP01-233",
        "concepto": "ALIMENTACION",
        "importe": 49.00,
    },
    {
        "nro": 12,
        "fecha": "09/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "LA OLLA DE BARRO E.I.R.L",
        "numero": "F002-12174",
        "concepto": "ALIMENTACION",
        "importe": 43.00,
    },
    {
        "nro": 13,
        "fecha": "09/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "AMAZON GOLD GRAIN E.I.R.L.",
        "numero": "FA01-00000952",
        "concepto": "HOSPEDAJE",
        "importe": 120.00,
    },
    {
        "nro": 14,
        "fecha": "10/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "INVERSIONES BGG ORIENTE SAC",
        "numero": "F021-00004515",
        "concepto": "ALIMENTACION",
        "importe": 34.50,
    },
    {
        "nro": 15,
        "fecha": "10/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "PALOMINO AGUILAR DAVID",
        "numero": "E001-4002",
        "concepto": "MOVILIDAD",
        "importe": 80.00,
    },
    {
        "nro": 16,
        "fecha": "10/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "W.E. CASTILLO S.A.C.",
        "numero": "F001-00000664",
        "concepto": "PASAJE",
        "importe": 75.00,
    },
    {
        "nro": 17,
        "fecha": "10/02/2026",
        "tipo_doc": "Factura",
        "razon_social": "EMPRESA SAN MARTIN SA",
        "numero": "FT27-526327",
        "concepto": "PASAJE",
        "importe": 40.00,
    },
]

# ============================================================
# COMPROBANTES DE PAGO — DOCUMENTO FUENTE (v2 LIMPIO)
# Cada registro = datos TAL CUAL del documento fuente
# NULL = dato no visible para el motor de extracción
# SIN inferencias, SIN cruces con Anexo 3, SIN correcciones
# ============================================================
COMPROBANTES_FUENTE = [
    # --- COMPROBANTE 1: E001-3990 (PyMuPDF, pág 38) ---
    {
        "pagina_pdf": 38,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "E001-3990",
        "ruc_emisor": "10257530171",
        "razon_social_emisor": "PALOMINO AGUILAR DAVID",
        "direccion_emisor": "JR. MARISCAL RAMON CASTILLA 348 URB. PLAYA RIMAC CALLAO - PROV. CONST. DEL CALLAO - PROV. CONST. DEL CALLAO",
        "fecha_emision": "07/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "SERVICIO DE TAXI DE LA VICTORIA AL AEROPUERTO DE LIMA",
        "valor_venta": "70.00",
        "igv": "0.00",
        "total": "70.00",
        "exonerado": "",
        "forma_pago": "Al Contado",
        "observaciones": "",
    },
    # --- COMPROBANTE 2: F205-00012200 (Qwen2.5-VL-7B 500DPI, pág 41) ---
    {
        "pagina_pdf": 41,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F205-00012200",
        "ruc_emisor": "20544822804",
        "razon_social_emisor": "PAPACHOS RESTAURANTES S.A.C.",
        "direccion_emisor": "AVENIDA NESTOR SARRIA 1570 30027 Prov. Const. del Callao Prov. Const. del Callao Callao",
        "fecha_emision": "07/02/2026",
        "ruc_comprador": "20380795007",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "SANGUICHE BUTITARRA (1 un) S/ 29.00 CAFE AMERICANO 12 oz (1 un) S/ 13.00",
        "valor_venta": "S/ 42.00",
        "igv": "S/ 6.15",
        "total": "S/ 42.00",
        "exonerado": "S/ 0.00",
        "forma_pago": "TC/TF - Visa",
        "observaciones": "",
    },
    # --- COMPROBANTE 3: F001-00007330 (Qwen2.5-VL-7B 500DPI, pág 44) ---
    {
        "pagina_pdf": 44,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F001-00007330",
        "ruc_emisor": "10403765452",
        "razon_social_emisor": "DIAZ RIOS ETY DORIS",
        "direccion_emisor": "Jr. BOLOGNESI 206 NUEVA CAJAMARCA RIOJA SAN MARTIN",
        "fecha_emision": "07/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "1.00 CECINA DE RES 20.00 / 1.00 AGUA SAN LUIS 2.00",
        "valor_venta": "20.00",
        "igv": "0.00",
        "total": "22.00",
        "exonerado": "22.00",
        "forma_pago": "CONTADO",
        "observaciones": "Bienes Transferidos en la Amazonia Para ser consumidos en la misma.",
    },
    # --- COMPROBANTE 4: F002-4814 (PyMuPDF, pág 47) ---
    # Una sola factura por 2 noches. Dato tal cual del documento fuente.
    {
        "pagina_pdf": 47,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F002-4814",
        "ruc_emisor": "20609843277",
        "razon_social_emisor": "GRUPO HOTELERO DORDEAN S.A.C.",
        "direccion_emisor": "JR. AMAZONAS NRO. 840 URB. CHACHAPOYAS AMAZONAS - CHACHAPOYAS - CHACHAPOYAS",
        "fecha_emision": "08/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "NUEVO SOL",
        "descripcion": "HOSPEDAJE (1.0 und, 300.0). Nro. habitacion: 09. Checkin: 07/02/2026 Checkout: 09/02/2026 Nro. noches: 2. COSTO DE HABITACION: S/. 150.00 (Por noche)",
        "valor_venta": "0.00",
        "igv": "0.00",
        "total": "300.00",
        "exonerado": "300.00",
        "forma_pago": "CONTADO",
        "observaciones": "",
    },
    # --- COMPROBANTE 5: FT05-00016765 (Qwen2.5-VL-7B 500DPI, pág 14) ---
    {
        "pagina_pdf": 14,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FT05-00016765",
        "ruc_emisor": "20284896506",
        "razon_social_emisor": "EMPRESA DE TRANSPORTES Y TURISMO RIOJA S.A.",
        "direccion_emisor": "JR. SANTO TORIBIO NRO. 743, SAN MARTIN-MOYOBAMBA-MOYOBAMBA",
        "fecha_emision": "07/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "Pasaje: TARAPOTO->MOYOBAMBA",
        "valor_venta": "50.00",
        "igv": "0.00",
        "total": "50.00",
        "exonerado": "0.00",
        "forma_pago": "CONTADO",
        "observaciones": "SERVICIOS TRANSFERIDOS EN LA AMAZONIA PARA SER CONSUMIDOS EN LA MISMA",
    },
    # --- COMPROBANTE 6: FT06-00002741 (PyMuPDF, pág 17) ---
    {
        "pagina_pdf": 17,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FT06-00002741",
        "ruc_emisor": "20487485463",
        "razon_social_emisor": "EMPRESA DE TRANSPORTE EVANGELIO PODER DE DIOS EMPRESA INDIVIDUAL DE RESPONSABILIDAD LIMITADA",
        "direccion_emisor": "AMAZONAS - CHACHAPOYAS - CHACHAPOYAS JR. TRIUNFO NRO. S/N - TERMINAL TERRESTRE STAN A-03",
        "fecha_emision": "07/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOL",
        "descripcion": "Pasaje: NVA. CAJAMARCA->CHACHAPOYAS (1 und, ZZ, 70.00)",
        "valor_venta": "70.00",
        "igv": "0.00",
        "total": "70.00",
        "exonerado": "70.00",
        "forma_pago": "CONTADO",
        "observaciones": "SERVICIOS PRESTADOS EN LA AMAZONIA REGION SELVA PARA SER CONSUMIDOS EN LA MISMA",
    },
    # --- COMPROBANTE 7: F002-4814 (misma factura que comp 4, línea 7 del Anexo 3) ---
    # El documento fuente es el MISMO de pág 47 — se repite tal cual
    {
        "pagina_pdf": 47,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F002-4814",
        "ruc_emisor": "20609843277",
        "razon_social_emisor": "GRUPO HOTELERO DORDEAN S.A.C.",
        "direccion_emisor": "JR. AMAZONAS NRO. 840 URB. CHACHAPOYAS AMAZONAS - CHACHAPOYAS - CHACHAPOYAS",
        "fecha_emision": "08/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "NUEVO SOL",
        "descripcion": "HOSPEDAJE (1.0 und, 300.0). Nro. habitacion: 09. Checkin: 07/02/2026 Checkout: 09/02/2026 Nro. noches: 2. COSTO DE HABITACION: S/. 150.00 (Por noche)",
        "valor_venta": "0.00",
        "igv": "0.00",
        "total": "300.00",
        "exonerado": "300.00",
        "forma_pago": "CONTADO",
        "observaciones": "",
    },
    # --- COMPROBANTE 8: E001-8818 (PyMuPDF, pág 50) ---
    {
        "pagina_pdf": 50,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "E001-8818",
        "ruc_emisor": "20487969231",
        "razon_social_emisor": "LA REAL CECINA E.I.R.L.",
        "direccion_emisor": "CAL. HERMOSURA 676 URB. CHACHAPOYAS CHACHAPOYAS - CHACHAPOYAS - AMAZONAS",
        "fecha_emision": "08/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "CECINA SHILPIDA (1 und, S/32.00) / INFUSION NATURAL (1 und, S/5.00)",
        "valor_venta": "37.00",
        "igv": "0.00",
        "total": "37.00",
        "exonerado": "",
        "forma_pago": "Al Contado",
        "observaciones": "",
    },
    # --- COMPROBANTE 9: FW01-00003756 (Qwen2.5-VL-7B 500DPI, pág 53) ---
    {
        "pagina_pdf": 53,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FW01-00003756",
        "ruc_emisor": "20602428452",
        "razon_social_emisor": "CORPORACION COMERCIAL Y DE SERVICIOS PARIS E.I.R.L.",
        "direccion_emisor": "JR. AMAZONAS NRO. 876 URB. CHACHAPOYAS",
        "fecha_emision": "08/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "PAPAYA: 1 Un x 7.00; CHEESECAKE DE FRESA: 1 Un x 13.00; AGUA MINERAL: 1 Un x 4.00",
        "valor_venta": "24.00",
        "igv": "0.00",
        "total": "24.00",
        "exonerado": "24.00",
        "forma_pago": "Tarjeta",
        "observaciones": "",
    },
    # --- COMPROBANTE 10: 001-005367 (Qwen2.5-VL-7B 500DPI, pág 56) ---
    {
        "pagina_pdf": 56,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "BOLETA DE VENTA",
        "serie_numero": "001-005367",
        "ruc_emisor": "10272647297",
        "razon_social_emisor": "Restaurant Clary",
        "direccion_emisor": "CARR. MARGINAL NRO. 912 POMACOCHAS - AMAZONAS - BONGARA - FLORIDA",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "25185850",
        "razon_social_comprador": "Señor: Rogelio Eduardo Baeza Luna Tadeo",
        "moneda": "SOLES",
        "descripcion": "Caldo Calli 25.00",
        "valor_venta": "25.00",
        "igv": "0.00",
        "total": "25.00",
        "exonerado": "",
        "forma_pago": "contado",
        "observaciones": "BIENES TRANSFERIDOS/SERVICIOS PRESTADOS EN LA AMAZONIA PARA SER CONSUMIDOS EN LA MISMA",
    },
    # --- COMPROBANTE 11: FP01-233 (Qwen2.5-VL-7B 500DPI, pág 60) ---
    {
        "pagina_pdf": 60,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FP01-233",
        "ruc_emisor": "20611761571",
        "razon_social_emisor": "GRUPO EMPRESARIAL GLORIA A DIOS S.A.C.",
        "direccion_emisor": "JR. PLAZA MAYOR NRO. 321 DPTO. 3 URB. CENTRO (FRENTE A LA PLAZA DE ARMAS DE TARAPOTO) SAN MARTIN - SAN MARTIN - TARAPOTO",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "[1] PAIACONES RELLENOS 25.00 [1] BATIDO DE FRESA 18.00 [1] AGUA CON Y SIN GAS 6.00",
        "valor_venta": "49.00",
        "igv": "0.00",
        "total": "49.00",
        "exonerado": "49.00",
        "forma_pago": "CONTADO",
        "observaciones": "BIENES TRANSFERIDOS A LA AMAZONIA PARA SER CONSUMIDOS EN LA MISMA",
    },
    # --- COMPROBANTE 12: F002-12174 (Qwen2.5-VL-7B 500DPI, pág 63) ---
    {
        "pagina_pdf": 63,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA DE VENTA ELECTRONICA",
        "serie_numero": "F002-12174",
        "ruc_emisor": "20531527080",
        "razon_social_emisor": "LA OLLA DE BARRO EIRL",
        "direccion_emisor": "Jr. Pedro Canaa 398 - Moyobamba - Telf. 925779705",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "1 mixto con tacacho 35.00 / 1 Uva / 1/2 jarra 8.00",
        "valor_venta": "43.00",
        "igv": "0.00",
        "total": "43.00",
        "exonerado": "43.00",
        "forma_pago": "CONTADO",
        "observaciones": "SERVICIOS PRESTADOS EN LA AMAZONIA REGION SELVA PARA SER DOS EN LA MISMA",
    },
    # --- COMPROBANTE 13: FA01-00000952 (Qwen2.5-VL-7B 500DPI, pág 66) ---
    {
        "pagina_pdf": 66,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FA01-00000952",
        "ruc_emisor": "20601209889",
        "razon_social_emisor": "AMAZON GOLD GRAIN E.I.R.L.",
        "direccion_emisor": "JR. RIOJA NRO. 357 SAN MARTIN - SAN MARTIN - TARAPOTO",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "SERVICIO DE ALOJAMIENTO DEL 09/02/2026 AL 10/02/2026",
        "valor_venta": "120.00",
        "igv": "0.00",
        "total": "120.00",
        "exonerado": "120.00",
        "forma_pago": "Efectivo",
        "observaciones": "",
    },
    # --- COMPROBANTE 14: F021-00004515 (Qwen2.5-VL-7B 500DPI, pág 69) ---
    {
        "pagina_pdf": 69,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F021-00004515",
        "ruc_emisor": "20541192833",
        "razon_social_emisor": "INVERSIONES BGG ORIENTE SAC",
        "direccion_emisor": 'Aeropuerto Cadete FAP "Guillermo del Castillo Paredes" Av. Aviacion, Tarapoto 22201, Region de San Martin, Peru',
        "fecha_emision": "10/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "1 CAFE AMERICANO 10.00 / 1 SANDWICH MIXTO DE JAMON AHUMADO 24.50",
        "valor_venta": "32.55",
        "igv": "1.95",
        "total": "34.50",
        "exonerado": "0.00",
        "forma_pago": "CONTADO",
        "observaciones": "",
    },
    # --- COMPROBANTE 15: E001-4002 (PyMuPDF, pág 72) ---
    {
        "pagina_pdf": 72,
        "motor": "PyMuPDF",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "E001-4002",
        "ruc_emisor": "10257530171",
        "razon_social_emisor": "PALOMINO AGUILAR DAVID",
        "direccion_emisor": "JR. MARISCAL RAMON CASTILLA 348 URB. PLAYA RIMAC CALLAO - PROV. CONST. DEL CALLAO - PROV. CONST. DEL CALLAO",
        "fecha_emision": "10/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "SERVICIO DE TAXI DEL AEROPUERTO DE LIMA A LA VICTORIA",
        "valor_venta": "80.00",
        "igv": "0.00",
        "total": "80.00",
        "exonerado": "",
        "forma_pago": "Al Contado",
        "observaciones": "",
    },
    # --- COMPROBANTE 16: F001-00000664 (Qwen2.5-VL-7B 500DPI, pág 20) ---
    {
        "pagina_pdf": 20,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "F001-00000664",
        "ruc_emisor": "20604862311",
        "razon_social_emisor": "EXPRESO Y TURISMO BORYS SIERRA NORTE S.A.C.",
        "direccion_emisor": "MOYOBAMBA, MOYOBAMBA - SAN MARTIN",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PAR. TODOS",
        "moneda": "SOLES",
        "descripcion": "1 Pasaje CHACHAPOYAS-MOYOBAMBA",
        "valor_venta": "75.00",
        "igv": "0.00",
        "total": "75.00",
        "exonerado": "75.00",
        "forma_pago": "CONTADO",
        "observaciones": "",
    },
    # --- COMPROBANTE 17: FT27-00027639 (Qwen2.5-VL-7B 500DPI, pág 23) ---
    {
        "pagina_pdf": 23,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
        "tipo_comprobante": "FACTURA ELECTRONICA",
        "serie_numero": "FT27-00027639",
        "ruc_emisor": "20531409478",
        "razon_social_emisor": "EMPRESA SAN MARTIN SA",
        "direccion_emisor": "SAN MARTIN-MOYOBAMBA-MOYOBAMBA JR. 20 DE ABRIL C-13",
        "fecha_emision": "09/02/2026",
        "ruc_comprador": "20380795907",
        "razon_social_comprador": "PROGRAMA EDUCACION BASICA PARA TODOS",
        "moneda": "SOLES",
        "descripcion": "Pasaje: MOYOBAMBA -> TARAPOTO 1 UND 40.00",
        "valor_venta": "40.00",
        "igv": "0.00",
        "total": "40.00",
        "exonerado": "0.00",
        "forma_pago": "CONTADO",
        "observaciones": "SERVICIOS TRANSFERIDOS EN LA AMAZONIA PARA SER CONSUMIDOS EN LA MISMA",
    },
]

# ============================================================
# DECLARACIÓN JURADA (Fuente: PDF Rendición, pág 3)
# Motor: PyMuPDF (texto digital)
# TEXTO LITERAL — sin completar palabras truncadas
# ============================================================
DJ_GASTOS = [
    {
        "fecha": "07/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD MOYOBAMBA - NUEVA",
        "importe": 15.00,
    },
    {
        "fecha": "07/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD TERMINAL CHACHAPOYAS - HOSPEDAJE",
        "importe": 10.00,
    },
    {
        "fecha": "08/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD HOSPEDAJE CHACHAPOYAS - COAR AMAZONAS",
        "importe": 15.00,
    },
    {
        "fecha": "08/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD COAR AMAZONAS - HOSPEDAJE CHACHAPOYAS",
        "importe": 15.00,
    },
    {
        "fecha": "09/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD HOSPEDAJE CHACHAPOYAS - TERMINAL",
        "importe": 10.00,
    },
    {
        "fecha": "09/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "TERMINAL TARAPOTO - HOSPEDAJE TARAPOTO",
        "importe": 15.00,
    },
    {
        "fecha": "10/02/2026",
        "concepto": "MOVILIDAD",
        "detalle": "MOVILIDAD HOSPEDAJE TARAPOTO - AEROPUERTO",
        "importe": 20.00,
    },
]

# ============================================================
# BOLETOS Y BOARDING PASS
# ============================================================
BOLETOS = [
    {
        "tipo": "TIQUETE AEREO",
        "aerolinea": "JETSMART AIRLINES PERU S.A.C.",
        "ruc_aerolinea": "20607393649",
        "nro_vuelo": "JA 7258",
        "pasajero": "MARTIRENA CARHUARUPAY VICTOR",
        "origen": "Lima",
        "destino": "Tarapoto",
        "fecha": "07/02/2026",
        "hora_salida": "09:00",
        "hora_llegada": "10:29",
        "codigo_reserva": "TG3DHE",
        "nro_orden": "041324071509",
        "tarifa": "93.00 USD",
        "total": "144.17 USD",
        "asiento": "",
        "pagina_pdf": "10-11",
        "motor": "PyMuPDF",
    },
    {
        "tipo": "TIQUETE AEREO",
        "aerolinea": "SKY AIRLINE PERU S.A.C.",
        "ruc_aerolinea": "20603446543",
        "nro_vuelo": "H2 5403",
        "pasajero": "VICTOR MARTIRENA CARHUARUPAY",
        "origen": "Tarapoto",
        "destino": "Lima",
        "fecha": "10/02/2026",
        "hora_salida": "07:30",
        "hora_llegada": "08:55",
        "codigo_reserva": "SVXHWL",
        "nro_orden": "bbbn206348",
        "tarifa": "94.00 USD",
        "total": "118.31 USD",
        "asiento": "",
        "pagina_pdf": "26-27",
        "motor": "PyMuPDF",
    },
    {
        "tipo": "BOARDING PASS",
        "aerolinea": "JetSMART Airlines S.A.C.",
        "ruc_aerolinea": "",
        "nro_vuelo": "JA 7258",
        "pasajero": "MR VICTOR MARTIARENA CARHUARUPAY",
        "origen": "Lima",
        "destino": "Tarapoto",
        "fecha": "07/02/2026",
        "hora_salida": "09:00",
        "hora_llegada": "10:29",
        "codigo_reserva": "TG3DHE",
        "nro_orden": "",
        "tarifa": "",
        "total": "",
        "asiento": "5C",
        "pagina_pdf": 36,
        "motor": "Qwen2.5-VL-7B (500 DPI)",
    },
    {
        "tipo": "BOARDING PASS",
        "aerolinea": "Sky Airline",
        "ruc_aerolinea": "",
        "nro_vuelo": "H2 5403",
        "pasajero": "Victor Martiarena Carhuarupay",
        "origen": "Tarapoto",
        "destino": "Lima",
        "fecha": "10/02/2026",
        "hora_salida": "07:30",
        "hora_llegada": "08:55",
        "codigo_reserva": "SVXHWL",
        "nro_orden": "",
        "tarifa": "",
        "total": "",
        "asiento": "15F",
        "pagina_pdf": 37,
        "motor": "PyMuPDF",
    },
]


# ============================================================
# GENERACIÓN DEL EXCEL
# ============================================================


def style_header(ws, row, max_col, fill_color="1F4E79"):
    """Aplica estilo de encabezado a una fila."""
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data(ws, start_row, end_row, max_col):
    """Aplica bordes y alineación a datos."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def crear_hoja_anexo3(wb):
    """Hoja 1: Anexo 3 — Planilla de Gastos."""
    ws = wb.active
    ws.title = "ANEXO_3"

    # Datos generales
    ws["A1"] = "ANEXO 3 — RENDICION DE CUENTAS POR COMISION DE SERVICIOS"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:G1")

    info_rows = [
        ("SINAD:", DATOS_EXPEDIENTE["sinad"]),
        ("Comisionado:", DATOS_EXPEDIENTE["comisionado"]),
        ("Nº Planilla:", DATOS_EXPEDIENTE["planilla"]),
        ("Exp SIAF:", DATOS_EXPEDIENTE["exp_siaf"]),
        ("Salida:", DATOS_EXPEDIENTE["salida"]),
        ("Regreso:", DATOS_EXPEDIENTE["regreso"]),
        ("Dias/Horas:", DATOS_EXPEDIENTE["dias_horas"]),
        ("Motivo:", DATOS_EXPEDIENTE["motivo"]),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=value)

    # Encabezados de tabla
    headers = ["Nº", "FECHA", "TIPO DOC.", "RAZON SOCIAL", "NUMERO", "CONCEPTO", "IMPORTE S/"]
    header_row = 12
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(headers))

    # Datos
    for i, c in enumerate(ANEXO3_COMPROBANTES, start=header_row + 1):
        ws.cell(row=i, column=1, value=c["nro"])
        ws.cell(row=i, column=2, value=c["fecha"])
        ws.cell(row=i, column=3, value=c["tipo_doc"])
        ws.cell(row=i, column=4, value=c["razon_social"])
        ws.cell(row=i, column=5, value=c["numero"])
        ws.cell(row=i, column=6, value=c["concepto"])
        ws.cell(row=i, column=7, value=c["importe"])
        ws.cell(row=i, column=7).number_format = "#,##0.00"

    last_data = header_row + len(ANEXO3_COMPROBANTES)
    style_data(ws, header_row + 1, last_data, len(headers))

    # Totales
    r = last_data + 1
    ws.cell(row=r, column=5, value="(1) GASTOS CON DOCUMENTACION").font = Font(bold=True)
    ws.cell(
        row=r, column=7, value=DATOS_EXPEDIENTE["total_comprobantes"]
    ).number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=5, value="(2) GASTOS SIN DOCUMENTACION (DJ)").font = Font(bold=True)
    ws.cell(row=r, column=7, value=DATOS_EXPEDIENTE["total_dj"]).number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=5, value="(3) TOTAL GASTADO").font = Font(bold=True, size=11)
    ws.cell(row=r, column=7, value=DATOS_EXPEDIENTE["total_gastado"])
    ws.cell(row=r, column=7).font = Font(bold=True, size=11)
    ws.cell(row=r, column=7).number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=5, value="(4) DEVOLUCION").font = Font(bold=True)
    ws.cell(row=r, column=7, value=DATOS_EXPEDIENTE["devolucion"]).number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=5, value="(5) MONTO RECIBIDO").font = Font(bold=True, size=11)
    ws.cell(row=r, column=7, value=DATOS_EXPEDIENTE["viatico_otorgado"])
    ws.cell(row=r, column=7).font = Font(bold=True, size=11)
    ws.cell(row=r, column=7).number_format = "#,##0.00"

    # Metadata
    r += 2
    ws.cell(row=r, column=1, value="Fuente: PDF Rendicion pag 1 | Motor: PyMuPDF (texto digital)")
    ws.cell(row=r, column=1).font = Font(italic=True, color="808080", size=8)

    # Anchos
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14


def crear_hoja_comprobantes(wb):
    """Hoja 2: Comprobantes de Pago — Documento Fuente."""
    ws = wb.create_sheet("COMPROBANTES_PAGO")

    ws["A1"] = "COMPROBANTES DE PAGO — DOCUMENTO FUENTE"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:R1")

    ws["A2"] = (
        "Datos extraidos TAL CUAL del documento fuente. NULL = no visible al motor. Sin inferir, sin cruzar con Anexo 3, sin corregir."
    )
    ws["A2"].font = Font(italic=True, color="FF0000", size=9)
    ws.merge_cells("A2:R2")

    headers = [
        "Nro",
        "PAG PDF",
        "MOTOR",
        "TIPO",
        "SERIE-NUMERO",
        "RUC EMISOR",
        "RAZON SOCIAL EMISOR",
        "DIRECCION EMISOR",
        "FECHA EMISION",
        "RUC COMPRADOR",
        "MONEDA",
        "DESCRIPCION",
        "VALOR VENTA",
        "IGV",
        "TOTAL",
        "EXONERADO",
        "FORMA PAGO",
        "OBSERVACIONES",
    ]
    header_row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(headers))

    for i, c in enumerate(COMPROBANTES_FUENTE, start=header_row + 1):
        ws.cell(row=i, column=1, value=i - header_row)
        ws.cell(row=i, column=2, value=c["pagina_pdf"])
        ws.cell(row=i, column=3, value=c["motor"])
        ws.cell(row=i, column=4, value=c["tipo_comprobante"])
        ws.cell(row=i, column=5, value=c["serie_numero"])
        ws.cell(row=i, column=6, value=c["ruc_emisor"])
        ws.cell(row=i, column=7, value=c["razon_social_emisor"])
        ws.cell(row=i, column=8, value=c.get("direccion_emisor"))
        ws.cell(row=i, column=9, value=c["fecha_emision"])
        ws.cell(row=i, column=10, value=c.get("ruc_comprador"))
        ws.cell(row=i, column=11, value=c["moneda"])
        ws.cell(row=i, column=12, value=c["descripcion"])
        ws.cell(row=i, column=13, value=c.get("valor_venta"))
        ws.cell(row=i, column=14, value=c["igv"])
        ws.cell(row=i, column=15, value=c["total"])
        ws.cell(row=i, column=16, value=c.get("exonerado"))
        ws.cell(row=i, column=17, value=c.get("forma_pago"))
        ws.cell(row=i, column=18, value=c.get("observaciones"))

        # Highlight NULL cells in red
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            if cell.value is None:
                cell.value = "NULL"
                cell.font = Font(color="FF0000", italic=True)

    last_data = header_row + len(COMPROBANTES_FUENTE)
    style_data(ws, header_row + 1, last_data, len(headers))

    # Anchos
    widths = [4, 8, 20, 20, 18, 14, 35, 45, 12, 14, 10, 45, 12, 8, 10, 10, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def crear_hoja_dj(wb):
    """Hoja 3: Declaracion Jurada."""
    ws = wb.create_sheet("DECLARACION_JURADA")

    ws["A1"] = "DECLARACION JURADA DE GASTOS (ANEXO 4)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    ws["A2"] = (
        "Comisionado: " + DATOS_EXPEDIENTE["comisionado"] + " | DNI: " + DATOS_EXPEDIENTE["dni"]
    )
    ws["A2"].font = Font(size=10)

    ws["A3"] = "TEXTO LITERAL del PDF (PyMuPDF). Sin completar palabras truncadas."
    ws["A3"].font = Font(italic=True, color="FF0000", size=9)

    headers = ["Nro", "FECHA", "CONCEPTO", "DETALLE", "IMPORTE S/"]
    header_row = 5
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(headers))

    for i, g in enumerate(DJ_GASTOS, start=header_row + 1):
        ws.cell(row=i, column=1, value=i - header_row)
        ws.cell(row=i, column=2, value=g["fecha"])
        ws.cell(row=i, column=3, value=g["concepto"])
        ws.cell(row=i, column=4, value=g["detalle"])
        ws.cell(row=i, column=5, value=g["importe"])
        ws.cell(row=i, column=5).number_format = "#,##0.00"

    last_data = header_row + len(DJ_GASTOS)
    style_data(ws, header_row + 1, last_data, len(headers))

    # Total
    r = last_data + 1
    ws.cell(row=r, column=4, value="TOTAL S/").font = Font(bold=True)
    ws.cell(row=r, column=5, value=sum(g["importe"] for g in DJ_GASTOS))
    ws.cell(row=r, column=5).font = Font(bold=True)
    ws.cell(row=r, column=5).number_format = "#,##0.00"

    # Metadata
    r += 2
    ws.cell(row=r, column=1, value="Fuente: PDF Rendicion pag 3 | Motor: PyMuPDF (texto digital)")
    ws.cell(row=r, column=1).font = Font(italic=True, color="808080", size=8)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 14


def crear_hoja_boletos(wb):
    """Hoja 4: Boletos Aereos y Boarding Pass."""
    ws = wb.create_sheet("BOLETOS_BOARDING")

    ws["A1"] = "BOLETOS AEREOS Y BOARDING PASS"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:O1")

    headers = [
        "Nro",
        "TIPO",
        "AEROLINEA",
        "RUC AEROLINEA",
        "Nro VUELO",
        "PASAJERO",
        "ORIGEN",
        "DESTINO",
        "FECHA",
        "HORA SALIDA",
        "HORA LLEGADA",
        "COD. RESERVA",
        "ASIENTO",
        "TOTAL",
        "PAG/MOTOR",
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(headers))

    for i, b in enumerate(BOLETOS, start=header_row + 1):
        ws.cell(row=i, column=1, value=i - header_row)
        ws.cell(row=i, column=2, value=b["tipo"])
        ws.cell(row=i, column=3, value=b["aerolinea"])
        ws.cell(row=i, column=4, value=b.get("ruc_aerolinea"))
        ws.cell(row=i, column=5, value=b["nro_vuelo"])
        ws.cell(row=i, column=6, value=b["pasajero"])
        ws.cell(row=i, column=7, value=b["origen"])
        ws.cell(row=i, column=8, value=b["destino"])
        ws.cell(row=i, column=9, value=b["fecha"])
        ws.cell(row=i, column=10, value=b["hora_salida"])
        ws.cell(row=i, column=11, value=b["hora_llegada"])
        ws.cell(row=i, column=12, value=b.get("codigo_reserva"))
        ws.cell(row=i, column=13, value=b.get("asiento"))
        ws.cell(row=i, column=14, value=b.get("total"))
        ws.cell(row=i, column=15, value="P" + str(b["pagina_pdf"]) + " / " + str(b["motor"]))

        # Highlight NULL cells
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            if cell.value is None:
                cell.value = "NULL"
                cell.font = Font(color="FF0000", italic=True)

    last_data = header_row + len(BOLETOS)
    style_data(ws, header_row + 1, last_data, len(headers))

    # Anchos
    widths = [4, 16, 30, 14, 10, 35, 12, 12, 12, 10, 10, 12, 8, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = openpyxl.Workbook()

    crear_hoja_anexo3(wb)
    crear_hoja_comprobantes(wb)
    crear_hoja_dj(wb)
    crear_hoja_boletos(wb)

    # Guardar
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(
        os.path.dirname(output_dir), "output", "RENDICION_DEBEDSAR2026-INT-0146130_v2.xlsx"
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print("=" * 70)
    print("Excel v2 generado: " + output_path)
    print("=" * 70)
    print("  Hoja 1: ANEXO_3 (" + str(len(ANEXO3_COMPROBANTES)) + " comprobantes)")
    print("  Hoja 2: COMPROBANTES_PAGO (" + str(len(COMPROBANTES_FUENTE)) + " documentos fuente)")
    print(
        "  Hoja 3: DECLARACION_JURADA ("
        + str(len(DJ_GASTOS))
        + " gastos, S/"
        + str(sum(g["importe"] for g in DJ_GASTOS))
        + ")"
    )
    print("  Hoja 4: BOLETOS_BOARDING (" + str(len(BOLETOS)) + " documentos)")
    print()
    print("Herramientas utilizadas:")
    pymupdf_count = sum(1 for c in COMPROBANTES_FUENTE if c["motor"] == "PyMuPDF")
    vlm_count = sum(1 for c in COMPROBANTES_FUENTE if "Qwen" in c["motor"])
    print("  PyMuPDF (texto digital): " + str(pymupdf_count) + " comprobantes")
    print("  Qwen2.5-VL-7B (500 DPI): " + str(vlm_count) + " comprobantes")
    print()
    print("REGLAS APLICADAS v2:")
    print("  - DJ linea 1: 'MOVILIDAD MOYOBAMBA - NUEVA' (literal, sin completar)")
    print("  - Comprobantes: datos TAL CUAL del documento fuente")
    print("  - SIN cruces con Anexo 3")
    print("  - SIN correcciones manuales")
    print("  - NULL solo si no visible al motor")
    print("  - Direccion emisor: incluida (500 DPI mejoro extraccion)")


if __name__ == "__main__":
    main()
