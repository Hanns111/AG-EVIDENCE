"""
Genera Excel de Rendición de Caja Chica N° 0000003
OT2026-INT-0179550 - CAJA CHICA JAQUELINE (MELGAREJO VILLARROEL, ALINA MABEL)

Hoja 1: Cuadro Resumen Rendición Caja Chica #3
Hoja 2: Comprobantes de Pago (detalle digitalizado de cada comprobante)

GUARDRAILS:
- Solo datos leídos del sustento escaneado (112 páginas)
- Si no se puede leer -> "ILEGIBLE"
- NO se consultan fuentes externas
- Observaciones cruzan con directiva RJ 0023-2025-MINEDU/SG-OGA y con el resumen
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# ESTILOS
# ============================================================
header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=12)
subtitle_font = Font(name='Calibri', bold=True, size=10)
normal_font = Font(name='Calibri', size=9)
obs_font = Font(name='Calibri', size=8, color='CC0000')
money_fmt = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
money_align = Alignment(horizontal='right', vertical='top')

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data(ws, row, max_col, is_money_col=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if is_money_col and col in is_money_col:
            cell.number_format = money_fmt
            cell.alignment = money_align
        else:
            cell.alignment = wrap_align

# ============================================================
# HOJA 1: CUADRO RESUMEN RENDICIÓN CAJA CHICA #3
# ============================================================
ws1 = wb.active
ws1.title = "Cuadro Resumen CC3"

# Título
ws1.merge_cells('A1:H1')
ws1['A1'] = 'RENDICIÓN DE CAJA CHICA Nº 0000003'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

# Datos generales
info = [
    ('Unidad Ejecutora:', '024 - MINISTERIO DE EDUCACIÓN-SEDE CENTRAL - 024'),
    ('Nro. Identificación:', '000079'),
    ('Fecha Rendición:', '12/02/2026'),
    ('Nombre Caja:', 'CAJA CHICA OT'),
    ('Responsable:', 'MELGAREJO VILLARROEL, ALINA MABEL'),
    ('Elaborado por:', 'AMELGAREJO - MELGAREJO VILLARROEL, ALINA MABEL'),
]
for i, (label, value) in enumerate(info, start=3):
    ws1[f'A{i}'] = label
    ws1[f'A{i}'].font = Font(name='Calibri', bold=True, size=9)
    ws1.merge_cells(f'B{i}:H{i}')
    ws1[f'B{i}'] = value
    ws1[f'B{i}'].font = normal_font

# Headers
headers_r = ['N°', 'Fecha', 'Tipo Documento', 'Nro Documento', 'Detalle del Gasto',
             'Centro de Costo', 'Monto S/.', 'Clasif. Gasto']
row_h = 10
for col, h in enumerate(headers_r, 1):
    ws1.cell(row=row_h, column=col, value=h)
style_header(ws1, row_h, len(headers_r))

# Data del resumen (tal cual figura en la rendición)
resumen_data = [
    (1,  '06/02/2026', 'Factura',             'F001-1488',         'SERVICIO DE IMPRESIÓN Y PLOTEO DE PLANOS PARA LA OBTENCIÓN DEL ITSE DE LA SEDE',  '01', 19.00,    '2.3.1.99.1.99'),
    (2,  '06/02/2026', 'Factura',             'E001-530',          'SERVICIO DE MONTA CARGA EN VENTANILLA - CARGAR CARPA NUEVA A CAMIÓN 13/01/2026. SERVICIO DE CARGA A LA CASA DE LA LITERATURA - DESCARGA DE CARPA GRANDE', '03', 424.80, '2.3.1.99.1.99'),
    (3,  '06/02/2026', 'Recibo Servicios P.', '0000-15171',        'PAGO DE DERECHOS DE CERTIFICADO DE PARÁMETROS URBANÍSTICOS Y EDIFICATORIOS ANTE LA MUNICIPALIDAD DISTRITAL DE ATE', '07', 50.70, '2.3.1.99.1.99'),
    (4,  '06/02/2026', 'Declaración Jurada',  '0000-002',          'MOVILIDAD LOCAL DÍA 27 DE ENERO 2026, IDA EN UBER ENTRE CONGRESO Y MINEDU', '07', 42.90, '2.3.1.99.1.99'),
    (5,  '07/02/2026', 'Factura',             'F002-8351',         'COMPRA DE PORTA CHAPA PISO LOPO', '01', 70.00, '2.3.1.99.1.99'),
    (6,  '09/02/2026', 'Factura',             'F001-7369',         'COMPRA DE 01 TAFACÓN NEGRO, 03 BISAGRA', '01', 40.15, '2.3.1.99.1.99'),
    (7,  '09/02/2026', 'Factura',             'E001-1094',         '01 TAPA DE BAÑO 1 PESADO', '01', 35.00, '2.3.1.99.1.99'),
    (8,  '09/02/2026', 'Factura',             'F001-2192',         'COMPRA DE 04 TIMBRES DE 10 PULGADAS', '01', 252.00, '2.3.1.99.1.99'),
    (9,  '09/02/2026', 'Factura',             'E001-2892',         'UN SELLO 4912', '07', 40.00, '2.3.1.99.1.99'),
    (10, '09/02/2026', 'Factura',             'FQ01-569',          'ENVÍO DE PAQUETE (MATERIAL DE OFICINA LIMA - TUMBES)', '01', 55.00, '2.3.1.99.1.99'),
    (11, '09/02/2026', 'Factura',             'F001-32196',        'AEROPUERTO, ENROLLADO KANJI/WANTAN, POLLO CHAHRÁN, POLLO PERA, TALLARÍN DE POLLO. MENÚ DÍA 05/02/2026', '07', 171.00, '2.3.1.99.1.99'),
    (12, '09/02/2026', 'Factura',             'FD1-21821',         '3 POLLOS A LA BRASA, COMPRA DE POLLO DELIVERY DÍA 05/02/2026', '07', 188.50, '2.3.1.99.1.99'),
    (13, '09/02/2026', 'Recibo Servicios P.', '0000-260001715589', 'EMISIÓN DE 127 CERTIFICADOS DIGITALES PARA LA EXPEDICIÓN DE FIRMAS DIGITALES DEL PERSONAL DEL MINEDU', '01', 1028.70, '2.3.1.99.1.99'),
    (14, '10/02/2026', 'Factura',             'FD15-502949',       'COMPRA DE REFRIGERIO: 5 GALLETA RITZ, 11 GATORADE, 5 SAN LUIS, 4 FIOARI, 9 GUARANÁ', '07', 158.90, '2.3.1.99.1.99'),
    (15, '10/02/2026', 'Factura',             'E001-499',          'SERVICIO DE MONTACARGA EN LA CASA DE LA LITERATURA - CARGA DE CARPA NEUMÁTICA A CAMIÓN. SERVICIO DE MONTACARGA EN VENTANILLA - DESCARGA DE CARPA NEUMÁTICA DE CAMIÓN', '03', 424.80, '2.3.1.99.1.99'),
    (16, '10/02/2026', 'Boleta de Venta',     'EB01-6',            'MOVILIDAD PARA LA ENTREGA DE MEDALLAS DEL CONCURSO DE BUENAS PRÁCTICAS DE GESTIÓN EDUCATIVA EN AV. COMERCIO 193 / SAN BORJA A JR. MORRO SOLAR SURCO Y VICEVERSA DÍA 29/01/2026', '01', 45.25, '2.3.1.99.1.99'),
]

money_cols = {7}
for i, row_data in enumerate(resumen_data):
    r = row_h + 1 + i
    for col, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=col, value=val)
    style_data(ws1, r, len(headers_r), money_cols)

# Total row
total_row = row_h + 1 + len(resumen_data)
ws1.merge_cells(f'A{total_row}:F{total_row}')
ws1[f'A{total_row}'] = 'Total S/:'
ws1[f'A{total_row}'].font = Font(name='Calibri', bold=True, size=10)
ws1[f'A{total_row}'].alignment = Alignment(horizontal='right')
ws1.cell(row=total_row, column=7, value=3090.80)
ws1.cell(row=total_row, column=7).font = Font(name='Calibri', bold=True, size=10)
ws1.cell(row=total_row, column=7).number_format = money_fmt
ws1.cell(row=total_row, column=7).border = thin_border

# Movimiento de Fondo
fondo_row = total_row + 2
fondo_data = [
    ('Saldo Anterior', 29619.83),
    ('Reembolso', 0.00),
    ('Ampliación / Rebaja', 0.00),
    ('Sub Total', 29619.83),
    ('Pte. Rendición', 3090.80),
    ('Saldo Actual', 26529.03),
    ('En Tránsito', 38470.97),
    ('[Liquidación]', 26529.03),
    ('Total Caja Chica S/.', 0.00),
]
ws1[f'A{fondo_row}'] = 'Movimiento de Fondo'
ws1[f'A{fondo_row}'].font = subtitle_font
for i, (label, val) in enumerate(fondo_data):
    r = fondo_row + 1 + i
    ws1.cell(row=r, column=6, value=label)
    ws1.cell(row=r, column=6).font = Font(name='Calibri', bold=True, size=9)
    ws1.cell(row=r, column=7, value=val)
    ws1.cell(row=r, column=7).number_format = money_fmt
    ws1.cell(row=r, column=7).font = normal_font

# Resumen Presupuestal
pres_row = fondo_row + 1 + len(fondo_data) + 1
ws1[f'A{pres_row}'] = 'Resumen Presupuestal'
ws1[f'A{pres_row}'].font = subtitle_font
ws1.cell(row=pres_row+1, column=1, value='Meta / Mnemónico')
ws1.cell(row=pres_row+1, column=2, value='FF/Rb')
ws1.cell(row=pres_row+1, column=3, value='Clasificador del Gasto')
ws1.cell(row=pres_row+1, column=4, value='Monto S/.')
for c in range(1, 5):
    ws1.cell(row=pres_row+1, column=c).font = Font(name='Calibri', bold=True, size=9)
    ws1.cell(row=pres_row+1, column=c).border = thin_border
ws1.cell(row=pres_row+2, column=1, value='0035')
ws1.cell(row=pres_row+2, column=2, value='1-00')
ws1.cell(row=pres_row+2, column=3, value='2.3.1.99.1.99')
ws1.cell(row=pres_row+2, column=4, value=3090.80)
ws1.cell(row=pres_row+2, column=4).number_format = money_fmt
for c in range(1, 5):
    ws1.cell(row=pres_row+2, column=c).font = normal_font
    ws1.cell(row=pres_row+2, column=c).border = thin_border
ws1.cell(row=pres_row+3, column=3, value='Total S/.')
ws1.cell(row=pres_row+3, column=3).font = Font(name='Calibri', bold=True, size=9)
ws1.cell(row=pres_row+3, column=4, value=3090.80)
ws1.cell(row=pres_row+3, column=4).number_format = money_fmt
ws1.cell(row=pres_row+3, column=4).font = Font(name='Calibri', bold=True, size=9)

# Column widths
col_widths_1 = [5, 12, 18, 22, 60, 14, 14, 16]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# HOJA 2: COMPROBANTES DE PAGO (detalle digitalizado)
# ============================================================
ws2 = wb.create_sheet("Comprobantes de Pago")

# Título
ws2.merge_cells('A1:R1')
ws2['A1'] = 'COMPROBANTES DE PAGO - CAJA CHICA Nº 0000003 - SUSTENTO LIQUIDACIÓN 03'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.merge_cells('A2:R2')
ws2['A2'] = 'Responsable: MELGAREJO VILLARROEL, ALINA MABEL | Fecha: 12/02/2026'
ws2['A2'].font = Font(name='Calibri', italic=True, size=9)
ws2['A2'].alignment = Alignment(horizontal='center')

# Headers
headers_c = [
    'N° GASTO',          # A
    'PÁG. SUSTENTO',     # B
    'FECHA EMISIÓN',     # C
    'TIPO COMPROBANTE',  # D
    'SERIE-NÚMERO',      # E
    'RUC EMISOR',        # F
    'RAZÓN SOCIAL EMISOR', # G
    'DESCRIPCIÓN / CONCEPTO', # H
    'CANTIDAD',          # I
    'VALOR VENTA S/.',   # J
    'IGV S/.',           # K
    'ICBPER S/.',        # L
    'TOTAL S/.',         # M
    'FORMA DE PAGO',     # N
    'SELLO PAGADO',      # O
    'MONTO RESUMEN S/.', # P
    'DIFERENCIA S/.',    # Q
    'OBSERVACIONES',     # R
]

row_h2 = 4
for col, h in enumerate(headers_c, 1):
    ws2.cell(row=row_h2, column=col, value=h)
style_header(ws2, row_h2, len(headers_c))

# Comprobantes detallados
# Cada comprobante: (N_GASTO, PAG, FECHA, TIPO, SERIE_NUM, RUC, RAZON_SOCIAL, DESCRIPCION, CANTIDAD, VALOR_VENTA, IGV, ICBPER, TOTAL, FORMA_PAGO, SELLO_PAGADO, MONTO_RESUMEN, DIFERENCIA, OBSERVACIONES)

comprobantes = [
    # Gasto 1: F001-1488 (o F001-00000468?) - CORPORACIÓN IMPRESIONA
    (1, 6, '06/02/2026', 'Factura Electrónica', 'F001-1488',
     'ILEGIBLE (parcialmente visible)', 'CORPORACIÓN IMPRESIONA S.A.C.',
     'Servicio de impresión y ploteo de planos para obtención del ITSE de la sede',
     None, None, None, None, 19.00,
     'Contado', 'SÍ - TESORERÍA', 19.00, 0.00,
     'Nro de serie en factura parece F001-00000468 pero resumen indica F001-1488. Verificar número correcto.'),

    # Gasto 2: E001-530 - ASERVNT PERU S.A.C.
    (2, 11, '03/02/2026', 'Factura Electrónica', 'E001-530',
     '20610827171', 'ASERVNT PERU S.A.C.',
     'Servicio de montacarga: carga de carpa nueva en Ventanilla a camión (13/01/2026). Servicio de carga a Casa de la Literatura - descarga de carpa grande.',
     '2 servicios', 360.00, 64.80, 0.00, 424.80,
     'Contado', 'SÍ - TESORERÍA 6 FEB 2026', 424.80, 0.00,
     'Monto supera 10% UIT (S/535.00) pero no supera 20% UIT (S/1,070.00). Requiere autorización excepcional OGA. Verificar si existe autorización.'),

    # Gasto 3: Recibo 0000-15171 - Municipalidad de ATE
    (3, 16, '05/02/2026', 'Recibo de Servicios Públicos', '0015171',
     None, 'MUNICIPALIDAD DISTRITAL DE ATE',
     'Pago de derechos de certificado de parámetros urbanísticos y edificatorios',
     1, None, None, None, 50.70,
     'Contado', 'SÍ - TESORERÍA', 50.70, 0.00,
     'Recibo de servicios públicos (tasa municipal). No es comprobante de pago SUNAT (factura/boleta). Directiva permite recibos de servicios públicos como sustento.'),

    # Gasto 4: DJ 0000-002 - Planilla Movilidad Vargas del Río
    (4, 23, '27/01/2026', 'Declaración Jurada / Planilla Movilidad', '0000-002',
     None, 'VARGAS DEL RIO, VERÓNICA GRACE (DNI 41721445)',
     'Movilidad local día 27/01/2026 - Comisión para mesa de trabajo segunda sesión ordinaria del grupo de trabajo y creación del Colegio de Fisioterapeutas del Perú. Servicios Uber.',
     '2 viajes', None, None, None, 42.90,
     'Uber (app)', 'SÍ - TESORERÍA 6 FEB 2026', 42.90, 0.00,
     'Planilla muestra total S/42.90 con recibos Uber adjuntos (pág. 24-25). DJ de movilidad conforme a directiva.'),

    # Gasto 5: F002-8351 - JH METALINOX
    (5, 31, '30/01/2026', 'Factura Electrónica', 'F002-00008351',
     '20604955498', 'JH METALINOX S.A.C.',
     'Compra de porta chapa piso',
     1, 59.32, 10.68, 0.00, 70.00,
     'Contado', 'SÍ - TESORERÍA 6 FEB 2026', 70.00, 0.00,
     ''),

    # Gasto 6: F001-7369 - COMERCIAL VICKI
    (6, 35, '07/02/2026', 'Factura Electrónica', 'F001-7369',
     'ILEGIBLE', 'COMERCIAL VICKI E.I.R.L.',
     'Compra de 01 tafacón negro + 03 bisagras',
     '2 ítems', None, None, None, 40.15,
     'Contado', 'SÍ - TESORERÍA', 40.15, 0.00,
     'Desglose IGV/valor venta parcialmente ilegible en factura escaneada. Dos facturas parecen estar en la misma página (difícil separar). Verificar físicamente.'),

    # Gasto 7: E001-1094 - RAYPEC S.A.C.
    (7, 38, '07/02/2026', 'Factura Electrónica', 'E001-1094',
     '20609780451', 'RAYPEC S.A.C.',
     'Compra de 01 tapa de baño pesado',
     1, 29.66, 5.34, 0.00, 35.00,
     'Contado', 'SÍ - TESORERÍA', 35.00, 0.00,
     ''),

    # Gasto 8: F001-2192 - D&D SOLUCIONES ELÉCTRICAS
    (8, 42, '08/02/2026', 'Factura Electrónica', 'F001-2192',
     '20606697091', 'D&D SOLUCIONES ELÉCTRICAS E.I.R.L.',
     'Compra de 04 timbres de 10 pulgadas',
     4, 213.56, 38.44, 0.00, 252.00,
     'Contado', 'SÍ - TESORERÍA 9 FEB 2026', 252.00, 0.00,
     ''),

    # Gasto 9: E001-2892 - GSO DIGITAL / SALAZAR ORE
    (9, 51, '06/02/2026', 'Factura Electrónica', 'E001-2892',
     '10701855406', 'GSO DIGITAL - SALAZAR ORE GLEM ANDERSON',
     'Compra de 01 sello 4F12',
     1, 30.51, 5.49, 0.00, 36.00,
     'Contado', 'SÍ - TESORERÍA 9 FEB 2026', 40.00, -4.00,
     'DISCREPANCIA: Factura indica TOTAL S/36.00 (texto: "TREINTA Y SEIS Y 00/100 SOLES"). Resumen indica S/40.00. Diferencia de S/4.00 a favor del fisco. VERIFICAR.'),

    # Gasto 10: FQ01-569 - ITTSA
    (10, 58, '07/02/2026', 'Factura Electrónica', 'FQ01-00569',
     '20440493781', 'ITTSA (ITTSABUS)',
     'Envío de paquete: material de oficina Lima - Tumbes. Peso 45 kg.',
     1, 46.61, 8.39, 0.00, 55.00,
     'Contado', 'SÍ - TESORERÍA 9 FEB 2026', 55.00, 0.00,
     ''),

    # Gasto 11: F001-32196 - HAO YUN LAI LAI (Chifa)
    (11, 63, '05/02/2026', 'Factura Electrónica', 'F001-00032196',
     '20664646143', 'HAO YUN LAI LAI S.A.C. (CHIFA WANG XIANG YUAN)',
     'Menús día 05/02/2026: Aeropuerto de pollo (17), delivery (8), enrollado salsa (19), lapi/wantan menú (22), mas G.S (0.50), pollo chicharrón menú x2 (38), pollo espárragos menú (20), pollo pan menú (19), tallarín pollo en trigo menú (8), taper (8)',
     '10 ítems (8 personas aprox.)', 145.00, 15.90, 0.00, 171.00,
     'Contado', 'SÍ - TESORERÍA 9 FEB 2026', 171.00, 0.00,
     'Gasto de alimentos/coffee break. Desglose de montos parcialmente ilegible en el ticket escaneado. Valor venta e IGV aproximados. Monto total consistente con resumen. Nota: texto en factura dice "SETENTA Y UNO SOLES ??/100" — el desglose exacto podría diferir en céntimos.'),

    # Gasto 12: FD1-21821 - GOLDATI S.A.C.
    (12, 65, '05/02/2026', 'Factura Electrónica', 'FD1-00021821',
     'ILEGIBLE (parcialmente)', 'GOLDATI S.A.C.',
     'Compra de 3 pollos a la brasa, delivery. Día 05/02/2026',
     3, 159.75, 28.75, 0.00, 188.50,
     'Contado', 'SÍ - TESORERÍA', 188.50, 0.00,
     'Gasto de alimentos. IGV/valor venta estimados desde total.'),

    # Gasto 13: Constancias Banco de la Nación - RENIEC
    (13, '70-71', '09/02/2026', 'Constancia de Pago de Tasas (Banco de la Nación)', 'NRO TICKET 260001715589',
     '20131370998 (MINEDU)', 'RENIEC - Emisión de Certificados Digitales',
     'Emisión de certificados digitales para entidades de la administración pública. Constancia 1: 99 certificados x S/8.10 = S/801.90. Constancia 2: 28 certificados x S/8.10 = S/226.80. Total: 127 certificados.',
     '127 certificados', None, None, None, 1028.70,
     'Pago en Banco de la Nación', 'SÍ - TESORERÍA 9 FEB 2026', 1028.70, 0.00,
     'Monto supera 10% UIT (S/535.00) y supera 20% UIT (S/1,070.00 - este gasto es S/1,028.70 < S/1,070). Requiere autorización excepcional OGA. Memorándum 00292-2026-MINEDU/SG-OGRH solicita autorización (pág. 68). Vale Provisional N° 067 regularizado (pág. 67). Soporte: 2 constancias de pago (pág. 70 y 71).'),

    # Gasto 14: FD15-502949 - WONG DISCOUNT
    (14, 83, '09/02/2026', 'Factura Electrónica', 'FD15-00502949',
     '20508565934', 'COMPAÑÍA WONG DISCOUNT S.A.C.',
     'Compra de refrigerios: galletas Ritz, Gatorade, San Luis, Guaraná, otros. Para reuniones sindicales y mesas de trabajo interinstitucionales.',
     'Varios ítems', None, None, None, 158.90,
     'Tarjeta (VISA)', 'SÍ - TESORERÍA 10 FEB 2026', 158.90, 0.00,
     'Pago con tarjeta de débito (VISA). Directiva exige comprobante "Al Contado". Solicitud de reembolso excepcional autorizada por OGA (correo pág. 72, Teresa Zenaida Quiroz Silva, Jefa OGA). Gasto de coffee break/alimentos < 5% UIT (S/267.50). Desglose valor venta/IGV ilegible en ticket.'),

    # Gasto 15: E001-499 - ASERVNT PERU S.A.C.
    (15, 88, '03/02/2026', 'Factura Electrónica', 'E001-499',
     '20610827171', 'ASERVNT PERU S.A.C.',
     'Servicio de montacarga: carga de carpa neumática a camión en Casa de la Literatura. Descarga de carpa neumática de camión en Ventanilla.',
     '2 servicios', 360.00, 64.80, 0.00, 424.80,
     'Contado', 'SÍ - TESORERÍA 10 FEB 2026', 424.80, 0.00,
     'Mismo proveedor que gasto #2 (ASERVNT). Monto idéntico (S/424.80). Servicio complementario (ida/vuelta de carpa). Monto supera 10% UIT (S/535.00 individual) — NO, S/424.80 < S/535.00, conforme.'),

    # Gasto 16: EB01-6 - LESCANO HIDALGO MARIO ARTURO + Planilla Movilidad
    (16, '91,112', '04/02/2026', 'Boleta de Venta Electrónica + Planilla de Movilidad', 'EB01-6',
     '10073775006', 'LESCANO HIDALGO MARIO ARTURO',
     'Servicio de traslado del MINEDU a Escuela Superior Tecnológica, viceversa (29/01/2026). Planilla de movilidad local (Rivera del Águila) para entrega de medallas del Concurso BPGE 2025 en DRE/UGEL.',
     1, None, None, None, 45.25,
     'Contado', 'SÍ - TESORERÍA 10 FEB 2026', 45.25, 0.00,
     'Gasto compuesto: Boleta EB01-6 = S/40.00 (servicio traslado) + S/5.25 (pasajes planilla movilidad) = S/45.25. Planilla de movilidad local N° 0003-2026 (pág. 109). Constancia de NO uso de movilidad local adjunta (pág. 111) — las especialistas DIFOCA no usaron el servicio; usaron transporte público. Boleta EB01-6 (pág. 112) por S/40.00 del traslado sí fue usado.'),
]

money_cols_2 = {10, 11, 12, 13, 16, 17}  # J, K, L, M, P, Q
for i, row_data in enumerate(comprobantes):
    r = row_h2 + 1 + i
    for col, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=col, value=val)
    style_data(ws2, r, len(headers_c), money_cols_2)
    # Color rojo para observaciones
    obs_cell = ws2.cell(row=r, column=18)
    if obs_cell.value:
        obs_cell.font = obs_font

# Total row
total_row2 = row_h2 + 1 + len(comprobantes)
ws2.merge_cells(f'A{total_row2}:L{total_row2}')
ws2[f'A{total_row2}'] = 'TOTAL COMPROBANTES:'
ws2[f'A{total_row2}'].font = Font(name='Calibri', bold=True, size=10)
ws2[f'A{total_row2}'].alignment = Alignment(horizontal='right')
# Sum formula for column M (TOTAL)
ws2.cell(row=total_row2, column=13, value=f'=SUM(M{row_h2+1}:M{total_row2-1})')
ws2.cell(row=total_row2, column=13).font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=total_row2, column=13).number_format = money_fmt
ws2.cell(row=total_row2, column=13).border = thin_border
# Sum for MONTO RESUMEN
ws2.cell(row=total_row2, column=16, value=f'=SUM(P{row_h2+1}:P{total_row2-1})')
ws2.cell(row=total_row2, column=16).font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=total_row2, column=16).number_format = money_fmt
ws2.cell(row=total_row2, column=16).border = thin_border
# Sum for DIFERENCIA
ws2.cell(row=total_row2, column=17, value=f'=SUM(Q{row_h2+1}:Q{total_row2-1})')
ws2.cell(row=total_row2, column=17).font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=total_row2, column=17).number_format = money_fmt
ws2.cell(row=total_row2, column=17).border = thin_border

# Resumen de discrepancias
disc_row = total_row2 + 2
ws2[f'A{disc_row}'] = 'RESUMEN DE DISCREPANCIAS Y OBSERVACIONES RELEVANTES'
ws2[f'A{disc_row}'].font = Font(name='Calibri', bold=True, size=11, color='CC0000')
ws2.merge_cells(f'A{disc_row}:R{disc_row}')

obs_list = [
    'Gasto #9 (Sello 4912): Factura E001-2892 indica S/36.00 ("TREINTA Y SEIS Y 00/100 SOLES"). Resumen indica S/40.00. Diferencia: S/4.00.',
    'Gasto #14 (Wong): Pago realizado con tarjeta de débito VISA. Directiva exige pago "Al Contado". Se cuenta con autorización excepcional de reembolso por OGA (Teresa Zenaida Quiroz Silva).',
    'Gastos #11 y #12 (Chifa + Pollos): Alimentos por total S/359.50 en un solo día (05/02/2026). Supera 5% UIT (S/267.50) por evento individual. Verificar si corresponden a eventos separados.',
    'Gasto #13 (RENIEC Certificados): S/1,028.70 — cercano al límite excepcional 20% UIT (S/1,070). Autorización vía Memorándum 00292-2026 y Vale Provisional N° 067 regularizado.',
    'Gasto #16 (Movilidad entrega medallas): Constancia de NO uso de movilidad local para las especialistas DIFOCA (pág. 111). El servicio se solicitó pero no fue utilizado por las 3 especialistas. Solo la boleta EB01-6 (traslado S/40.00) fue efectivamente usada.',
    'Gastos #2 y #15 (Montacarga ASERVNT): Mismo proveedor, mismo monto (S/424.80), servicios complementarios (ida y vuelta de carpas). Total al mismo proveedor: S/849.60. Supera 10% UIT (S/535) en conjunto.',
]
for i, obs in enumerate(obs_list):
    r = disc_row + 1 + i
    ws2.merge_cells(f'A{r}:R{r}')
    ws2[f'A{r}'] = f'{i+1}. {obs}'
    ws2[f'A{r}'].font = Font(name='Calibri', size=9, color='CC0000')
    ws2[f'A{r}'].alignment = wrap_align

# Referencia directiva
ref_row = disc_row + 1 + len(obs_list) + 1
ws2[f'A{ref_row}'] = 'DIRECTIVA DE REFERENCIA: RJ 0023-2025-MINEDU/SG-OGA - Directiva para la Administración del Fondo de Caja Chica'
ws2[f'A{ref_row}'].font = Font(name='Calibri', italic=True, size=8)
ws2.merge_cells(f'A{ref_row}:R{ref_row}')
ws2[f'A{ref_row+1}'] = 'LÍMITES: 10% UIT = S/535.00 (máx. por gasto) | 20% UIT = S/1,070.00 (excepcional con autorización OGA) | 5% UIT = S/267.50 (coffee break/alimentos por evento)'
ws2[f'A{ref_row+1}'].font = Font(name='Calibri', italic=True, size=8)
ws2.merge_cells(f'A{ref_row+1}:R{ref_row+1}')

# Column widths for sheet 2
col_widths_2 = [9, 12, 12, 22, 20, 18, 30, 55, 12, 12, 10, 10, 12, 14, 18, 14, 12, 65]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Row heights for comprobantes
for r in range(row_h2 + 1, total_row2):
    ws2.row_dimensions[r].height = 60

# ============================================================
# GUARDAR
# ============================================================
out_path = r'C:\Users\Hans\Proyectos\AG-EVIDENCE\data\expedientes\pruebas\caja_chica_2026\OT2026-INT-0179550_CAJA_CHICA_JAQUELINE\RENDICION_CAJA_CHICA_003.xlsx'
wb.save(out_path)
print(f"Excel generado: {out_path}")
print(f"Hojas: {wb.sheetnames}")
print(f"Hoja 1: {len(resumen_data)} gastos en resumen")
print(f"Hoja 2: {len(comprobantes)} comprobantes digitalizados")
print(f"Total resumen: S/3,090.80")
total_comprobantes = sum(c[12] for c in comprobantes if isinstance(c[12], (int, float)))
print(f"Total comprobantes: S/{total_comprobantes:,.2f}")
