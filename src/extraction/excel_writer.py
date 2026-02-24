# -*- coding: utf-8 -*-
"""
Excel Writer — Hoja DIAGNOSTICO
================================
Tarea #20 del Plan de Desarrollo (Fase 2: Contrato + Router)

Genera la hoja DIAGNOSTICO del Excel de rendición con:
  - Banner resumen del expediente (SINAD, status, confianza, acción)
  - Secciones del diagnóstico del ConfidenceRouter (6 pasos)
  - Detalle por campo: nombre, confianza, status con color
  - Métricas globales del expediente

Consume:
  - DiagnosticoExpediente (de IntegrityCheckpoint)
  - ResultadoRouter (para métricas detalladas por campo)
  - DecisionCheckpoint (para decisión + diagnóstico integrado)

Produce:
  - Hoja "DIAGNOSTICO" en un Workbook openpyxl (existente o nuevo)

Diseño:
  - Recibe Workbook ya existente; solo agrega hoja DIAGNOSTICO.
  - Colores semáforo Excel estándar:
      Verde:   #C6EFCE fondo + #006100 texto = LEGIBLE / OK
      Amarillo: #FFEB9C fondo + #9C5700 texto = INCOMPLETO / WARNING
      Rojo:    #FFC7CE fondo + #9C0006 texto = ILEGIBLE / CRITICAL / ABSTENCION
      Gris:    #D9D9D9 fondo + #404040 texto = SKIP / sin evaluar
  - Compatible con openpyxl ≥3.0. Sin dependencia de pandas.
  - NO crea Workbook nuevo: se integra en el Excel del pipeline.

Gobernanza:
  - Regla 6: Excel solo consume contrato (DiagnosticoExpediente)
  - Regla 7: Cada fila lleva fuente, confianza, motor
  - Art. 3: Anti-alucinación (muestra datos tal cual, sin inferir)
  - Art. 17: Trazabilidad (timestamp, versión del router)

Uso:
    from src.extraction.excel_writer import EscritorDiagnostico

    escritor = EscritorDiagnostico()
    escritor.escribir_hoja_diagnostico(workbook, decision_checkpoint)

    # O directamente desde DiagnosticoExpediente + ResultadoRouter:
    escritor.escribir_desde_componentes(workbook, diagnostico, resultado)

Versión: 1.0.0
Fecha: 2026-02-23
"""

import sys
import os
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple

# Asegurar que el directorio raíz del proyecto esté en el path
_PROJECT_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side,
    )
    from openpyxl.utils import get_column_letter

    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

from src.extraction.confidence_router import (
    DiagnosticoExpediente,
    SeccionDiagnostico,
    ResultadoRouter,
    DecisionCheckpoint,
    VERSION_ROUTER,
)
from src.extraction.abstencion import (
    CampoExtraido,
    EvidenceStatus,
    ResultadoAbstencion,
)
from src.extraction.expediente_contract import (
    IntegridadStatus,
    ConfianzaGlobal,
)


# ==============================================================================
# CONSTANTES
# ==============================================================================

VERSION_EXCEL_WRITER = "1.0.0"
"""Versión del módulo excel_writer."""

NOMBRE_HOJA = "DIAGNOSTICO"
"""Nombre de la hoja que se crea en el Workbook."""


# ==============================================================================
# PALETA DE COLORES SEMÁFORO
# ==============================================================================

@dataclass(frozen=True)
class ColorSemaforo:
    """Un color semáforo con fondo y texto."""
    fondo: str
    texto: str
    nombre: str


# Colores estándar Excel para formato condicional
VERDE = ColorSemaforo(fondo="C6EFCE", texto="006100", nombre="verde")
AMARILLO = ColorSemaforo(fondo="FFEB9C", texto="9C5700", nombre="amarillo")
ROJO = ColorSemaforo(fondo="FFC7CE", texto="9C0006", nombre="rojo")
GRIS = ColorSemaforo(fondo="D9D9D9", texto="404040", nombre="gris")
BLANCO = ColorSemaforo(fondo="FFFFFF", texto="000000", nombre="blanco")
AZUL_HEADER = ColorSemaforo(fondo="4472C4", texto="FFFFFF", nombre="azul_header")
AZUL_BANNER = ColorSemaforo(fondo="D6E4F0", texto="1F3864", nombre="azul_banner")


def _color_por_status(status: str) -> ColorSemaforo:
    """
    Mapea un string de status al color semáforo correspondiente.

    Args:
        status: Uno de OK, WARNING, CRITICAL, SKIP, LEGIBLE,
                INCOMPLETO, ILEGIBLE.

    Returns:
        ColorSemaforo correspondiente.
    """
    mapa = {
        "OK": VERDE,
        "LEGIBLE": VERDE,
        "WARNING": AMARILLO,
        "INCOMPLETO": AMARILLO,
        "CRITICAL": ROJO,
        "ILEGIBLE": ROJO,
        "ABSTENCION": ROJO,
        "SKIP": GRIS,
    }
    return mapa.get(status.upper(), GRIS)


def _color_por_confianza(confianza: float) -> ColorSemaforo:
    """
    Mapea un nivel de confianza (0.0-1.0) al color semáforo.

    Umbrales:
      - >= 0.70: Verde (LEGIBLE)
      - >= 0.40: Amarillo (INCOMPLETO)
      - < 0.40: Rojo (ILEGIBLE/ABSTENCION)

    Args:
        confianza: Nivel de confianza 0.0-1.0.

    Returns:
        ColorSemaforo correspondiente.
    """
    if confianza >= 0.70:
        return VERDE
    elif confianza >= 0.40:
        return AMARILLO
    else:
        return ROJO


# ==============================================================================
# ESTILOS OPENPYXL
# ==============================================================================

def _crear_fill(color: ColorSemaforo) -> "PatternFill":
    """Crea PatternFill de openpyxl desde un ColorSemaforo."""
    if not OPENPYXL_DISPONIBLE:
        raise ImportError("openpyxl no está instalado")
    return PatternFill(start_color=color.fondo, end_color=color.fondo, fill_type="solid")


def _crear_font(color: ColorSemaforo, bold: bool = False, size: int = 11) -> "Font":
    """Crea Font de openpyxl desde un ColorSemaforo."""
    if not OPENPYXL_DISPONIBLE:
        raise ImportError("openpyxl no está instalado")
    return Font(name="Calibri", size=size, bold=bold, color=color.texto)


def _borde_delgado() -> "Border":
    """Borde delgado estándar para celdas."""
    if not OPENPYXL_DISPONIBLE:
        raise ImportError("openpyxl no está instalado")
    lado = Side(style="thin", color="B4B4B4")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


# ==============================================================================
# CLASE PRINCIPAL: EscritorDiagnostico
# ==============================================================================

class EscritorDiagnostico:
    """
    Escritor de la hoja DIAGNOSTICO para Excel de rendición.

    Tarea #20 del Plan de Desarrollo (Fase 2).

    Genera una hoja Excel con:
      1. Banner resumen (SINAD, status global, confianza, acción)
      2. Tabla de secciones de diagnóstico (6 filas del router)
      3. Tabla de detalle por campo (nombre, confianza, color)
      4. Métricas globales pie de página

    Integración:
      - Recibe un Workbook ya existente (de las otras hojas del Excel)
      - Solo agrega la hoja DIAGNOSTICO
      - Si la hoja ya existe, la reemplaza

    Uso:
        escritor = EscritorDiagnostico()
        escritor.escribir_hoja_diagnostico(wb, decision)
    """

    def __init__(self, nombre_hoja: str = NOMBRE_HOJA):
        """
        Inicializa el escritor.

        Args:
            nombre_hoja: Nombre de la hoja a crear. Default: "DIAGNOSTICO".
        """
        if not OPENPYXL_DISPONIBLE:
            raise ImportError(
                "openpyxl no está instalado. "
                "Instalar con: pip install openpyxl"
            )
        self.nombre_hoja = nombre_hoja

    # ------------------------------------------------------------------
    # INTERFAZ PÚBLICA
    # ------------------------------------------------------------------

    def escribir_hoja_diagnostico(
        self,
        wb: "Workbook",
        decision: DecisionCheckpoint,
    ) -> "Worksheet":
        """
        Escribe la hoja DIAGNOSTICO completa desde un DecisionCheckpoint.

        Este es el método principal. Extrae DiagnosticoExpediente y
        ResultadoRouter del DecisionCheckpoint y delega a las secciones.

        Args:
            wb: Workbook de openpyxl (ya debe tener las otras hojas).
            decision: DecisionCheckpoint del IntegrityCheckpoint.

        Returns:
            Worksheet creado con la hoja DIAGNOSTICO.

        Raises:
            ValueError: Si el DecisionCheckpoint no tiene diagnóstico.
        """
        if decision.diagnostico is None:
            raise ValueError(
                "DecisionCheckpoint no tiene diagnóstico. "
                "Ejecutar IntegrityCheckpoint.evaluar() primero."
            )

        ws = self._preparar_hoja(wb)

        fila = 1

        # Banner resumen
        fila = self._escribir_banner(ws, decision, fila)

        # Espacio
        fila += 1

        # Tabla de secciones
        fila = self._escribir_secciones(ws, decision.diagnostico, fila)

        # Espacio
        fila += 1

        # Tabla de detalle por campo
        if decision.resultado:
            fila = self._escribir_detalle_campos(
                ws, decision.resultado, fila
            )
            fila += 1

        # Pie de página con métricas
        fila = self._escribir_pie(ws, decision, fila)

        # Ajustar anchos de columna
        self._ajustar_anchos(ws)

        return ws

    def escribir_desde_componentes(
        self,
        wb: "Workbook",
        diagnostico: DiagnosticoExpediente,
        resultado: Optional[ResultadoRouter] = None,
        accion: str = "",
    ) -> "Worksheet":
        """
        Escribe la hoja DIAGNOSTICO desde componentes individuales.

        Útil cuando no se tiene un DecisionCheckpoint completo
        (por ejemplo, en tests o procesamiento parcial).

        Args:
            wb: Workbook de openpyxl.
            diagnostico: DiagnosticoExpediente con las secciones.
            resultado: ResultadoRouter opcional (para detalle por campo).
            accion: Acción del checkpoint (CONTINUAR/ALERTAS/DETENER).

        Returns:
            Worksheet creado.
        """
        # Construir DecisionCheckpoint parcial
        decision = DecisionCheckpoint(
            accion=accion or "N/A",
            resultado=resultado,
            diagnostico=diagnostico,
            timestamp=diagnostico.timestamp,
        )
        return self.escribir_hoja_diagnostico(wb, decision)

    # ------------------------------------------------------------------
    # SECCIÓN 1: BANNER RESUMEN
    # ------------------------------------------------------------------

    def _escribir_banner(
        self,
        ws: "Worksheet",
        decision: DecisionCheckpoint,
        fila_inicio: int,
    ) -> int:
        """
        Escribe el banner resumen en la parte superior de la hoja.

        Contenido:
          Fila 1: "DIAGNÓSTICO DE EXPEDIENTE" (título)
          Fila 2: SINAD | Status | Confianza | Acción
          Fila 3: Timestamp | Versión Router

        Returns:
            Siguiente fila disponible.
        """
        fila = fila_inicio
        diag = decision.diagnostico
        resultado = decision.resultado

        # --- Título ---
        ws.merge_cells(
            start_row=fila, start_column=1,
            end_row=fila, end_column=6,
        )
        celda_titulo = ws.cell(row=fila, column=1)
        celda_titulo.value = "DIAGNÓSTICO DE EXPEDIENTE"
        celda_titulo.font = _crear_font(AZUL_BANNER, bold=True, size=14)
        celda_titulo.fill = _crear_fill(AZUL_BANNER)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = 30
        # Aplicar fill a todas las celdas del merge
        for col in range(2, 7):
            ws.cell(row=fila, column=col).fill = _crear_fill(AZUL_BANNER)
        fila += 1

        # --- Fila de datos principales ---
        labels = ["SINAD", "Status", "Confianza", "Acción", "Campos OK", "Abstención"]
        valores = [
            diag.sinad if diag else "",
            resultado.status.value if resultado else "N/A",
            resultado.confianza_global.value if resultado else "N/A",
            decision.accion,
            (
                f"{resultado.campos_legibles}/{resultado.campos_evaluados}"
                if resultado and resultado.campos_evaluados > 0
                else "N/A"
            ),
            (
                f"{resultado.tasa_abstencion:.1%}"
                if resultado
                else "N/A"
            ),
        ]

        # Labels
        for col_idx, label in enumerate(labels, 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.value = label
            celda.font = _crear_font(AZUL_HEADER, bold=True, size=10)
            celda.fill = _crear_fill(AZUL_HEADER)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()
        fila += 1

        # Valores con color por status
        for col_idx, valor in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.value = valor
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()

            # Color especial para columna Status
            if col_idx == 2 and resultado:
                color = _color_por_status(resultado.status.value)
                celda.fill = _crear_fill(color)
                celda.font = _crear_font(color, bold=True)
            # Color especial para columna Acción
            elif col_idx == 4:
                color_accion = {
                    "CONTINUAR": VERDE,
                    "CONTINUAR_CON_ALERTAS": AMARILLO,
                    "DETENER": ROJO,
                }.get(decision.accion, GRIS)
                celda.fill = _crear_fill(color_accion)
                celda.font = _crear_font(color_accion, bold=True)
            else:
                celda.font = Font(name="Calibri", size=11, bold=True)
        fila += 1

        # --- Fila metadata ---
        ws.cell(row=fila, column=1).value = "Generado:"
        ws.cell(row=fila, column=1).font = Font(
            name="Calibri", size=9, italic=True, color="808080"
        )
        ws.cell(row=fila, column=2).value = (
            diag.timestamp[:19] if diag and diag.timestamp else ""
        )
        ws.cell(row=fila, column=2).font = Font(
            name="Calibri", size=9, color="808080"
        )
        ws.cell(row=fila, column=3).value = f"Router v{diag.version_router}" if diag else ""
        ws.cell(row=fila, column=3).font = Font(
            name="Calibri", size=9, color="808080"
        )
        ws.cell(row=fila, column=4).value = f"Writer v{VERSION_EXCEL_WRITER}"
        ws.cell(row=fila, column=4).font = Font(
            name="Calibri", size=9, color="808080"
        )
        fila += 1

        return fila

    # ------------------------------------------------------------------
    # SECCIÓN 2: TABLA DE SECCIONES DEL DIAGNÓSTICO
    # ------------------------------------------------------------------

    def _escribir_secciones(
        self,
        ws: "Worksheet",
        diagnostico: DiagnosticoExpediente,
        fila_inicio: int,
    ) -> int:
        """
        Escribe la tabla de secciones del diagnóstico.

        Columnas: Sección | Status | Mensaje | Detalle

        Cada fila corresponde a una SeccionDiagnostico del router:
          1. evaluacion_campos
          2. enforcement
          3. completitud
          4. unicidad
          5. aritmetica
          6. decision

        Returns:
            Siguiente fila disponible.
        """
        fila = fila_inicio

        # Subtítulo
        ws.merge_cells(
            start_row=fila, start_column=1,
            end_row=fila, end_column=6,
        )
        celda = ws.cell(row=fila, column=1)
        celda.value = "Secciones del Diagnóstico"
        celda.font = Font(name="Calibri", size=12, bold=True, color="1F3864")
        for col in range(2, 7):
            ws.cell(row=fila, column=col)  # touch para merge
        fila += 1

        # Headers
        headers = ["Sección", "Status", "Mensaje", "Detalle 1", "Detalle 2", "Detalle 3"]
        for col_idx, h in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.value = h
            celda.font = _crear_font(AZUL_HEADER, bold=True, size=10)
            celda.fill = _crear_fill(AZUL_HEADER)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()
        fila += 1

        # Filas de secciones
        rows = diagnostico.to_rows()

        # Agrupar rows por sección para evitar duplicación
        secciones_vistas: Dict[str, bool] = {}
        for seccion in diagnostico.secciones:
            if seccion.nombre in secciones_vistas:
                continue
            secciones_vistas[seccion.nombre] = True

            color = _color_por_status(seccion.status)

            # Nombre de sección legible
            nombres_legibles = {
                "evaluacion_campos": "Evaluación de Campos",
                "enforcement": "Estándar Probatorio",
                "completitud": "Completitud",
                "unicidad": "Unicidad",
                "aritmetica": "Aritmética (Grupo J)",
                "decision": "Decisión Final",
            }
            nombre_mostrar = nombres_legibles.get(seccion.nombre, seccion.nombre)

            # Columna 1: Nombre
            celda = ws.cell(row=fila, column=1)
            celda.value = nombre_mostrar
            celda.font = Font(name="Calibri", size=11, bold=True)
            celda.border = _borde_delgado()

            # Columna 2: Status con color
            celda = ws.cell(row=fila, column=2)
            celda.value = seccion.status
            celda.fill = _crear_fill(color)
            celda.font = _crear_font(color, bold=True)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()

            # Columna 3: Mensaje
            celda = ws.cell(row=fila, column=3)
            celda.value = seccion.mensaje
            celda.font = Font(name="Calibri", size=10)
            celda.border = _borde_delgado()
            celda.alignment = Alignment(wrap_text=True)

            # Columnas 4-6: Primeros 3 detalles
            for d_idx, detalle in enumerate(seccion.detalles[:3]):
                celda = ws.cell(row=fila, column=4 + d_idx)
                celda.value = detalle
                celda.font = Font(name="Calibri", size=9, color="404040")
                celda.border = _borde_delgado()
                celda.alignment = Alignment(wrap_text=True)

            # Si hay más de 3 detalles, indicar
            if len(seccion.detalles) > 3:
                celda = ws.cell(row=fila, column=6)
                celda.value = f"... +{len(seccion.detalles) - 2} más"
                celda.font = Font(name="Calibri", size=9, italic=True, color="808080")

            fila += 1

        return fila

    # ------------------------------------------------------------------
    # SECCIÓN 3: DETALLE POR CAMPO (confianza + color)
    # ------------------------------------------------------------------

    def _escribir_detalle_campos(
        self,
        ws: "Worksheet",
        resultado: ResultadoRouter,
        fila_inicio: int,
    ) -> int:
        """
        Escribe tabla detallada por campo con confianza y color.

        Columnas: Campo | Valor | Confianza | Status | Motor | Archivo

        Cada fila es un CampoExtraido del ResultadoRouter, con su
        color semáforo basado en confianza:
          - Verde: confianza >= 0.70 (LEGIBLE)
          - Amarillo: 0.40 <= confianza < 0.70 (INCOMPLETO)
          - Rojo: confianza < 0.40 (ILEGIBLE/ABSTENCION)

        Returns:
            Siguiente fila disponible.
        """
        fila = fila_inicio

        # Subtítulo
        ws.merge_cells(
            start_row=fila, start_column=1,
            end_row=fila, end_column=6,
        )
        celda = ws.cell(row=fila, column=1)
        celda.value = "Detalle por Campo"
        celda.font = Font(name="Calibri", size=12, bold=True, color="1F3864")
        for col in range(2, 7):
            ws.cell(row=fila, column=col)
        fila += 1

        # Headers
        headers = ["Campo", "Valor", "Confianza", "Status", "Motor", "Archivo:Pág"]
        for col_idx, h in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.value = h
            celda.font = _crear_font(AZUL_HEADER, bold=True, size=10)
            celda.fill = _crear_fill(AZUL_HEADER)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()
        fila += 1

        # Recoger campos desde resultados de abstención
        campos_escritos = 0

        for res_abs in resultado.resultados_abstencion:
            campo = res_abs.campo
            status = campo.clasificar_status()
            color = _color_por_status(status.value)

            # Col 1: Nombre del campo
            celda = ws.cell(row=fila, column=1)
            celda.value = campo.nombre_campo
            celda.font = Font(name="Calibri", size=10)
            celda.border = _borde_delgado()

            # Col 2: Valor (truncado a 60 chars)
            celda = ws.cell(row=fila, column=2)
            valor_str = str(campo.valor) if campo.valor is not None else "(abstención)"
            celda.value = valor_str[:60] + ("..." if len(valor_str) > 60 else "")
            celda.font = Font(name="Calibri", size=10)
            celda.border = _borde_delgado()
            celda.alignment = Alignment(wrap_text=True)
            if campo.valor is None:
                celda.font = Font(
                    name="Calibri", size=10, italic=True, color="9C0006"
                )

            # Col 3: Confianza con color
            celda = ws.cell(row=fila, column=3)
            celda.value = f"{campo.confianza:.0%}"
            color_conf = _color_por_confianza(campo.confianza)
            celda.fill = _crear_fill(color_conf)
            celda.font = _crear_font(color_conf, bold=True)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()

            # Col 4: Status con color
            celda = ws.cell(row=fila, column=4)
            celda.value = status.value
            celda.fill = _crear_fill(color)
            celda.font = _crear_font(color, bold=True)
            celda.alignment = Alignment(horizontal="center")
            celda.border = _borde_delgado()

            # Col 5: Motor
            celda = ws.cell(row=fila, column=5)
            motor = campo.motor_ocr or campo.metodo.value
            celda.value = motor
            celda.font = Font(name="Calibri", size=9, color="404040")
            celda.border = _borde_delgado()
            celda.alignment = Alignment(horizontal="center")

            # Col 6: Archivo:Página
            celda = ws.cell(row=fila, column=6)
            archivo_corto = os.path.basename(campo.archivo) if campo.archivo else ""
            celda.value = f"{archivo_corto}:{campo.pagina}" if campo.pagina > 0 else archivo_corto
            celda.font = Font(name="Calibri", size=9, color="404040")
            celda.border = _borde_delgado()

            fila += 1
            campos_escritos += 1

        # Si no hay campos de abstención, mostrar mensaje
        if campos_escritos == 0:
            celda = ws.cell(row=fila, column=1)
            celda.value = "(Sin campos evaluados por AbstencionPolicy)"
            celda.font = Font(name="Calibri", size=10, italic=True, color="808080")
            ws.merge_cells(
                start_row=fila, start_column=1,
                end_row=fila, end_column=6,
            )
            fila += 1

        return fila

    # ------------------------------------------------------------------
    # SECCIÓN 4: PIE DE PÁGINA CON MÉTRICAS
    # ------------------------------------------------------------------

    def _escribir_pie(
        self,
        ws: "Worksheet",
        decision: DecisionCheckpoint,
        fila_inicio: int,
    ) -> int:
        """
        Escribe pie de página con métricas globales y alertas.

        Returns:
            Siguiente fila disponible.
        """
        fila = fila_inicio
        resultado = decision.resultado

        # Línea separadora
        for col in range(1, 7):
            celda = ws.cell(row=fila, column=col)
            celda.border = Border(top=Side(style="medium", color="4472C4"))
        fila += 1

        # Alertas (si hay)
        if resultado and resultado.alertas:
            celda = ws.cell(row=fila, column=1)
            celda.value = "⚠ ALERTAS:"
            celda.font = Font(name="Calibri", size=11, bold=True, color="9C5700")
            fila += 1

            for alerta in resultado.alertas:
                celda = ws.cell(row=fila, column=1)
                celda.value = f"  • {alerta}"
                celda.font = Font(name="Calibri", size=10, color="9C5700")
                ws.merge_cells(
                    start_row=fila, start_column=1,
                    end_row=fila, end_column=6,
                )
                fila += 1

            fila += 1

        # Métricas resumen
        metricas = self._recopilar_metricas(decision)
        if metricas:
            celda = ws.cell(row=fila, column=1)
            celda.value = "Métricas:"
            celda.font = Font(name="Calibri", size=10, bold=True, color="404040")
            fila += 1

            for clave, valor in metricas.items():
                celda = ws.cell(row=fila, column=1)
                celda.value = f"  {clave}:"
                celda.font = Font(name="Calibri", size=9, color="808080")
                celda = ws.cell(row=fila, column=2)
                celda.value = str(valor)
                celda.font = Font(name="Calibri", size=9, color="404040")
                fila += 1

        return fila

    # ------------------------------------------------------------------
    # UTILIDADES INTERNAS
    # ------------------------------------------------------------------

    def _preparar_hoja(self, wb: "Workbook") -> "Worksheet":
        """
        Prepara la hoja DIAGNOSTICO en el Workbook.

        Si ya existe, la elimina y crea una nueva.

        Returns:
            Worksheet limpia lista para escribir.
        """
        if self.nombre_hoja in wb.sheetnames:
            del wb[self.nombre_hoja]

        ws = wb.create_sheet(title=self.nombre_hoja)

        # Configuración de impresión
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        return ws

    def _ajustar_anchos(self, ws: "Worksheet") -> None:
        """
        Ajusta anchos de columna según contenido.

        Anchos fijos optimizados para la estructura de 6 columnas.
        """
        anchos = {
            1: 28,   # Campo / Sección
            2: 25,   # Valor / Status
            3: 14,   # Confianza / Mensaje
            4: 14,   # Status / Detalle 1
            5: 16,   # Motor / Detalle 2
            6: 20,   # Archivo:Pág / Detalle 3
        }
        for col, ancho in anchos.items():
            ws.column_dimensions[get_column_letter(col)].width = ancho

    @staticmethod
    def _recopilar_metricas(decision: DecisionCheckpoint) -> Dict[str, str]:
        """
        Recopila métricas globales para el pie de página.

        Returns:
            Dict con métricas legibles.
        """
        metricas: Dict[str, str] = {}
        r = decision.resultado

        if r:
            metricas["Campos evaluados"] = str(r.campos_evaluados)
            metricas["Campos legibles"] = str(r.campos_legibles)
            metricas["Campos incompletos"] = str(r.campos_incompletos)
            metricas["Campos abstenidos"] = str(r.campos_abstenidos)
            metricas["Tasa abstención"] = f"{r.tasa_abstencion:.1%}"
            metricas["Obs. degradadas"] = str(len(r.observaciones_degradadas))
            metricas["Errores aritmét."] = str(len(r.errores_aritmeticos))
            metricas["Problemas completitud"] = str(len(r.problemas_completitud))
            metricas["Duplicados"] = str(len(r.comprobantes_duplicados))

        metricas["Versión router"] = decision.diagnostico.version_router if decision.diagnostico else "N/A"
        metricas["Versión writer"] = VERSION_EXCEL_WRITER

        return metricas


# ==============================================================================
# FUNCIÓN DE CONVENIENCIA
# ==============================================================================

def escribir_diagnostico(
    wb: "Workbook",
    decision: DecisionCheckpoint,
    nombre_hoja: str = NOMBRE_HOJA,
) -> "Worksheet":
    """
    Función de conveniencia para escribir la hoja DIAGNOSTICO.

    Crea un EscritorDiagnostico y ejecuta la escritura.

    Args:
        wb: Workbook de openpyxl.
        decision: DecisionCheckpoint del IntegrityCheckpoint.
        nombre_hoja: Nombre de la hoja. Default: "DIAGNOSTICO".

    Returns:
        Worksheet creado.

    Raises:
        ImportError: Si openpyxl no está instalado.
        ValueError: Si el DecisionCheckpoint no tiene diagnóstico.
    """
    escritor = EscritorDiagnostico(nombre_hoja=nombre_hoja)
    return escritor.escribir_hoja_diagnostico(wb, decision)
