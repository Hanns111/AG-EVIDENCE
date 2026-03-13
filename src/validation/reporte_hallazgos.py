# -*- coding: utf-8 -*-
"""
Reporte de Hallazgos — Hoja HALLAZGOS en Excel
================================================
Tarea #29 del Plan de Desarrollo (Fase 4: Validaciones)

Genera la hoja HALLAZGOS en el Workbook Excel con todos los hallazgos
detectados por el validador aritmético (#27) y reglas de viáticos (#28).

Columnas:
  - #: Número correlativo
  - Severidad: CRÍTICO / ALTO / MEDIO / BAJO / INFO
  - Tipo: ARITMÉTICO / NORMATIVO / DOCUMENTAL / CRUZADO
  - Descripción: Texto del hallazgo
  - Acción requerida: Qué debe hacer el revisor
  - Referencia normativa: Regla o directiva
  - Archivo: PDF fuente
  - Página: Número de página
  - Comprobante: Serie-número afectado
  - Confianza: Nivel de confianza de la evidencia

Diseño:
  - Banner resumen con contadores por severidad
  - Tabla con colores semáforo por severidad
  - Pie con total y distribución

Consume:
  - List[Observacion] de validador_expediente.py y reglas_viaticos.py
  - ResultadoValidacion y ResultadoReglasViaticos para estadísticas

Produce:
  - Hoja "HALLAZGOS" en Workbook openpyxl existente

Versión: 1.0.0
Fecha: 2026-03-13
"""

import os
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

# Asegurar que el directorio raíz del proyecto esté en el path
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

from config.settings import NivelObservacion, Observacion

# ==============================================================================
# CONSTANTES
# ==============================================================================

VERSION_REPORTE = "1.0.0"
"""Versión del módulo reporte_hallazgos."""

NOMBRE_HOJA = "HALLAZGOS"
"""Nombre de la hoja en el Workbook."""

# Mapeo de NivelObservacion a etiqueta de severidad para Excel
MAPA_SEVERIDAD = {
    NivelObservacion.CRITICA: "CRÍTICO",
    NivelObservacion.MAYOR: "ALTO",
    NivelObservacion.MENOR: "MEDIO",
    NivelObservacion.INFORMATIVA: "BAJO",
    NivelObservacion.INCIERTO: "INFO",
}

# Mapeo de regla_aplicada a tipo de hallazgo
TIPOS_HALLAZGO = {
    "VAL_IGV": "ARITMÉTICO",
    "VAL_TOTAL": "ARITMÉTICO",
    "VAL_SUMA_ITEMS": "ARITMÉTICO",
    "VAL_NOCHES": "ARITMÉTICO",
    "VAL_DUPLICIDAD": "CRUZADO",
    "VAL_SUMA_VS_ANEXO3": "CRUZADO",
    "VAL_CAMPOS": "DOCUMENTAL",
    "VAL_EXPEDIENTE": "DOCUMENTAL",
    "VIAT_DOC": "DOCUMENTAL",
    "VIAT_TOPE": "NORMATIVO",
    "VIAT_FECHA": "NORMATIVO",
    "VIAT_MONTO": "NORMATIVO",
    "VIAT_COBERTURA": "NORMATIVO",
    "VIAT_BOLETA": "NORMATIVO",
    "VIAT_DEVOLUCION": "NORMATIVO",
}


def _clasificar_tipo(regla: str) -> str:
    """Clasifica el tipo de hallazgo basado en la regla_aplicada."""
    if not regla:
        return "OTRO"
    for prefijo, tipo in TIPOS_HALLAZGO.items():
        if regla.startswith(prefijo):
            return tipo
    return "OTRO"


# ==============================================================================
# COLORES SEMÁFORO
# ==============================================================================


@dataclass(frozen=True)
class _Color:
    fondo: str
    texto: str


_CRITICO = _Color(fondo="FFC7CE", texto="9C0006")
_ALTO = _Color(fondo="FFC7CE", texto="9C0006")
_MEDIO = _Color(fondo="FFEB9C", texto="9C5700")
_BAJO = _Color(fondo="C6EFCE", texto="006100")
_INFO = _Color(fondo="D9D9D9", texto="404040")
_HEADER = _Color(fondo="4472C4", texto="FFFFFF")
_BANNER = _Color(fondo="D6E4F0", texto="1F3864")
_BLANCO = _Color(fondo="FFFFFF", texto="000000")

_COLORES_SEVERIDAD = {
    "CRÍTICO": _CRITICO,
    "ALTO": _ALTO,
    "MEDIO": _MEDIO,
    "BAJO": _BAJO,
    "INFO": _INFO,
}


# ==============================================================================
# DATOS PARA EXCEL
# ==============================================================================


@dataclass
class FilaHallazgo:
    """Una fila de la tabla de hallazgos."""

    numero: int
    severidad: str
    tipo: str
    descripcion: str
    accion_requerida: str
    referencia_normativa: str
    archivo: str
    pagina: int
    comprobante: str
    confianza: float

    def to_list(self) -> list:
        return [
            self.numero,
            self.severidad,
            self.tipo,
            self.descripcion,
            self.accion_requerida,
            self.referencia_normativa,
            self.archivo,
            self.pagina if self.pagina > 0 else "",
            self.comprobante,
            f"{self.confianza:.0%}" if self.confianza > 0 else "",
        ]


@dataclass
class ResumenHallazgos:
    """Resumen estadístico para el banner."""

    total: int = 0
    criticos: int = 0
    altos: int = 0
    medios: int = 0
    bajos: int = 0
    info: int = 0

    def to_dict(self) -> Dict[str, int]:
        return {
            "total": self.total,
            "criticos": self.criticos,
            "altos": self.altos,
            "medios": self.medios,
            "bajos": self.bajos,
            "info": self.info,
        }


# ==============================================================================
# CLASE PRINCIPAL — EscritorHallazgos
# ==============================================================================


class EscritorHallazgos:
    """
    Genera la hoja HALLAZGOS en un Workbook Excel.

    Recibe observaciones de validaciones y las presenta en formato
    de tabla con colores semáforo por severidad.
    """

    def __init__(self, nombre_hoja: str = NOMBRE_HOJA):
        self._nombre_hoja = nombre_hoja

    @property
    def version(self) -> str:
        return VERSION_REPORTE

    # ==========================================================================
    # MÉTODO PRINCIPAL
    # ==========================================================================

    def escribir_hoja_hallazgos(
        self,
        wb: "Workbook",
        observaciones: List[Observacion],
        sinad: str = "",
    ) -> Optional["Worksheet"]:
        """
        Escribe la hoja HALLAZGOS en el Workbook.

        Parameters
        ----------
        wb : Workbook
            Workbook openpyxl donde agregar la hoja.
        observaciones : List[Observacion]
            Hallazgos de validaciones (#27 y #28).
        sinad : str
            Identificador SINAD para el banner.

        Returns
        -------
        Worksheet or None
            La hoja creada, o None si openpyxl no está disponible.
        """
        if not OPENPYXL_DISPONIBLE:
            return None

        # Crear hoja (eliminar si ya existe)
        if self._nombre_hoja in wb.sheetnames:
            del wb[self._nombre_hoja]
        ws = wb.create_sheet(title=self._nombre_hoja)

        # Preparar datos
        filas = self._observaciones_a_filas(observaciones)
        resumen = self._calcular_resumen(observaciones)

        # Escribir secciones
        fila_actual = 1
        fila_actual = self._escribir_banner(ws, fila_actual, sinad, resumen)
        fila_actual += 1  # Línea en blanco
        fila_actual = self._escribir_tabla(ws, fila_actual, filas)
        fila_actual += 1
        self._escribir_pie(ws, fila_actual, resumen)

        # Ajustar anchos
        self._ajustar_anchos(ws)

        return ws

    # ==========================================================================
    # CONVERSIÓN DE OBSERVACIONES
    # ==========================================================================

    def _observaciones_a_filas(
        self,
        observaciones: List[Observacion],
    ) -> List[FilaHallazgo]:
        """Convierte observaciones a filas de tabla."""
        filas = []

        # Ordenar por severidad (más grave primero)
        orden = {
            NivelObservacion.CRITICA: 0,
            NivelObservacion.MAYOR: 1,
            NivelObservacion.MENOR: 2,
            NivelObservacion.INFORMATIVA: 3,
            NivelObservacion.INCIERTO: 4,
        }
        obs_ordenadas = sorted(observaciones, key=lambda o: orden.get(o.nivel, 5))

        for i, obs in enumerate(obs_ordenadas, 1):
            severidad = MAPA_SEVERIDAD.get(obs.nivel, "INFO")
            tipo = _clasificar_tipo(obs.regla_aplicada)
            referencia = obs.regla_aplicada or ""

            # Extraer archivo, página, comprobante de evidencia
            archivo = ""
            pagina = 0
            comprobante = ""
            confianza = 0.0

            ev = obs.get_evidencia_principal()
            if ev:
                archivo = ev.archivo or ""
                pagina = ev.pagina
                confianza = ev.confianza
                # Extraer serie-número del snippet o valor_detectado
                if ev.valor_detectado and any(
                    c in ev.valor_detectado for c in ["-", "F0", "B0", "E0"]
                ):
                    comprobante = ev.valor_detectado

            filas.append(
                FilaHallazgo(
                    numero=i,
                    severidad=severidad,
                    tipo=tipo,
                    descripcion=obs.descripcion[:200],
                    accion_requerida=obs.accion_requerida[:150] if obs.accion_requerida else "",
                    referencia_normativa=referencia,
                    archivo=archivo,
                    pagina=pagina,
                    comprobante=comprobante,
                    confianza=confianza,
                )
            )

        return filas

    def _calcular_resumen(self, observaciones: List[Observacion]) -> ResumenHallazgos:
        """Calcula contadores por severidad."""
        resumen = ResumenHallazgos(total=len(observaciones))
        for obs in observaciones:
            if obs.nivel == NivelObservacion.CRITICA:
                resumen.criticos += 1
            elif obs.nivel == NivelObservacion.MAYOR:
                resumen.altos += 1
            elif obs.nivel == NivelObservacion.MENOR:
                resumen.medios += 1
            elif obs.nivel == NivelObservacion.INFORMATIVA:
                resumen.bajos += 1
            else:
                resumen.info += 1
        return resumen

    # ==========================================================================
    # ESCRIBIR BANNER
    # ==========================================================================

    def _escribir_banner(
        self,
        ws: "Worksheet",
        fila: int,
        sinad: str,
        resumen: ResumenHallazgos,
    ) -> int:
        """Escribe banner resumen en la parte superior."""
        fill_banner = PatternFill(
            start_color=_BANNER.fondo, end_color=_BANNER.fondo, fill_type="solid"
        )
        font_titulo = Font(name="Calibri", size=14, bold=True, color=_BANNER.texto)
        font_dato = Font(name="Calibri", size=11, color=_BANNER.texto)

        # Título
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)
        celda = ws.cell(row=fila, column=1)
        celda.value = f"REPORTE DE HALLAZGOS — {sinad}" if sinad else "REPORTE DE HALLAZGOS"
        celda.font = font_titulo
        celda.fill = fill_banner
        celda.alignment = Alignment(horizontal="center")
        fila += 1

        # Resumen por severidad
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)
        celda = ws.cell(row=fila, column=1)
        celda.value = (
            f"Total: {resumen.total} | "
            f"Críticos: {resumen.criticos} | "
            f"Altos: {resumen.altos} | "
            f"Medios: {resumen.medios} | "
            f"Bajos: {resumen.bajos}"
        )
        celda.font = font_dato
        celda.fill = fill_banner
        celda.alignment = Alignment(horizontal="center")
        fila += 1

        return fila

    # ==========================================================================
    # ESCRIBIR TABLA
    # ==========================================================================

    def _escribir_tabla(
        self,
        ws: "Worksheet",
        fila: int,
        filas_datos: List[FilaHallazgo],
    ) -> int:
        """Escribe la tabla de hallazgos con headers y datos."""
        headers = [
            "#",
            "Severidad",
            "Tipo",
            "Descripción",
            "Acción Requerida",
            "Referencia",
            "Archivo",
            "Página",
            "Comprobante",
            "Confianza",
        ]

        # Header
        fill_header = PatternFill(
            start_color=_HEADER.fondo, end_color=_HEADER.fondo, fill_type="solid"
        )
        font_header = Font(name="Calibri", size=10, bold=True, color=_HEADER.texto)
        borde = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.font = font_header
            celda.fill = fill_header
            celda.border = borde
            celda.alignment = Alignment(horizontal="center", wrap_text=True)
        fila += 1

        # Datos
        font_dato = Font(name="Calibri", size=10)
        for fila_dato in filas_datos:
            valores = fila_dato.to_list()
            color = _COLORES_SEVERIDAD.get(fila_dato.severidad, _INFO)
            fill = PatternFill(start_color=color.fondo, end_color=color.fondo, fill_type="solid")
            font_color = Font(name="Calibri", size=10, color=color.texto)

            for col, valor in enumerate(valores, 1):
                celda = ws.cell(row=fila, column=col)
                celda.value = valor
                celda.border = borde
                if col <= 2:  # # y Severidad con color
                    celda.fill = fill
                    celda.font = font_color
                else:
                    celda.font = font_dato
                # Wrap text para descripción y acción
                if col in (4, 5):
                    celda.alignment = Alignment(wrap_text=True, vertical="top")
            fila += 1

        return fila

    # ==========================================================================
    # ESCRIBIR PIE
    # ==========================================================================

    def _escribir_pie(
        self,
        ws: "Worksheet",
        fila: int,
        resumen: ResumenHallazgos,
    ) -> int:
        """Escribe pie con totales y versión."""
        font_pie = Font(name="Calibri", size=9, italic=True, color="808080")

        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)
        celda = ws.cell(row=fila, column=1)
        celda.value = (
            f"AG-EVIDENCE Validador v{VERSION_REPORTE} | "
            f"{resumen.total} hallazgos detectados | "
            f"Distribución: {resumen.criticos}C/{resumen.altos}A/{resumen.medios}M/{resumen.bajos}B"
        )
        celda.font = font_pie
        fila += 1

        return fila

    # ==========================================================================
    # AJUSTAR ANCHOS
    # ==========================================================================

    def _ajustar_anchos(self, ws: "Worksheet"):
        """Ajusta anchos de columna para legibilidad."""
        anchos = {
            1: 5,  # #
            2: 12,  # Severidad
            3: 14,  # Tipo
            4: 60,  # Descripción
            5: 40,  # Acción
            6: 30,  # Referencia
            7: 25,  # Archivo
            8: 8,  # Página
            9: 20,  # Comprobante
            10: 12,  # Confianza
        }
        for col, ancho in anchos.items():
            ws.column_dimensions[get_column_letter(col)].width = ancho


# ==============================================================================
# FUNCIÓN DE CONVENIENCIA
# ==============================================================================


def escribir_hallazgos(
    wb: "Workbook",
    observaciones: List[Observacion],
    sinad: str = "",
    nombre_hoja: str = NOMBRE_HOJA,
) -> Optional["Worksheet"]:
    """
    Función de conveniencia para escribir hoja HALLAZGOS.

    Parameters
    ----------
    wb : Workbook
        Workbook openpyxl.
    observaciones : List[Observacion]
        Hallazgos a reportar.
    sinad : str
        SINAD para el banner.
    nombre_hoja : str
        Nombre de la hoja (default: HALLAZGOS).

    Returns
    -------
    Worksheet or None
    """
    escritor = EscritorHallazgos(nombre_hoja=nombre_hoja)
    return escritor.escribir_hoja_hallazgos(wb, observaciones, sinad)
